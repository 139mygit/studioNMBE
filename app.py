import os
import openai
from flask import Flask, request, jsonify, send_file,session,redirect
from flask_cors import CORS

from PyPDF2 import PdfReader
import pandas as pd

from azure.identity import DefaultAzureCredential
from azure.cosmos import CosmosClient, PartitionKey
from azure.storage.blob import BlobServiceClient

import logging
from datetime import datetime,timezone,timedelta,UTC
import uuid
import json
import io
import re
import fitz  # PyMuPDF
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import time
import threading

import zipfile
import lxml.etree as ET
import os
import io
import ast  # 상단에 임포트 추가
from azure.cosmos.exceptions import CosmosResourceNotFoundError, CosmosHttpResponseError
import secrets
from flask_session import Session
from werkzeug.security import generate_password_hash, check_password_hash
import urllib.parse
from io import StringIO
from asgiref.wsgi import WsgiToAsgi

import asyncio
#新追加的包 530
import requests
import pdfplumber
from openpyxl.utils import get_column_letter
from copy import copy
from difflib import SequenceMatcher
import jaconv
import regex as regcheck

# 로그 형식 정의 (시간, 로그 레벨, 메시지)
log_format = '%(asctime)sZ: [%(levelname)s] %(message)s'

# 로그 기본 설정: 시간 형식, 로그 레벨 및 출력 형식 설정
logging.basicConfig(
    level=logging.INFO,  # 로그 레벨 설정 (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format=log_format,   # 로그 형식 설정
    handlers=[logging.StreamHandler()]  # 콘솔로 출력 (파일로도 출력 가능)
)

# Managed Identity Auth
credential = DefaultAzureCredential()
token_OPENAI = credential.get_token("https://cognitiveservices.azure.com/.default")
token_COSMOS = credential.get_token("https://cosmos.azure.com/.default")

# Flask app init
app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # 安全密钥
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 会话有效期30分钟

# 🔹 Flask 세션 설정 (세션을 파일 시스템에 저장하여 지속성 유지)
app.config["SESSION_TYPE"] = "filesystem"  # 서버가 재시작되어도 세션 유지
app.config["SESSION_COOKIE_SECURE"] = False  # 개발 환경에서는 False, 운영 환경에서는 True
app.config["SESSION_COOKIE_HTTPONLY"] = True  # JavaScript에서 접근 불가능 (XSS 방지)
app.config["SESSION_COOKIE_SAMESITE"] = "None"  # CSRF 보호 (CORS 요청 가능)
app.config["SESSION_COOKIE_NAME"] = "secure_session"  # 세션 쿠키 이름


Session(app)

# CORS(app, resources={r"/api/*": {"origins": "*"}})
CORS(app, supports_credentials=True, resources={
    r"*": {
        "origins": "*"  # 실제 프론트엔드 도메인으로 변경
    }
})



# 模拟用户数据库
users = {
    "admin": {"password": "123"},
    "user": {"password": "123"}
}

#-----------------------------------------------------------------
# Azure OpenAI Setting
# openai.api_type = "azure"
# openai.api_key = os.getenv("AZURE_OPENAI_KEY")  # Get ENV API Key

# COSMOS_DB_KEY = os.getenv("COSMOS_DB_KEY")  # Cosmos DB Key
#-----------------------------------------------------------------


# AzureTokenCache 클래스 정의
class AzureTokenCache:
    def __init__(self):
        self._lock = threading.Lock() # 스레드 안전을 위한 락
        self.credential = DefaultAzureCredential()
        self.scope = "https://cognitiveservices.azure.com/.default"
        
        self.cached_token = None
        self.token_expires = 0  # 토큰 만료 시간 (Unix 타임스탬프)
        self.last_refreshed = 0 # 마지막 체크 시간 (Unix 타임스탬프)
        
        # 초기 토큰 획득
        self._refresh_token()
        
        # 백그라운드 갱신 쓰레드 시작
        self._start_refresh_thread()

    def get_token(self):
        with self._lock:
            # 토큰이 10분 이내로 만료되면 갱신
            if time.time() >= self.token_expires - 600:  # 10분 전에 갱신
                self._refresh_token()
            return self.cached_token

    def _acquire_new_token(self):
        """새 토큰 획득"""
        return self.credential.get_token(self.scope)

    def _refresh_token(self):
        """토큰 갱신"""
        new_token = self._acquire_new_token()
        with self._lock:
            self.cached_token = new_token.token
            self.token_expires = new_token.expires_on  # 실제 토큰 만료 시간 사용
            self.last_refreshed = time.time()
        print(f"🔄Updated Token (END of at:,haha, {self._format_time(self.token_expires)})")

    def _start_refresh_thread(self):
        """백그라운드에서 1분마다 토큰을 확인하고 만료되면 자동 갱신하는 쓰레드 실행"""
        thread = threading.Thread(target=self._refresh_loop, daemon=True)
        thread.start()

    def _refresh_loop(self):
        while True:
            time.sleep(30)  # 30초마다 실행하여 만료 여부 확인
            if time.time() >= self.token_expires - 600:  # 10분 전에 갱신
                self._refresh_token()

    def _format_time(self, timestamp):
        """디버깅용 시간 포맷팅"""
        local_time = time.localtime(timestamp)
        adjusted_time = time.mktime(local_time) + (8 * 3600)  # 8시간을 초 단위로 추가
        return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(adjusted_time))
# -------------------------------------------------------------------
token_cache = AzureTokenCache()
#---------

# token method
openai.api_type = "azure_ad"
openai.api_base = os.getenv("AZURE_OPENAI_ENDPOINT")  # Get Env
openai.api_version = os.getenv("AZURE_OPENAI_API_VERSION")  # API Version
deployment_id = os.getenv("AZURE_OPENAI_MODEL")
_deployment_id = os.getenv("AZURE_OPENAI_MODEL_4")

# Cosmos DB 연결 설정
COSMOS_DB_URI = os.getenv("COSMOS_DB_URI")
DATABASE_NAME = os.getenv("DATABASE_NAME")
CONTAINER_NAME = os.getenv("CONTAINER_NAME")  # debug not used

# Azure Storage 계정 URL & 컨테이너명
ACCOUNT_URL = os.getenv("ACCOUNT_URL")
STORAGE_CONTAINER_NAME = os.getenv("STORAGE_CONTAINER_NAME")


# Cosmos DB
def get_db_connection(CONTAINER):
    # Cosmos DB 클라이언트 연결
    client = CosmosClient(COSMOS_DB_URI, credential=credential)
    database = client.get_database_client(DATABASE_NAME)
    container = database.get_container_client(CONTAINER)
    print("Connected to Azure Cosmos DB SQL API")
    logging.info("Connected to Azure Cosmos DB SQL API")
    return container  # Cosmos DB의 컨테이너 객체 반환

#-----------------------------------------------------------------
LOG_RECORD_CONTAINER_NAME = "log_record"
FILE_MONITOR_ITEM = "file_monitor_item"
TENBREND_CONTAINER_NAME = 'tenbrend_history'
PROXYINFO_CONTAINER_NAME = 'proxyInfo'
#-----------------------------------------------------------------
# List proxy
@app.route('/api/proxyinfo', methods=['GET'])
def get_proxyinfos():
    # Cosmos DB 연결
    container = get_db_connection(PROXYINFO_CONTAINER_NAME)
    
    query = "SELECT * FROM c"
    users = list(container.query_items(query=query, enable_cross_partition_query=True))
    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

# Create proxy
@app.route('/api/proxyinfo', methods=['POST'])
def create_proxyuser():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Missing username or password"}), 400

    # Cosmos DB 연결
    container = get_db_connection(PROXYINFO_CONTAINER_NAME)

    # 기존 사용자 확인
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]
    existing_users = list(container.query_items(
        query=query, 
        parameters=params, 
        enable_cross_partition_query=True
    ))

    if existing_users:
        return jsonify({"error": "Username already exists"}), 409  # HTTP 409 Conflict

    user_item = {
        'id': str(uuid.uuid4()),
        'username': username,
        'password': password  # 비밀번호 해싱
    }
    container.create_item(body=user_item)
    # 응답에 code: 200 추가
    response = {
        "code": 200,
        "data": user_item
    }

    return jsonify(response), 201

# update proxy
@app.route('/api/proxyinfo', methods=['PUT'])
def update_proxyuser():
    try:
        data = request.get_json()
        new_username = data.get('username')
        new_password = data.get('password')

        # 1. 필수 필드 검증
        if not all([new_username, new_password]):
            return jsonify({"error": "Required fields: proxyuserName and Password"}), 400

        container = get_db_connection(PROXYINFO_CONTAINER_NAME)

        # 2. 기존 사용자 조회 (ID로 조회 + 크로스 파티션 쿼리)
        try:
            query = f"SELECT * FROM c"
            existing_user = list(container.query_items(
                query=query,
                enable_cross_partition_query=True
            ))[0]
        except IndexError:
            return jsonify({"error": "Find error error"}), 404

        proxy_data = dict(username=new_username, password=new_password)
        if existing_user:
            existing_user.update(proxy_data)
            container.upsert_item(existing_user)
        else:
            proxy_data.update(id=str(uuid.uuid4()))
            container.upsert_item(proxy_data)

        return jsonify({
            "username": new_username,
            "code": 200
        }), 200

    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB Error: {str(e)}")
        return jsonify({"error": "DB Error"}), 500
    except Exception as e:
        logging.error(f"server error: {str(e)}")
        return jsonify({"error": "server error"}), 500
    
USERINFO_CONTAINER_NAME = 'userInfo'
#----------------------User CRUD--------
@app.route('/api/users', methods=['GET'])
def get_users():
    # Cosmos DB 연결
    container = get_db_connection(USERINFO_CONTAINER_NAME)

    query = "SELECT * FROM c"
    users = list(container.query_items(query=query, enable_cross_partition_query=True))
    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

@app.route('/api/users', methods=['POST'])
def create_user():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Missing username or password"}), 400

    container = get_db_connection(USERINFO_CONTAINER_NAME)

    # 기존 사용자 확인
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]
    existing_users = list(container.query_items(
        query=query, 
        parameters=params, 
        enable_cross_partition_query=True
    ))

    if existing_users:
        return jsonify({"error": "Username already exists"}), 409  # HTTP 409 Conflict

    user_item = {
        'id': str(uuid.uuid4()),
        'username': username,
        'password': generate_password_hash(password)  # 비밀번호 해싱
    }
    container.create_item(body=user_item)
    # 응답에 code: 200 추가
    response = {
        "code": 200,
        "data": user_item
    }

    return jsonify(response), 201

@app.route('/api/users/<user_id>', methods=['PUT'])
def update_user(user_id):
    try:
        data = request.get_json()
        new_username = data.get('username')
        new_password = data.get('password')

        # 1. 필수 필드 검증
        if not all([new_username, new_password]):
            return jsonify({"error": "사용자명과 비밀번호 필수 입력"}), 400

        container = get_db_connection(USERINFO_CONTAINER_NAME)

        # 2. 기존 사용자 조회 (ID로 조회 + 크로스 파티션 쿼리)
        try:
            query = f"SELECT * FROM c WHERE c.id = '{user_id}'"
            existing_user = list(container.query_items(
                query=query,
                enable_cross_partition_query=True
            ))[0]
        except IndexError:
            return jsonify({"error": "사용자를 찾을 수 없음"}), 404

        # 3. 사용자명 중복 검사
        if existing_user['username'] != new_username:
            dup_query = f"SELECT * FROM c WHERE c.username = '{new_username}'"
            if list(container.query_items(dup_query, enable_cross_partition_query=True)):
                return jsonify({"error": "이미 사용중인 이름"}), 409

        # 4. 문서 업데이트
        updated_item = {
            "id": user_id,
            "username": new_username,
            "password": generate_password_hash(new_password),
            # 기존 필드 유지
            **{k: v for k, v in existing_user.items() if k not in ['username', 'password']}
        }

        # 5. 기존 문서 삭제 후 새 문서 생성 (파티션 키 변경 대응)
        container.delete_item(item=user_id, partition_key=existing_user['id'])
        container.create_item(body=updated_item)

        return jsonify({
            "id": updated_item['id'],
            "username": updated_item['username']
        }), 200

    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB 오류: {str(e)}")
        return jsonify({"error": "데이터베이스 오류"}), 500
    except Exception as e:
        logging.error(f"서버 오류: {str(e)}")
        return jsonify({"error": "내부 서버 오류"}), 500
            

@app.route('/api/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    container = get_db_connection(USERINFO_CONTAINER_NAME)
    
    container.delete_item(item=user_id, partition_key=user_id)
    return jsonify({"message": "User deleted"}), 200

#--------------------------------------
@app.before_request
def check_session():
    # 检查会话有效期
    if 'user_id' in session:
        last_activity = session.get('last_activity')
        session.modified = True  # 자동 갱신 활성화
        if last_activity and (datetime.now() - datetime.fromisoformat(last_activity)) > app.config['PERMANENT_SESSION_LIFETIME']:
            session.clear()
            return jsonify({"status": "error", "message": "Session expired"}), 401
        # 更新最后活动时间
        session['last_activity'] = datetime.now().isoformat()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip().lower()  # 대소문자 통일 및 공백 제거
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({"status": "error", "message": "ユーザー名またはパスワードが間違っています"}), 400

    container = get_db_connection(USERINFO_CONTAINER_NAME)
    
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]

    items = list(container.query_items(
        query=query,
        parameters=params,
        enable_cross_partition_query=True
    ))

    if not items:
        return jsonify({"success": "false", "message": "User not found"}), 404

    user = items[0]
    if not check_password_hash(user['password'], password):
        return jsonify({"success": "false", "message": "Invalid password"}), 401

    session.clear()
    session['user_id'] = user['id']
    session['username'] = username

    return jsonify({"success": "true", "message": "ログイン成功！"}), 200

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"status": "success", "message": "ログアウト"}), 200

@app.route('/api/protected', methods=['GET'])  # 이미 GET 요청을 처리함
def protected():
    if not session.get('session_id'):
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    return jsonify({
        "status": "success",
        "message": "Protected content",
        "secure_session": session.get('session_id')
    }), 200


# 모니터링 상태 컨테이너 설정
CHECK_SESSION_COOKIE = "session_cookie"

# Cosmos DB 상태 확인 엔드포인트 (수정 버전)
@app.route('/api/session_cookie', methods=['GET'])
def get_session_cookie():
    try:
        container = get_db_connection(CHECK_SESSION_COOKIE)
        
        # Cosmos DB에서 데이터를 쿼리하여 가져오기
        query = "SELECT * FROM c"  # SQL 쿼리 (전체 데이터를 가져옴)
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        # 각 아이템의 '_id'는 이미 문자열로 되어 있기 때문에 추가적인 변환이 필요하지 않음
        for item in items:
            item['id'] = item['id']  # Cosmos DB에서 _id는 id로 제공

        return jsonify(items), 200  # JSON 형식으로 반환
        
    except CosmosResourceNotFoundError:
        logging.error("Monitoring status document not found")
        return jsonify({"error": "Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    

# 상태 업데이트 엔드포인트 (수정 버전)
@app.route('/api/session_cookie', methods=['PUT'])
def update_session_cookie():
    try:
        container = get_db_connection(CHECK_SESSION_COOKIE)
        
        # 쿠키에서 secure_session 값 추출
        # session_value = request.cookies.get('secure_session', 'none')
        session_value = request.json.get('status', 'off')
        
        # 문서 업데이트
        status_item = {
            'id': 'session_cookie',
            'type': 'control',  # 파티션 키
            'session_value': session_value,
            "timestamp": datetime.utcnow().isoformat()
        }
        
        container.upsert_item(body=status_item)
        logging.info(f"Session value updated: {session_value}")
        return jsonify({
            'message': 'Session value updated',
            'session_value': session_value
        }), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Database operation failed"}), 500
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
#-----------API--------------------
def remove_code_blocks(text):
    # 정규 표현식을 사용하여 ```html 및 ``` 제거
    text = re.sub(r'```html', '', text)
    text = re.sub(r'```', '', text)
    return text.strip()  # 앞뒤 공백 제거

def remove_code_blocks_enhance(text):
    # 정규 표현식을 사용하여 ```html\n 및 ``` 제거
    text = re.sub(r'```html\n?', '', text)  # ```html 또는 ```html\n 제거
    text = re.sub(r'```', '', text)          # ``` 제거
    text = re.sub(r'\n\n\*\*NG\*\*\n```', '', text)  # \n\n**NG**\n``` 제거
    return text.strip()  # 앞뒤 공백 제거


@app.route('/api/dic_search_db', methods=['POST'])
def dic_search_db():
    try:
        # 요청 데이터 파싱
        data = request.json

        original = data.get('original')
        corrected = data.get('corrected')

        # Cosmos DB 컨테이너 클라이언트 가져오기
        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        # DB에서 일치하는 데이터 찾기
        query = f"SELECT * FROM c WHERE c.original = '{original}' AND c.corrected = '{corrected}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            # 일치하는 데이터의 result 출력
            results = [{"original": item["original"], "corrected": item["corrected"]} for item in items]
            # results = [item if item.get("original") else {"corrected": item["corrected"], } for item in items]
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 404

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/ask_gpt', methods=['POST'])
def ask_gpt():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        prompt = data.get("input", "")

        if not prompt:
            return jsonify({"success": False, "error": "No input provided"}), 400

        # db get map
        corrected_map = fetch_and_convert_to_dict()

        # 3. apply_corrections를 사용하여 교정 적용
        corrected = apply_corrections(prompt, corrected_map)


        prompt_result = f"""
        You are a professional Japanese text proofreading assistant. 
        Please carefully proofread the content of a Japanese report following the rules below. 
        This includes not only Japanese text but also English abbreviations (英略語), foreign terms (外来語),
        and specialized terminology (専門用語). Ensure that all language elements are reviewed according to the guidelines and corrected where necessary.:

        **Report Content to Proofread:**
        {corrected}

        **Proofreading Requirements:**
        1. **Check for typos and missing characters (誤字脱字がないこと):**
        - Ensure there are no **spelling errors** or **missing characters** in the report. 
        - あなたの役割は、日本語の誤字・脱字・表記ミスを修正し、不完全な単語や文章に適切な語を補完することです。  
        以下のルールに従い、入力されたテキストを校正してください。
            **Common Mistakes Examples (誤字・脱字の例)**:
            Example:
            - `リテール投家` → `リテール投資家` (誤字: 家 → 資)
            - `長国債` → `長期国債` (脱字: 期を追加)
            - `識された` → `意識された` (表記の統一)
            - `金緩和期待` → `金融緩和期待` (誤字: 金 → 金融)
            - `見方が動し` → `見方が変動し` (誤字: 動し → 動き)
            - `視する` → `重視する`  
            - `経成長` → `経済成長`  
            - `送配電備` → `送配電設備`  
            - `業見通し` → `業績見通し`  
            - `常増益` → `経常増益`  
            - `財政策` → `財政政策`  
            - `方` → `方針`  
            - `手Eコマース` → `大手Eコマース`  
            - `響しました` → `影響しました`  
            - `施され` → `実施された`  
            - `企業の合併・収` → `企業の合併・回収`  
            - `本とします` → `基本とします`  
            - `務状況` → `財務状況`
            - `内投資信託` → `国内投資信託`  
            - `持しました` → `維持しました`  
            - `マイナス因` → `マイナス要因`  
            - `反される` → `反映される`  
            - `替ヘッジ` → `為替ヘッジ`  
            - `比は` → `比率は`
            - `規緩和` → `規制緩和`
            - `景済指標` → `経済指標`
            - `剤` → `経済`
            - `昇するなどまちまちでした。` → `異なる動きとなりました。` (Ensure that the original text is not directly modified but follows this guideline.)
            - `積極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `消極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `（割安に）放置` → `割安感のある`
            - `限定的` → `他の適切な表現に修正 （修正理由: 効果や影響がプラスかマイナスか不明瞭なため）`
            - `利益確定の売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `利食い売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `必ず～` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～になる` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～である` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
  
            **Disambiguation Rule**:
            - 「沈静」＝自然に落ち着く (natural calming down; happens over time)
            - 「鎮静」＝人為的におさめる (intentional suppression; medically or artificially done)

            **Correction Policy**:
            1. Detect whether the context implies a natural or artificial calming.
            2. If the usage does not match the context, correct it using the appropriate word.
            3. Highlight the correction using the format below:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`
            4. Do **not** modify the original sentence structure or paragraph formatting.
            5. Only apply the correction when the term is clearly misused.
            6. If the current usage is correct, do not change or annotate it.

            **Example**:
            - Input: 市場は徐々に鎮静していった。
            - Output: 市場は徐々に <span style="color:red;">沈静</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">鎮静</s> → 沈静</span>) していった。


        - 表現の使用制限:
            Expression Usage Restrictions:
            Restricted Expressions:

            - 魅力的な
            - 投資妙味
            - 割高感
            - 割安感

            Usage Conditions:
            The above expressions can be used if evidence is provided.

            However, these expressions should not be used in contexts where the word "fund" (ファンド) or any related reference is mentioned. In any sentence or context where "fund" or "ファンド" appears, these expressions should be avoided.

            使用例:
            魅力的な: 根拠に基づいて使用することは可能ですが、ファンドについては使用しないようにしてください。
            投資妙味: 投資妙味があることを示す場合でも、ファンドに対する言及は避け、他の投資対象に適用するようにしてください。
            割高感: 割高感について述べる場合、ファンド以外の投資対象に対して適用してください。
            割安感: 割安感について言及する場合も、ファンドに対して使用することは不可です。

            ✅ 出力フォーマット:
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)
            ✅ Exsample1:
            Input:ファンドは魅力的な投資先として紹介された。

            Output:
            ファンドは
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対する使用は不可</span>)投資先として紹介された。
            ✅ Exsample2:
            Input:
            このファンド銘柄には投資妙味がある。
            
            Output:
            この銘柄には
            <span style="color:red;">投資妙味</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)がある。


        - 数字やパーセント（％）を含む文章の誤りをチェックする
            文脈を理解し、数値・割合の前後の語句が適切か確認すること。
            例:
            月末時点（20日判定）での為替ヘッジのターゲット比は48％です。
            → 正しい表記: 月末時点（20日判定）での為替ヘッジのターゲット比率は48％です。
            市場の成長率は10%の見込みです。 ✅ (問題なし)
            インフレ率は2上昇しました。 ❌ (誤り: 2%上昇しました。 に修正)
            販売シェアは15の拡大が予想されます。 ❌ (誤り: 販売シェアは15%の拡大が予想されます。)
        - 校正ルール
            1. **「行って来い」の適切な置き換え**(Ensure that the original text is not directly modified but follows this guideline.)
            - 文章全体を分析し、「行って来い」が何を指しているのかを判断してください。
            - **価格・指数・レートなどが上昇した意味の場合** → 「行って来い」を「上昇した」に変換
            - **価格・指数・レートなどが下落した意味の場合** → 「行って来い」を「下落した」に変換

            2. **文脈を考慮した校正**
            - 修正の際、周辺の文脈を理解し、自然な形に調整してください。
            
            3. **「横ばい」の適切な置き換え**
            - 文章全体を分析し、「横ばい」の前後の文脈を考慮してください。
            - **期中の変動幅が小さい場合** → 「横ばい」を維持
            - **変動幅が大きく、結果的に同程度となった場合** → 「ほぼ変わらず」または「同程度となる」に変更
            - 周囲の文章に合わせて、より適切な表現を使用してください。

            4. **「出遅れ感」の適切な修正**
            - 対応方針:

            「出遅れ感」 は主観的な相場観が含まれるため、必ず「…と考えます。」に修正してください。

            修正方法: 文脈に応じて自然に「〜と考えます」形に修正します。

            修正後の表現: 文の流れに合わせて自然に表現を調整し、読みやすさを考慮します。

            修正理由: 相場観が含まれているため、主観的な表現を客観的な表現に変えることで文章の信頼性を向上させます。

            出力フォーマット:
            <span style="color:red;">出遅れ感</span>
            (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 「〜と考えます」に修正</span>)

            Exsample:
                Input: この銘柄には出遅れ感がある
                Output: この銘柄には
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある

                Input: 一部のセクターには出遅れ感があると感じられる
                Output: 一部のセクターには
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある


            5. **「上昇要因」・「下落要因」の適切な説明追加**
            - **文脈を分析し、具体的な要因を追加してください**。
            - **「上昇要因」がある場合** → 上昇の理由（例: 企業決算の改善、政策の発表、需給バランスの変化など）を補足。
            - **「下落要因」がある場合** → 下落の理由（例: 景気後退懸念、金融引き締め、地政学リスクなど）を補足。
            - **修正後も文章の流れがスムーズになるように調整してください。

            6. **「予想」「心理」の適切な修正**
            - **「予想」 がある場合:**  
                - **誰の予想か明確でない場合** → 「市場予想」に修正  
            - **「心理」 がある場合:**  
                - **主語が曖昧な場合** → 「市場心理」に修正

    
        2. **Follow the "Fund Manager Comment Terminology Guide" (ファンドマネージャコメント用語集に沿った記載となっていること):**
        - **Consistent Terminology (表記の統一):**
            - Ensure the **writing format** of financial terms is **consistent throughout the report**.
            Example:
            - `相対に低かった` → `相対的に低かった` (文法修正)
            - `東証33業種分では` → `東証33業種分類では` (表記の統一)
        - **Common Mistakes and Corrections (誤記と修正例)**:
            Example:
            - `政府支の` → `政府支出の遅れ、1回はのを出、2回はのを削除` (誤字: の → 出)
            - `投資比率を維する` → `投資比率を維持する` (一致性不足 動産 → 不動産)
            - `（配当こみ）` → `（配当込み）` (表記の統一: こみ → 込み)
            - `いっぽうで` → `一方で` (表記の統一: いっぽう → 一方)
            - `ウクライナとロシアをめぐる` → `ウクライナとロシアを巡る` (表記の統一: めぐる → 巡る)
            - `東証33業種でみると` → `東証33業種で見ると` (文法修正: みる → 見る)
            - `ひき続き` → `引き続き` (表記の統一: ひき → 引き)
            - `電気機器、銀行業、保険業等` → `電気機器、銀行業、保険業など` (表記の統一: 等 → など)
            - `じき` → `次期`
            - `底いれ後` → `底入れ後`
            - `なかから` → `中から`
            - `おもな` → `主要な`
            - `はやく` → `早急に`
            - `かいつけした` → `買い付けした`
            - `など` → `等`
            - `のぞく` → `除く`
            - `くみいれ` → `組み入れ`
            - `じょうき` → `上記`
            - `とうファンド` → `当ファンド`

        - **Prohibited Words and Phrases (禁止（NG）ワード及び文章の注意事項):**
            - Check if any prohibited words or phrases are used in the report and correct them as per the guidelines.
        - **Replaceable and Recommended Terms/Expressions (置き換えが必要な用語/表現、置き換えを推奨する用語/表現):**
            - If you find terms or expressions that need to be replaced, revise them according to the provided rules.
            - ハト派／タカ派の表記（金融政策に関する）:
                -Exsample:
                - 金融緩和重視  → 金融引き締め重視
                - 金融緩和に前向き  → 金融引き締めに積極的
            - Exsample:
             - 織り込む  → 反映され
             - 相場  → 市場/価格
             - 連れ高  → 影響を受けて上昇
             - 伝播  → 広がる
             - トレンド  → 傾向
             - レンジ  → 範囲
        - **〇％を上回る（下回る）マイナスの表記:**
                - 〇％を上回る  → 〇％を超える
                - 〇％を下回る  → 縮小
                - 〇％を上回る  → 下回るマイナス幅
                - 〇％を下回る  → 縮小


        - **Use of Hiragana (ひらがなを表記するもの):**
            - Ensure the report follows the rules for hiragana notation, replacing content that does not conform to commonly used kanji.
        - **Kana Notation for Non-Standard Kanji (一部かな書き等で表記するもの):**
            - Ensure non-standard kanji are replaced with kana as the standard writing format.
        - **Correct Usage of Okurigana (一般的な送り仮名など):**
            - Ensure the correct usage of okurigana is applied.
        - **English Abbreviations, Loanwords, and Technical Terms (英略語、外来語、専門用語など):**
            - Check if English abbreviations, loanwords, and technical terms are expressed correctly.
        - **Identify and mark any 常用外漢字 (Hyōgai kanji):**
        - **Identify and mark any **常用外漢字 (Hyōgai kanji)** in the following text**
        - **常用外漢字** refers to characters **not included** in the [常用漢字表 (Jōyō Kanji List)](https://ja.wikipedia.org/wiki/常用漢字), which is Japan’s official list of commonly used kanji.
        - Refer to the [Wikipedia list of Hyōgai kanji](https://ja.wikipedia.org/wiki/常用漢字) to determine if a character falls into this category.
        - **For any detected 常用外漢字**, apply the following formatting:
        - **Highlight the incorrect character in red** (`<span style="color:red;">`).
        - **Strike through the incorrect character and provide the reason in yellow highlight.**

        ---

        ### **💡 Output Format (出力フォーマット)**
        - **Incorrect characters should be displayed in red (`<span style="color:red;">`)**.
        - **Corrected text should be marked with a strikethrough (`<s>`) and highlighted in yellow (`background:yellow`)** to show the correction.
        - Use the following structure:
            ```html
            <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **Example Correction**:
            ```plaintext
            鴉 → <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **For multiple Hyōgai kanji**, apply the same structure to each character.

        ---

        ### **✅ Example Input:**
        ```plaintext
        彼は鴉が空を飛ぶのを見た。

        ### **✅ Example output:**
        ```plaintext
        彼は <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>) が空を飛ぶのを見た。

        - **Foreign Exchange Market Trend Analysis**
            In the foreign exchange market (`為替市場`), determine whether `"円だか"` should be revised to `"円高"` (Yen Appreciation) or `"円安"` (Yen Depreciation) based on the **context**.

            #### **** Criteria for Yen Appreciation (円高)**
            - **Yen appreciation (`円高`) occurs when the value of the yen increases relative to other currencies.**  
            The following situations indicate yen appreciation:
            1. **"多くの通貨が対円で下落した"** (Many currencies declined against the yen) → Change `円だか` to **円高**.
            2. **"ドル円が下落した"** (USD/JPY exchange rate declined) → Change `円だか` to **円高**.
            3. **"対米ドルで円の価値が上昇した"** (The yen appreciated against the US dollar) → Change `円だか` to **円高**.

            #### **** Criteria for Yen Depreciation (円安)**
            - **Yen depreciation (`円安`) occurs when the value of the yen declines relative to other currencies.**  
            The following situations indicate yen depreciation:
            1. **"多くの通貨が対円で上昇した"** (Many currencies rose against the yen) → Change `円だか` to **円安**.
            2. **"ドル円が上昇した"** (USD/JPY exchange rate increased) → Change `円だか` to **円安**.
            3. **"対米ドルで円の価値が下落した"** (The yen depreciated against the US dollar) → Change `円だか` to **円安**.


        3. **Replaceable and Recommended Terms/Expressions (推奨される表現の修正)**
        - Use the correct **kanji, hiragana, and katakana** combinations based on standard Japanese financial terms.
            Example:
            - `が好された輸送用機器など` → `が好感された輸送用機器など` (修正理由: 適切な表現)

        - **Task**: Header Date Format Validation & Correction  
        - **Target Area**: Date notation in parentheses following "今後運用方針 (Future Policy Decision Basis)"  
        ---
        ### Validation Requirements  
        1. **Full Format Compliance Check**:  
        - Must follow "YYYY年MM月DD日現在" (Year-Month-Day as of)  
        - **Year**: 4-digit number (e.g., 2024)  
        - **Month**: 2-digit (01-12, e.g., April → 4)  
        - **Day**: 2-digit (01-31, e.g., 5th → 5)  
        - **Suffix**: Must end with "現在" (as of)  

        2. **Common Error Pattern Detection**:  
        ❌ "1月0日" → Missing month leading zero + invalid day 0  
        ❌ "2024年4月1日" → 2024年4月1日
        ❌ "2024年12月" → Missing day value  
        ❌ "2024-04-05現在" → Incorrect separator usage (hyphen/slash)  
        ---
        ### Correction Protocol  
        1. **Leading Zero Enforcement**  
        - Add leading zeros to single-digit months/days (4月 → 4月, 5日 → 5日)  

        2. **Day 0 Handling**  
        - Replace day 0 with YYYYMMDD Date Format  
        - Example: 2024年4月0日 → 2024年4月00日

        3. **Separator Standardization**  
        - Convert hyphens/slashes to CJK characters:  
            `2024/04/05` → `2024年4月5日`  

        ---
        ### Output Format Specification  
        ```html
        <Correction Example>
        <span style="color:red;">（2024年4月0日現在）</span> 
        → 
        <span style="color:green;">（2024年04月00日現在）</span>
        修正理由:
        ①日付0をYYYYMMDD日付フォーマットに置換
        ---

        3. **Consistency with Report Data Section (レポートのデータ部との整合性確認):**
        - Ensure the textual description in the report is completely consistent with the data section, without any logical or content-related discrepancies.

        4. **Eliminate language fluency(単語間の不要なスペース削除):**
        - Ensure that there are no extra spaces.
            -Example:
            input:景気浮揚が意 識されたことで
            output:景気浮揚が意識されたことで
        
        5.  **Layout and Formatting Rules (レイアウトに関する統一):**
            - **文頭の「○」印と一文字目の間隔を統一:**

                - When a sentence starts with the symbol ○, make sure there is no space (half-width or full-width) between it and the first character. That is, use ○文字 instead of ○ 文字 or ○　文字.
                    - Any whitespace (half-width or full-width spaces) after ○ must be removed.
                    - This spacing rule must be applied consistently throughout the document.

            半角括弧を全角括弧に統一:
                - Convert all half-width parentheses () to full-width parentheses （）.
                - Example: (注) → （注）
                - Example input: 
                    ○ 世界の高配当株式指数(注)は月間では上昇しました。
                - Exsample output: 
                    <span style="color:red;">○世界</span> (<span>修正理由: 文頭の「○」印と一文字目の間隔を統一 <s style="background:yellow;color:red">○ 世界</s> → ○世界</span>)
                    <span style="color:red;">（注）</span> (<span>修正理由: 半角括弧を全角括弧に統一 <s style="background:yellow;color:red">(注)</s> → （注）</span>)


            - **文章の間隔の統一:**
                - If a sentence begins with "○", ensure that the spacing within the frame remains consistent.
            - **上位10銘柄 コメント欄について、枠内に適切に収まっているかチェック:**
                - If the stock commentary contains a large amount of text, confirm whether it fits within the designated frame. 
                - If the ranking changes in the following month, adjust the frame accordingly.
                - **Check point**
                    1. **文字数制限内に収まっているか？**
                    - 1枠あたりの最大文字数を超えていないか？
                    - 適切な行数で収まっているか？

                    2. **次月の順位変動に伴う枠調整の必要性**
                    - 順位が変更されると枠調整が必要なため、調整が必要な箇所を特定

                    3. **枠内に収まらない場合の修正提案**
                    - 必要に応じて、短縮表現や不要な情報の削除を提案
                    - 重要な情報を損なわずに適切にリライト

                    output Format:
                    - **コメントの枠超過チェック**
                    - (枠超過しているか: はい / いいえ)
                    - (超過している場合、オーバーした文字数)

                    - **順位変動による枠調整の必要性**
                    - (調整が必要なコメントリスト)

                    - **修正提案**
                    - (枠内に収めるための修正後のコメント)

            **Standardized Notation (表記の統一):**
            - **基準価額の騰落率:**
            When there are three decimal places, round off using the round-half-up method to the second decimal place. If there are only two decimal places, keep the value unchanged.
                Make modifications directly in this article and explain the reasons for the modifications.

                exsample:
                0.546％（×） → 0.55％（○）
                修正理由: 小数点以下の桁数の丸め（0.546％ → 0.55％）
                If the value is 0.00％, replace it with "前月末から変わらず" or "前月末と同程度" instead of stating "騰落率は変わらず".
                修正理由: 「騰落率は変わらず」という表記はNG。代わりに「基準価額（分配金再投資）は前月末から変わらず」や「前月末と同程度」と記載します。

                exsample:
                0.00％となり（×） → 前月末から変わらず（○）

                騰落率は変わらず（×） → 基準価額（分配金再投資）は前月末から変わらず（○）

                When comparing the performance of the fund with the benchmark (or reference index), the comparison must be made using rounded numbers.

                修正理由: 比較は丸めた数字で行うこと。
                If the fund and benchmark (or reference index) have the same rate of return, use the phrase "騰落率は同程度となりました" instead of saying "騰落率は同じでした".
                修正理由: 同じという表現は避け、代わりに「同程度」と記載すること。

                exsample:
                「騰落率は同じでした」（×） → 「騰落率は同程度となりました」（○）
                If the fund's rate of return is greater than the benchmark's, use the phrase "上回りました" to indicate the fund outperformed the benchmark.
                修正理由: 上回った場合、「上回りました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-2.221％（ベンチマーク）の場合、値の差は0.79％となるため、「上回りました」と記載します。

                If the fund's rate of return is lower than the benchmark's, use the phrase "下降しました" to indicate the fund underperformed the benchmark.
                修正理由: 下降した場合、「下降しました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-0.221％（ベンチマーク）の場合、基金のパフォーマンスは「下降しました」と記載します。

            - **「今後の運用方針」作成日付のルール:**
                - 前月末（営業日）現在で作成。
                - 翌月初の日付になる場合は、作成した日付を入れる。
                - クライアント・サービス部へ送信する以降の日付（先日付）は入れない。
                - 「参考月」より後であり、「チェック期間」より前の日付のみ使用可。
                    - Example:
                    - OK: 参考月＝2024年2月 → 作成日が2024年2月28日（営業日） or 3月1日（翌月初）
                    - NG: 参考月＝2024年2月 → 作成日が3月5日（クライアント・サービス部送信後の先日付）

            - **％（パーセント）、カタカナ:**
                - **半角カタカナ → 全角カタカナ**（例:「ｶﾀｶﾅ」→「カタカナ」）
                - **半角記号 → 全角記号**（例:「%」→「％」、「@」→「＠」）
                    Example:
                        input: ﾍﾞﾝﾁﾏｰｸ (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して 
                        output: ベンチマーク (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して
                    Example:
                        input: ｶﾀｶﾅ 
                        output: カタカナ
                    Example:
                        input: %
                        output: ％ 
                    Example:
                        input: @
                        output: ＠ 

            - **数字、アルファベット、「＋」・「－」:**
                - **全角数字・アルファベット → 半角数字・アルファベット**（例:「１２３」→「123」、「ＡＢＣ」→「ABC」）
                - **全角「＋」「－」 → 半角「+」「-」**（例:「＋－」→「+-」
                    Example:
                        input: １２３ ＡＢＣ ｱｲｳ ＋－
                        output: 123 ABC アイウ +-

            - **スペースは変更なし**  

            - **「※」の使用:**
                - 「※」は可能であれば **上付き文字（superscript）※** に変換してください。
                - 出力形式の例:
                - 「重要事項※」 → 「重要事項<sup>※</sup>」

            - **（カッコ書き）:**
                - Parenthetical notes should only be included in their first occurrence in a comment.
                    For the following Japanese text, check if parentheses ("（ ）") are used appropriately.
                    If a parenthetical note appears more than once, remove the parentheses for subsequent occurrences.
                    The first occurrence should retain the parentheses, but any further appearances should have the parentheses removed.
                    Modification reason: Parentheses are redundant after the first mention, so the text is cleaned up for consistency and readability.

                **Check point**
                    1. **カッコ書きは、コメントの初出のみに記載されているか？**
                    - 同じカッコ書きが2回以上登場していないか？
                    - 初出ページ以降のコメントにカッコ書きが重複して記載されていないか？

                    2. **ディスクロのページ番号順に従ってルールを適用**
                    - シートの順番ではなく、実際のページ番号を基準にする。

                    3. **例外処理**
                    - 「一部例外ファンドあり」とあるため、例外的にカッコ書きが複数回登場するケースを考慮する。
                    - 例外として認められるケースを判断し、適切に指摘。

                    output Format:
                    - **カッコ書きの初出リスト**（どのページに最初に登場したか）
                    - **重複チェック結果**（どのページで二重記載されているか）
                    - **修正提案**（どのページのカッコ書きを削除すべきか）
                    - **例外ファンドが適用される場合、補足情報**

            - **会計期間の表記:**
                - The use of "～" is prohibited; always use "-".
                - Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 6～8月期（×） → 6-8月期（○）

                - 暦年を採用している国の年度表記:
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                カッコ書きで暦年の期間を明記する。
                - Example:
                    ブラジルの2021年度予算（×） → ブラジルの2021年度（2021年1月-12月）予算（○）

                - 決算期間は「●-●月期」に統一し、日付は省略する。
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 第1四半期（5月21日～8月20日）（×） → 5-8月期（○）
                    イレギュラーなケースも含め、原則「●-●月期」と表記。

            - **「TOPIX」または「東証株価指数」が含まれている場合、以下のルールを適用:**
                文中で使用する場合: 「TOPIX（東証株価指数）」と表記することを指示。
                「文中では『TOPIX（東証株価指数）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                ベンチマーク（BM）や参考指数として使用する場合: 「東証株価指数（TOPIX）（配当込み）」と表記することを指示。
                「BMや参考指数で使用する場合は、『東証株価指数（TOPIX）（配当込み）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」

            - **年をまたぐディスクロコメントの年度表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - When specifying the fiscal year in disclosure comments that span multiple years, always use "yyyy年度".
                - Similarly, for disclosures based on the January-March period, specify the corresponding year.
                - Example:
                    - For a disclosure with a December-end reference, released in January:
                    - 今年度（×） → 2021年度（○）
                    - 来年度（×） → 2022年度（○）
            - **Benchmark, Index, and Reference Index Name Formatting:(ベンチマーク・インデックス・参考指数の名称の表記)
                - Ensure Consistency in Index Terminology:
                    Read the context and identify terms related to "index" (指数) within the text. Ensure that these terms are unified and consistently referred to using the correct and standardized terminology.
                    It is important to carefully analyze each mention of "index" to make sure the terminology is consistent throughout the text.
                    Do not modify the original text directly. Instead, provide comments that explain the reasoning behind the proposed changes, especially when identifying inconsistencies or clarifications needed.
                Example Formatting Guidelines:

                    Incorrect format (×): "ISM非製造業景況"
                    Correct format (○): "ISM非製造業景況指数"
                    Incorrect format (×): "MSCIインド"
                    Correct format (○): "MSCIインド・インデックス"
                    If multiple terms are used to refer to the same index, they should be unified under the correct term. For example, if "MSCIインド指数" and "MSCIインド・インデックス" are used in different places, they should be unified as "MSCIインド・インデックス" in the final report to maintain consistency.
                Handling Multiple Terms Referring to the Same Index:

                    If it can be clearly determined that different terms refer to the same index (e.g., "MSCIインド指数" and "MSCIインド・インデックス"), do not modify them but mark them accordingly. These terms should be noted as referring to the same index.
                    Example:
                    Original: "MSCIインド指数" and "MSCIインド・インデックス"
                    Comment: These are different ways of referring to the same index, so no change is needed.
                Handling Uncertainty in Index Terminology:
                    
                    If there is uncertainty about whether multiple terms refer to the same index (e.g., it is unclear whether "ISM非製造業景況" and "ISM非製造業景況指数" refer to the same index), mark them without modification. Additionally, note that these terms might refer to the same index, but the exact nature of the index should be verified.
                    Example:
                    Original: "ISM非製造業景況" and "ISM非製造業景況指数"
                    Comment: These terms are potentially referring to the same index but require further clarification. Therefore, no changes are made in this case.
                Key Notes:

                Always ensure that consistency is maintained across the report. Even if different names are used for the same index, it is essential to mark them properly and explain that they are different terms for the same entity.
                Consistency applies not only to the formatting of the terms but also to how the terms are presented across the entire document. All references to a given index must follow the same format from the first mention to the last.


            - **上昇 or 下落に関する要因を明記:**
                文章内に 「上昇」 または 「下落」 という単語が含まれている場合、その要因を特定し、明記してください。
                output:
                「●●は上昇（または下落）しました。(理由: ○○)」
            - **指定用語の表記ルールを提示:**
                独Ifo企業景況感指数 または 独Ifo景況感指数 が含まれている場合、以下のメッセージを表示する。
                独Ifo企業景況感指数 または独Ifo景況感指数 の表記について、月報内での統一ルールを確認してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                
                「独ZEW景気期待指数」または「独ZEW景況感指数」 が含まれている場合、以下のメッセージを表示する。
                「『独ZEW景気期待指数』または『独ZEW景況感指数』の表記を使用してください。ZEW単独使用の場合は括弧書き付き、または『欧州経済研究センター』のみとし、『ZEW』単独使用を避けてください。(Ensure that the original text is not directly modified but follows this guideline.)」
            - **特定の金融用語に対する表記ルールを確認・適用:**

                「リフレーション」 が含まれている場合、以下のメッセージを表示。
                「リフレーションとは、デフレーションから抜けて、まだインフレーションにはなっていない状況を指します。」

                「リスクプレミアム」 が含まれている場合、以下のメッセージを表示。
                「リスクプレミアムとは、あるリスク資産の期待収益率が、同期間の無リスク資産（国債など）の収益率を上回る幅を指します。」

                「モメンタム」 が含まれている場合、以下のメッセージを表示。
                「『モメンタム』の使用は避け、相場の『勢い』や『方向性』などの言葉に置き換えてください。」

                「ベージュブック」 が含まれている場合、以下のメッセージを表示。
                「『ベージュブック』は『FRB（米連邦準備制度理事会）が発表したベージュブック（地区連銀経済報告）』と明記してください。スペースが限られる場合は、『ベージュブック（米地区連銀経済報告）』としてください。」

                「フリーキャッシュフロー」 が含まれている場合、以下のメッセージを表示。
                「フリーキャッシュフローとは、税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたものです。」
                
                「システミック・リスク」 が含まれている場合、以下のメッセージを表示。
                「システミック・リスクとは、個別の金融機関の支払不能や特定の市場・決済システム等の機能不全が、他の金融機関、市場、または金融システム全体に波及するリスクを指します。」

                「クレジット（信用）市場」 が含まれている場合、以下のメッセージを表示。
                「クレジット（信用）市場とは、信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称であり、企業の信用リスクを取引する市場です。」
            - **特定の金融用語に対し、欄外に注記を加える指示を表示:**
                「格付別」 が含まれている場合:
                    格付別 -> 格付別
                「格付機関」 が含まれている場合:
                    格付機関 -> 格付機関
                「組入比率」 が含まれている場合:
                    組入比率 -> 組入比率
                「引締策」 が含まれている場合:
                    引締策 -> 引締策
                「国債買入れオペ」 が含まれている場合:
                    国債買入れオペ -> 国債買入オペ

                「投資適格債」 が含まれている場合:
                「※欄外に注記: 投資適格債とは、格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債を指します。」

                「デュレーション」 が含まれている場合:
                「※欄外に注記: デュレーションとは、金利が一定の割合で変動した場合、債券の価格がどの程度変化するかを示す指標です。この値が大きいほど、金利変動に対する債券価格の変動率が大きくなります。」

                「デフォルト債」 が含まれている場合:
                「※欄外に注記: デフォルトとは、一般的に債券の利払いおよび元本返済の不履行、または遅延などを指し、このような状態にある債券を『デフォルト債』といいます。」

                「ディストレス債券」 が含まれている場合:
                「※欄外に注記: ディストレス債券とは、信用事由などにより価格が著しく下落した債券を指します。」
                
                「イールドカーブ」 が含まれている場合:
                「※欄外に注記: イールドカーブ（利回り曲線）とは、横軸に残存年数、縦軸に利回りをとった座標に、債券利回りを点描して結んだ曲線のことを指します。」

            - **組入上位10銘柄」について記述がある場合、以下のルールを適用:**
                「組入上位10銘柄を超える保有銘柄（個別銘柄の特定が可能な子会社名等を含む）は原則として開示禁止である」ことを明示。
                ただし、社内規程に基づき開示が認められているファンドは例外とすることを伝える。

            - **年度表記:**
                - Use four-digit notation for years.(Ensure that the original text is not directly modified but follows this guideline.)
                - Example: 22年（×） → 2022年（○）

            - **前年比 or 前年同月（同期）比の統一:**
                - 「前年同月（同期）比」に統一。
                - 通年の比較には「前年比」の使用可。
                - Ensure that the original text is not directly modified but follows this guideline.
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example:
                        前年比+3.0%（×） → 前年同月比+3.0%（○）
                        2023年のGDPは前年比+3.0％（○）

            - **年をまたいだ経済指標の記載:**
                - コメント内の初出のみに記載する。(In the case where there is a description of the economic indicator over the year, it is described only in the first comment.)
                    - Example:
                        - 2023年12月のCPIは～（○）
                        - 一方2024年1月のユーロ圏PMIは～ （○）
                        - 10-12月期のGDPは～（○）
            - **経済指標について:**
                    -"加速" の対象を明確に記載すること。文脈を考慮し、"加速" の対象を適切に補う。
                        文脈に応じて、"何が加速したのか" を判断し、適切な単語に修正してください。
                        修正ルール:

                        - 前月から加速しました（×） → 何が加速したのかを明記（○）
                        Exsample: 「前月から上昇が加速しました」 → 「景気回復の加速（○）」
                        - 前月から上昇が加速しました（×） → 具体的な経済活動を明記（○）
                        Exsample: 「景気加速（○）」「消費の回復が加速（○）」「投資の拡大が加速（○）」
                        - 経済（×） → 景気（○）（"経済" ではなく "景気" を用いる）
                        Exsample: 
                        前月から加速しました。（×）-> 企業の設備投資が加速しました。（○）
                        経済加速が見られます。（×）-> 景気加速が見られます。（○）
                        消費が前月から加速しました。（×）-> 個人消費の拡大が加速しました。（○）
                        インフレが前月から加速しました。（×）-> 物価上昇のスピードが加速しました。（○）


                    - 日付・国名の明記:
                        いつのものか特定できる場合、○月のものかを明記する。例: 「10月の製造業PMI（購買担当者景気指数）は～」
                        必要に応じて国名も記載する。（文脈に応じて記載していれば、位置は問わない）
                    - 日付・国名の変換ルール:

                        **「下旬」「上旬」**と記載されている場合、文脈から適切な日付に変更する。
                        **「ユーロ圏」**と記載されている場合、文脈に応じて適切な国名に置き換える。
                        Exsample:

                        修正前: 「下旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月下旬のドイツ総合PMI（購買担当者景気指数）が…」

                        修正前: 「上旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月上旬のフランス総合PMI（購買担当者景気指数）が…」

            - **業界指数の表記:**
                - 必ず対象となる「月」を明記する。
                - 月がない場合、最近3ヶ月以内か、3ヶ月以上前のものかを確認する。
                - 必要に応じて国名も記載する。（文脈に応じて位置は自由）
                    - Example:
                        - 製造業PMI（購買担当者景気指数）（×） → 10月の製造業PMI（○）
                        - 直近3ヶ月以内の指数は明示的に「○月」と記載する。（例:12月のCPI）
                        - 3ヶ月以上前の指数は、比較の文脈を明確にする。（例:2023年10月のGDP成長率）
                        - ユーロ圏の10月PMIは～（○）
            - **カタカナ表記の統一:**
                Katakana representation of foreign words should be unified within the document.
                    Ensure that the Katakana form is consistent throughout the text, and choose one version for the entire document.
                    Modification reason: To maintain consistency in the usage of Katakana for foreign words.
                    Example of text modifications:

                    サステナブル (×) → サスティナブル (○)

                    エンターテイメント (×) → エンターテインメント (○)
            
            - **レンジの表記について表記:**
                - Always append "%" when indicating a range.(Ensure that the original text is not directly modified but follows this guideline.)
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: -1～0.5%（×） → -1%～0.5%（○）
            - **償還に関する記載:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 最終リリースの1ヵ月程前より、償還に関する内容を入れること。
                - 例）当ファンドは、●●月●●日に信託の終了日（償還日or繰上償還日）を迎える予定です。
                - ※（）内は、定時償還の場合には償還日、繰上償還の場合には繰上償還日とする。
            - **個別企業名の表記:**
                - 投資環境等においては、個別企業の名称を使わない表現を心掛ける。
                - 例:スイス金融大手クレディ・スイス（×） → スイスの大手金融グループ（○）
            - **プラスに寄与/影響の表記:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 「プラスに寄与」または「プラスに影響」どちらも可。
                - または、「～プラス要因となる」と表記。
            - **マイナスに影響の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「マイナスに寄与」（×）→「マイナスに影響」（○）
                - マイナスの際は「寄与」は使用しない。
                - または、「～マイナス要因となる」と表記。
            - **利回りの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 利回りは「上昇（価格は下落）」または「低下（価格は上昇）」と表記。
            - **低下と下落の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 債券利回りは「低下（○）」と表記し、「下落（×）」は使用しない。
                - 価格は「下落（○）」と表記し、「低下（×）」は使用しない。
                - 金利の「低下（〇）」と表記し、「下落（×）」は使用しない。

            - **資金流出入の表記:**
                - Do notMake modifications directly in this article and explain the reasons for the modifications.
                - 「外国人投資家の資金流出」を「外国人投資家からの資金流入」と記載。

            - **（金利の）先高感/先高観 の表記統一:**
                - 文中に「先高観」という表記がある場合でも、原文は修正しないでください。
                - その代わり、「先高観」の直後に「修正提案」として、「先高感」への統一理由を提示してください。
                - 表記がすでに「先高感」である場合は、何も追記せずそのままにしてください。

                - 表示フォーマット（例）:
                    先高観<span style="color:red;">（修正理由: 用語の統一 <s style="background:yellow;color:red">先高観</s> → 先高感）</span>

                - 必ず原文の構成と文脈を保持し、構文を壊さず、修正理由は補足的に後ろに追記してください。
                
            - **ポートフォリオの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「●●への組み入れ（×）」ではなく、「●●の組み入れ（○）」と表記。
                - 「への投資比率」は使用可能。
                
            - **構成比の0％の表記:
                - 「0％程度」or「ゼロ％程度」の表記を使用すること
                - 変更前表記: 構成比は0％である
                - 統一後表記: 構成比は0％程度
                - Append a correction reason in the following format:
                        `<span style="color:red;">変更前表記</span> (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: 構成比は0％である
                    Output: 
                    <span style="color:red;">構成比は0％である</span> 
                    (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">構成比は0％である</s> → 構成比は0％程度である。</span>)で売られています。

                
            - **'投資環境の記述:** 
                - Make modifications directly in this article and explain the reasons for the modifications.
                **「先月の投資環境」**の部分で「先月末」の記述が含まれる場合、「前月末」に変更してください。
                - Ensure that the original text is directly modified and follows this guideline.
                Example:
                修正前: 先月末の市場動向を分析すると…
                修正後: 前月末の市場動向を分析すると…

            - **通貨表記の統一:**
                - Standardize currency notation across the document.
                    - The first appearance of any currency symbol (e.g., ドル, $, 円, JPY) will be the standard.
                    - All following occurrences of that currency must match this format.

                    - For example, if "100ドル" appears first, then all future "$100" will be rewritten to "100ドル" for consistency.
                    - If "$100" appears first, then "100ドル" should be rewritten as "$100".

                    - Always apply this rule in the direction of "first-appeared" format.
                    - Append a correction reason in the following format:
                        `<span style="color:red;">統一後表記</span> (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: このバッグは100ドルですが、アメリカでは$100で売られています。
                    Output:
                    このバッグは100ドルですが、アメリカでは
                    <span style="color:red;">100ドル</span>
                    (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">$100</s> → 100ドル</span>)で売られています。


            **Preferred and Recommended Terminology (置き換えが必要な用語/表現):**
            - **第1四半期:**
                - Ensure the period is clearly stated.
                - Example: 18年第4四半期（×） → 2018年10-12月期（○）
            - **約○％程度:**
                - Do not use "約" (approximately) and "程度" (extent) together. Choose either one.
                - Example: 約○％程度（×） → 約○％ or ○％程度（○）
            - **大手企業表記の明確化**  
                **Correction Rule:**
                - If a sentence contains vague expressions like「大手○○」, analyze the context to determine what type of company is being referred to.
                - Rewrite it in the format:「大手○○会社」/「大手○○企業」/「大手○○メーカー」depending on the company’s nature.
                - Use context clues (e.g., product type, industry references) to guess the appropriate company category (e.g., 不動産, 自動車, 電機, 金融).
                - Append a correction reason in this format:
                `<span style="color:red;">Changed Expression</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Expression</s> → Changed Expression</span>)`

                **Example Input:**
                - 大手は業界全体に影響力を持つ。
                - 大手が新しい半導体を発表した。

                **Example Output:**
                - <span style="color:red;">大手不動産会社</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手不動産会社</span>) は業界全体に影響力を持つ。
                - <span style="color:red;">大手半導体メーカー</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手半導体メーカー</span>) が新しい半導体を発表した。

                **Important Notes:**
                - Always preserve the original sentence structure and paragraph formatting.
                - Only make corrections when「○○大手」is ambiguous and can be clarified using contextual information.
                - Do not modify proper nouns or known company names (e.g., トヨタ, ソニー).

            - **入力例:**  
                - 「大手メーカー/会社/企業」  
                - **出力:** 「大手不動産会社、大手半導体メーカー」  
            - **The actual company name must be found and converted in the article
            - **先月/前月の表記:
                - 1ヵ月前について言及する場合は、「前月」を使用。
            前期比○％の表記:

            - **前期比年率○％:**
                - 基本的に、期間比較の伸率は「年率」を記載してください。
                - 主に経済統計等で一般的に前期比で年率換算されているものについては、「前期比年率○％」と表記。
            - **第○四半期の表記:**
                **ルール:
                - If the input contains a format like "18年第4四半期", infer it as:
                    - "18年" → "2018年"
                    - "第1四半期" → "1-3月期"
                    - "第2四半期" → "4-6月期"
                    - "第3四半期" → "7-9月期"
                    - "第4四半期" → "10-12月期"
                - Modify the expression accordingly, converting the year to a 4-digit format and specifying the exact month range.
                - Add a correction reason in this format:
                `<span style="color:red;">修正後</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">修正前</s> → 修正後</span>)`

                ---

                **Example:**
                - Input: 18年第4四半期の売上が好調だった。
                - Output: 
                <span style="color:red;">2018年10-12月期</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">18年第4四半期</s> → 2018年10-12月期</span>) の売上が好調だった。

                ---

                **Additional Notes:**
                - Do not modify any proper names, organizations, or if the date range is already correct.
                - Apply to all similar shorthand expressions like "20年第2四半期", "21年第1四半期" etc.
                - Keep the structure and formatting of the original document.


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).

        **Output Requirements:**
        1. **Highlight the original incorrect text in red and include additional details:**
        - For corrected parts:
            - Highlight the original incorrect text in red using `<span style="color:red;">`.
            - Append the corrected text in parentheses, marked with a strikethrough using `<s>` tags.
            - Provide the reason for the correction and indicate the change using the format `123 → 456`.
            - Example:
            `<span style="color:red;">123</span> (<span>修正理由: 一致性不足 <s style="background:yellow;color:red">123</s> → 456</span>)`
        
        2. **Preserve the original structure and formatting of the document:**
        - Maintain paragraph breaks, headings, and any existing structure in the content.

        3. **Use the uploaded correction rules for reference:**
        - {corrected}

        4. **Do not provide any explanations or descriptions in the output. Only return the corrected HTML content.**

         **Corrected Terminology Map (修正された用語リスト):
            {corrected_map}
        - Replace only when the **original** term in `corrected_map` appears in the input text.
        - Do **not** replace anything if the input already contains the `corrected` term (it is already correct).
        - Do **not** perform any reverse replacements (`corrected → original` は禁止).
        - Modify the original text only when the `original` term is found.

        - If the `corrected` term appears in the input, **do not modify it** (it is already correct).
        - Do **not** reverse substitutions (i.e., never convert corrected → `original`).
        
        - After replacing, add the reason in this format:
        Original Term (修正理由: 用語の統一 Original Term → Corrected Term)
        Example:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`
        
        Example:
        Input: 中銀
        Output: 
        `<span style="color:red;">中央銀行</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">中銀</s> → 中央銀行</span>)`
        ※ Note: Do **not** convert 中央銀行 → 中銀. All replacements must follow the direction from `original` to `corrected` only.

        Input: 中央銀行  
        Output:  
        中央銀行 ← (No correction shown because it is already the correct term)

        If the input already contains the corrected term, it should remain unchanged.
        For English abbreviations or foreign terms, the rule is the same: replace the original term with the corrected term and format as follows:
        Example:
        Input: BOE
        Output: <span style="color:red;">BOE（英中央銀行、イングランド銀行）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BOE</s> → BOE（英中央銀行、イングランド銀行）</span>)
        Input: AAA
        Output: <span style="color:red;">AAA（全米自動車協会）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AAA</s> → AAA（全米自動車協会）</span>)

        Input: インバウンド
        Output: <span style="color:red;">インバウンド（観光客の受け入れ）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">インバウンド</s> → インバウンド（観光客の受け入れ）</span>)

        
        **Except Original Term
        Input: 等
        Output: 
        `<span style="color:red;">等</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">等</s> → など</span>)`

        Input: ローン
        Output: 
        `<span style="color:red;">ローン</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">ローン</s> → 貸し付け</span>)`
                            
        Input: ％を上回る
        Output: 
        `<span style="color:red;">％を上回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を上回る</s> → ％を超える</span>)`
                            
        Input: ％を下回る
        Output: 
        `<span style="color:red;">％を下回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を下回る</s> → ％を下回るマイナス幅</span>)`
        
        Input: 伝播（でんぱ）
        Output: 
        `<span style="color:red;">伝播（でんぱ）</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播（でんぱ）</s> → 広ります</span>)`
        
        Input: 伝播しています
        Output:
        `<span style="color:red;">伝播しています</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播しています</s> → 広がるしています</span>)`
        
        Input: 連れ高
        Output:
        `<span style="color:red;">連れ高</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">連れ高</s> → 影響を受けて上昇</span>)`
        
        Input: 相場
        Output:
        `<span style="color:red;">相場</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">相場</s> → 市場/価格</span>)`
        
        Input: ハト派
        Output:
        `<span style="color:red;">ハト派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">ハト派</s> → 金融緩和重視、金融緩和に前向き</span>)`
        
        Input: タカ派
        Output:
        `<span style="color:red;">タカ派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">タカ派</s> → 金融引き締め重視、金融引き締めに積極的</span>)`
        
        Input: 織り込む
        Output: 
        `<span style="color:red;">織り込む</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">織り込む</s> → 反映され</span>)`
        
        Input: 積極姿勢とした
        Output: 
        `<span style="color:red;">積極姿勢とした</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">積極姿勢とした</s> → 長めとした</span>)`
        
        Input: 限定的
        Output: 
        `<span style="color:red;">限定的</span> (<span>修正理由: 効果や影響がプラスかマイナスか不明瞭なため <s style="background:yellow;color:red">限定的</s> → 他の適切な表現に修正</span>)`
        
        Input: 利益確定の売り
        Output: 
        `<span style="color:red;">利益確定の売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利益確定の売り</s> → が出たとの見方</span>)`
        
        Input: 利食い売り
        Output: 
        `<span style="color:red;">利食い売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利食い売り</s> → が出たとの見方</span>)`
        
        Input: ABS
        Output: 
        `<span style="color:red;">ABS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ABS</s> → ABS（資産担保証券、各種資産担保証券）</span>)`
        
        Input: AI
        Output: 
        `<span style="color:red;">AI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AI</s> → AI（人工知能</span>)`
        
        Input: BRICS（5ヵ国）
        Output: 
        `<span style="color:red;">BRICS（5ヵ国）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BRICS（5ヵ国）</s> → BRICS（ブラジル、ロシア、インド、中国、南アフリカ）</span>)`
        
        Input: CMBS
        Output: 
        `<span style="color:red;">CMBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">CMBS</s> → CMBS（商業用不動産ローン担保証券）</span>)`
        
        Input: ISM
        Output: 
        `<span style="color:red;">ISM</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ISM</s> → ISM（全米供給管理協会）</span>)`
        
        Input: IT
        Output: 
        `<span style="color:red;">IT</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">IT</s> → IT（情報技術）</span>)`
        
        Input: MBS
        Output: 
        `<span style="color:red;">MBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">MBS</s> → MBS（住宅ローン担保証券）</span>)`
        
        Input: PMI
        Output: 
        `<span style="color:red;">PMI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">PMI</s> → PMI（購買担当者景気指数）</span>)`
        
        Input: S&P
        Output: 
        `<span style="color:red;">S&P</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">S&P</s> → S&P（スタンダード・アンド・プアーズ）社</span>)`
        
        Input: アセットアロケーション
        Output: 
        `<span style="color:red;">アセットアロケーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">アセットアロケーション</s> → アセットアロケーション（資産配分）</span>)`
        
        Input: E-コマース
        Output: 
        `<span style="color:red;">Eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">E-コマース</s> → Eコマース（電子商取引）</span>)`
             
        Input: e-コマース
        Output: 
        `<span style="color:red;">eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">eコマース</s> → eコマース（電子商取引）</span>)`
           
        Input: EC
        Output: 
        `<span style="color:red;">EC（電子商取引）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">EC</s> → EC（電子商取引）</span>)`
        
        Input: イールドカーブ
        Output: 
        `<span style="color:red;">イールドカーブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">イールドカーブ</s> → イールドカーブ（利回り曲線）</span>)`
        
        Input: エクスポージャー
        Output: 
        `<span style="color:red;">エクスポージャー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">エクスポージャー</s> → ＊積極的に使用しない。　（価格変動リスク資産の配分比率、割合）</span>)`
        
        Input: クレジット（信用）市場
        Output: 
        `<span style="color:red;">クレジット（信用）市場</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">クレジット（信用）市場</s> → 信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称。　企業の信用リスクを取引する市場。</span>)`
        
        Input: システミック・リスク
        Output: 
        `<span style="color:red;">システミック・リスク</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">システミック・リスク</s> → 個別の金融機関の支払不能等や、特定の市場または決済システム等の機能不全が、他の金融機関、他の市場、または金融システム全体に波及するリスク</span>)`
        
        Input: ディストレス債券
        Output: 
        `<span style="color:red;">ディストレス債券</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディストレス債券</s> → 信用事由などにより、価格が著しく下落した債券</span>)`
        
        Input: ディフェンシブ
        Output: 
        `<span style="color:red;">ディフェンシブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディフェンシブ</s> → ディフェンシブ（景気に左右されにくい）</span>)`
        
        Input: テクニカル
        Output: 
        `<span style="color:red;">テクニカル</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">テクニカル</s> → テクニカル（過去の株価の動きから判断すること</span>)`
        
        Input: デフォルト
        Output: 
        `<span style="color:red;">デフォルト</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト</s> → デフォルト（債務不履行）</span>)`
        
        Input: デフォルト債
        Output: 
        `<span style="color:red;">デフォルト債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト債</s> → デフォルトとは一般的には債券の利払いおよび元本返済の不履行、もしくは遅延などをいい、このような状態にある債券をデフォルト債といいます。</span>)`
        
        Input: デュレーション
        Output: 
        `<span style="color:red;">デュレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デュレーション</s> → デュレーション（金利感応度）</span>)`
        
        Input: 投資適格債
        Output: 
        `<span style="color:red;">投資適格債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">投資適格債</s> → 格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債</span>)`
        
        Input: ファンダメンタルズ
        Output: 
        `<span style="color:red;">ファンダメンタルズ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ファンダメンタルズ</s> → ファンダメンタルズ（賃料や空室率、需給関係などの基礎的条件）※REITファンドで使用する</span>)`
        
        Input: フリーキャッシュフロー
        Output: 
        `<span style="color:red;">フリーキャッシュフロー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">フリーキャッシュフロー</s> → 税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたもの</span>)`
        
        Input: ベージュブック
        Output: 
        `<span style="color:red;">ベージュブック</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ベージュブック</s> → スペースがない場合は、ベージュブック（米地区連銀経済報告）</span>)`
        
        Input: モメンタム
        Output: 
        `<span style="color:red;">モメンタム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">モメンタム</s> →  相場の勢い)が強く、投資家たちは短期的な利益を狙っています。</span>)`
        
        Input: リオープン
        Output: 
        `<span style="color:red;">リオープン</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リオープン</s> → リオープン/リオープニング（経済活動再開）</span>)`
        
        Input: リスクプレミアム
        Output: 
        `<span style="color:red;">リスクプレミアム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リスクプレミアム</s> → 同じ投資期間内において、あるリスク資産の期待収益率が、無リスク資産（国債など）の収益率を上回る幅のこと。</span>)`
        
        Input: リフレーション
        Output: 
        `<span style="color:red;">リフレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リフレーション</s> → リフレーション**デフレーションから抜けて、まだ、インフレーションにはなっていない状況のこと。</span>)`
        
        **【例外用語 – 修正しないこと】**
        - コロナ禍
        - コロナショック
        - 新型コロナ禍
        - 住宅ローン
        - 引き締め策
        - 引き締め政策
        - 組入比率
        - 格付機関
        - 格付別
        - 国債買入オペ

        **特定表現の言い換えルール（文脈判断を伴う修正）:
        文脈に応じて、具体的な表現に言い換えてください。
        「まちまち」の使用
        「まちまち」という曖昧な表現が出現した場合は、その語をそのまま保持した上で、後に「修正理由: 曖昧表現の明確化」を補足してください。

        変換語「異なる動き」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        <span style="color:red;">まちまち</span>
        (<span>修正理由: 曖昧表現の明確化 <s style="background:yellow;color:red">まちまち</s> → 異なる動き</span>)

        -「行って来い」の表現

        文脈に応じて、「上昇（下落）したのち下落（上昇）」のように明確にしてください。
        Exsample:

        Input: 相場は行って来いの展開となった
        Output: 相場は上昇したのち下落する展開となった
        
        変換語「行って来い」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        
        修正理由: 表現の明確化
        <span style="color:red;">行って来い</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Term</s> → 行って来い</span>)

        -「横ばい」表現の適正使用

        小幅な変動であれば「横ばい」を使用可能。
        大きな変動の末に同水準で終了した場合は、「ほぼ変わらず」「同程度となる」などに修正。
        
        変換語「横ばい」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 用語の適正使用
        <span style="color:red;">横ばい</span> (<span>修正理由: 用語の適正使用 <s style="background:yellow;color:red">横ばい</s> → ほぼ変わらず</span>)

        -「（割安に）放置」表現の修正

        「割安感のある」など、より適切な表現に修正してください。

        Exsample:
        Input: 株価は割安に放置された
        Output: 株価には割安感がある状態が続いた

        変換語「（割安に）放置」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 表現の明確化と客観性の向上
        <span style="color:red;">（割安に）放置</span> (<span>修正理由: 表現の明確化と客観性の向上 <s style="background:yellow;color:red"（割安に）放置</s> → 割安感のある</span>)

        """  
        # ChatCompletion Call
        response = openai.ChatCompletion.create(
        # OpenAI API 호출을 asyncio에서 비동기로 실행
        # loop = asyncio.get_event_loop()
        # response = await loop.run_in_executor(None, lambda: openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."
                "This includes not only Japanese text but also English abbreviations (英略語), "
                "foreign terms (外来語),and specialized terminology (専門用語)."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=32768,
            temperature=0,
            seed=42  # 재현 가능한 결과를 위해 seed 설정
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)
        
        return jsonify({"success": True, "corrected_text": re_answer})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def convert_logs(items):
    converted_data = {
        "code": 200,
        "message": "成功",
        "data": []
    }
    
    # 주어진 아이템들을 변환
    for idx, item in enumerate(items):
        log_entries = item.get("logEntries", [])
        
        for log_idx, log_entry in enumerate(log_entries):
            log_parts = log_entry.split(" - ")
            timestamp_str = log_parts[0] if len(log_parts) > 1 else ""
            message = log_parts[1] if len(log_parts) > 1 else ""
            
            # 로그 항목을 변환하여 새로운 형식에 맞추기
            log_data = {
                "id": idx * len(log_entries) + log_idx + 1,  # ID를 고유하게 설정
                "name": message,  # message[:30] 메시지 첫 30자만을 'name'으로 사용
                "status": "完了(修正あり)" if "✅ SUCCESS" in message else "エラー",  # 성공 메시지에 따라 상태 결정
                "timeclock": timestamp_str,
                "progress": "成功" if "✅ SUCCESS" in message else "エラー",  # 성공 여부로 진행 상태 결정
                "timestamp": timestamp_str,
                "selected": False  # 기본적으로 'selected' 값은 False
            }
            
            converted_data["data"].append(log_data)
    
    # 반환할 JSON 포맷 구조에 맞추어 반환
    return converted_data

# appLog
APPLOG_CONTAINER_NAME='appLog'
@app.route('/api/applog', methods=['GET'])
def get_applog():
    # Cosmos DB 연결
    container = get_db_connection(APPLOG_CONTAINER_NAME)

    # Cosmos DB에서 데이터를 쿼리하여 가져오기
    query = "SELECT * FROM c"  # SQL 쿼리 (전체 데이터를 가져옴)
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    # 각 아이템의 '_id'는 이미 문자열로 되어 있기 때문에 추가적인 변환이 필요하지 않음
    for item in items:
        item['id'] = item['id']  # Cosmos DB에서 _id는 id로 제공

    # return jsonify(items)  # JSON 형식으로 반환
    # 변환 실행
    converted_logs = convert_logs(items)
    return jsonify(converted_logs)  # JSON 형식으로 반환

# azure Cosmos DB
@app.route('/api/faqs', methods=['GET'])
def get_faq():
    # Cosmos DB 클라이언트 연결,ENV에서 받아온다다
    container=get_db_connection()

    # Cosmos DB에서 데이터를 쿼리하여 가져오기
    query = "SELECT * FROM c"  # SQL 쿼리 (전체 데이터를 가져옴)
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    # 각 아이템의 '_id'는 이미 문자열로 되어 있기 때문에 추가적인 변환이 필요하지 않음
    for item in items:
        item['id'] = item['id']  # Cosmos DB에서 _id는 id로 제공

    return jsonify(items)  # JSON 형식으로 반환


# 527 debug
@app.route('/api/tenbrend', methods=['POST'])
def tenbrend():
    data = request.get_json() or {}

    raw_fcode = data.get('fcode', '').strip()
    months = data.get('month', '').strip()
    stocks = data.get('stock', '').strip()
    fund_type = data.get('fundType', 'public').strip()  # 默认为公募

    # 根据 fundType 选择容器（即 Cosmos DB 的表）

    if fund_type == 'private':
        TENBREND_CONTAINER_NAME = 'tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME = 'tenbrend'

    # Cosmos DB 클라이언트 연결
    container = get_db_connection(TENBREND_CONTAINER_NAME)
    parameters = []
    if not raw_fcode:
        query = "SELECT * FROM c"

    else:
        # 构建 SQL 查询
        if '-' in raw_fcode:
            # 带 `-` 的直接用字符串模糊匹配
            query = "SELECT * FROM c WHERE CONTAINS(c.fcode, @fcode)"
            parameters.append({"name": "@fcode", "value": raw_fcode})
        else:
            try:
                fcode_num = int(raw_fcode)
                query = "SELECT * FROM c WHERE c.fcode = @fcode"
                parameters.append({"name": "@fcode", "value": fcode_num})
            except ValueError:
                # fallback 到字符串查询
                query = "SELECT * FROM c WHERE CONTAINS(c.fcode, @fcode)"
                parameters.append({"name": "@fcode", "value": raw_fcode})

        if months:
            query += " AND c.months = @months"
            parameters.append({"name": "@months", "value": months})

        if stocks:
            query += " AND CONTAINS(c.stocks, @stocks)"
            parameters.append({"name": "@stocks", "value": stocks})

    items = list(container.query_items(
        query=query,
        parameters=parameters,
        enable_cross_partition_query=True
    ))

    filtered_items = [item for item in items if item.get('id')]
    return jsonify({"code": 200, "data": filtered_items})


@app.route('/api/tenbrend/months', methods=['POST'])
def tenbrend_months():
    data = request.get_json() or {}

    fcode = data.get('fcode', '').strip()
    stocks = data.get('stock', '').strip() if data.get('stock') else ''
    fund_type = data.get('fundType', 'public').strip()

    if not fcode:
        return jsonify({"code": 400, "message": "fcode is required"}), 400

    # ✅ 根据 fundType 切换容器（表）
    if fund_type == 'private':
        TENBREND_CONTAINER_NAME ='tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME='tenbrend'

    # Cosmos DB 클라이언트 연결
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    query = "SELECT c.months FROM c WHERE CONTAINS(c.fcode, @fcode)"
    parameters = [{"name": "@fcode", "value": fcode}]

    if stocks:
        query += " AND CONTAINS(c.stocks, @stocks)"
        parameters.append({"name": "@stocks", "value": stocks})

    try:
        items = list(container.query_items(
            query=query,
            parameters=parameters,
            enable_cross_partition_query=True
        ))

        months = sorted({item.get('months') for item in items if item.get('months')})
        return jsonify({"code": 200, "data": months})
    except Exception as e:
        print("❌ Cosmos DB query failed:", e)
        return jsonify({"code": 500, "message": "internal error"}), 500


@app.route('/api/tenbrend/stocks', methods=['POST'])
def tenbrend_stocks():
    data = request.get_json() or {}

    fcode = data.get('fcode', '').strip()
    months = data.get('month', '').strip() if data.get('month') else ''
    fund_type = data.get('fundType', 'public').strip()

    if not fcode:
        return jsonify({"code": 400, "message": "fcode is required"}), 400

    if fund_type == 'private':
        TENBREND_CONTAINER_NAME ='tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME='tenbrend'

    # Cosmos DB 클라이언트 연결
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    query = "SELECT c.stocks FROM c WHERE CONTAINS(c.fcode, @fcode)"
    parameters = [{"name": "@fcode", "value": fcode}]

    if months:
        query += " AND c.months = @months"
        parameters.append({"name": "@months", "value": months})

    try:
        items = list(container.query_items(
            query=query,
            parameters=parameters,
            enable_cross_partition_query=True
        ))

        stocks = sorted({item.get('stocks') for item in items if item.get('stocks')})
        return jsonify({"code": 200, "data": stocks})
    except Exception as e:
        print("❌ Cosmos DB query failed:", e)
        return jsonify({"code": 500, "message": "internal error"}), 500




@app.route('/api/tenbrend/template', methods=['GET'])
def download_excel_template():
    data = request.get_json() or {}
    # 默认是“公募”
    fund_type = data.get('fundType', 'public').strip()

    # 根据类型拼接路径
    if fund_type == '私募':
        file_url = "https://nriazureaistudio.blob.core.windows.net/1225-container/10銘柄マスタ管理_私募.xlsx"
    else:
        file_url = "https://nriazureaistudio.blob.core.windows.net/1225-container/10銘柄マスタ管理_公募.xlsx"

    try:
        # 注意:send_file 不能直接下载远程链接，改为重定向
        return redirect(file_url)
    except Exception as e:
        return jsonify({"code": 500, "message": str(e)}), 500



# Data transfer
def transform_data(items,fund_type):
    menu_data = {
        "公募": [],
        "私募": []
    }

    for item in items:
        if fund_type == 'public':
            fund_category = menu_data["公募"]
        elif fund_type == 'private':
            fund_category = menu_data["私募"]
        else:
            continue  # 잘못된 fund_type은 무시

        # 데이터 구조에 맞게 변환
        reference = {
            "id": "reference",
            "name": "📁 参照ファイル",
            "children": [
                {
                    "id": "report_data",
                    "name": "📂 レポートデータ",
                    "children": []
                },
                {
                    "id": "mingbing_data",
                    "name": "📂 10銘柄解説一覧表",
                    "children": []
                }
            ]
        }

        # report_data에 아이템 추가
        reference["children"][0]["children"].append({
            "id": item.get('id'),
            "name": item.get('fileName'),
            "icon": "📄",
            "file": item.get('fileName'),
            "pdfPath": extract_pdf_path(item.get('link')),
        })

        # mingbing_data에 아이템 추가
        reference["children"][1]["children"].append({
            "id": item.get('id'),
            "name": item.get('fileName'),
            "icon": "📄",
            "file": item.get('fileName'),
            "pdfPath": extract_pdf_path(item.get('link'))
        })

        fund_category.append(reference)

        # checked_files 섹션 추가
        checked_files = {
            "id": "checked_files",
            "name": "📁 チェック対象ファイル",
            "children": [
                {
                    "id": "individual_comments",
                    "name": "📂 共通コメントファイル",
                    "children": []
                },
                {
                    "id": "kobetsucomment",
                    "name": "📂 個別コメントファイル",
                    "children": []
                }
            ]
        }

        # individual_comments에 아이템 추가
        checked_files["children"][0]["children"].append({
            "id": item.get('id'),  
            "name": item.get('fileName'),  
            "icon": "⚠️",
            "file": item.get('fileName'),  
            "status": item.get('comment_status'),  
            "readStatus": item.get('comment_readStatus'),  
            "pdfPath": extract_pdf_path(item.get('link'))  
        })

        # kobetsucomment에 아이템 추가
        checked_files["children"][1]["children"].append({
            "id": item.get('id'),  
            "name": item.get('fileName'),
            "icon": "❌",
            "file": item.get('fileName'),
            "status": item.get('individual_status'),
            "readStatus": item.get('individual_readStatus'),
            "pdfPath": extract_pdf_path(item.get('link'))  
        })

        fund_category.append(checked_files)

    return menu_data

def extract_pdf_path(link):
    """ <a href="URL"> 에서 URL을 추출하는 함수 """
    match = re.search(r'href="([^"]+)"', link)
    return match.group(1) if match else ""

def extract_base_name(file_path):
    """파일 경로에서 확장자 앞의 이름 추출"""
    # 파일 이름 추출 (경로 제거)
    file_name = os.path.basename(file_path)
    
    # 확장자 분리
    base_name, _ = os.path.splitext(file_name)
    
    return base_name

# public_Fund and private_Fund
@app.route('/api/fund', methods=['POST'])
def handle_fund():
    # 파라미터 유효성 검사
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # 쿼리 실행
        query = "SELECT * FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # 결과 필터링
        # filtered_items = [item for item in items if item and item.get('id')]
        
        # return jsonify(filtered_items)
        # 데이터 변환 (트리 구조)
        formatted_data = transform_data(items,fund_type)

        return jsonify(formatted_data)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# 625 tenbrend
def convert_to_tenbrend(items):
    corrections = []

    for item in items:
        old_text = item.get("元組入銘柄解説", "").strip()
        new_text = item.get("新組入銘柄解説", "").strip()

        if old_text != new_text:
            corrections.append({
                "check_point": "組入銘柄解説",
                "comment": f"{old_text} → {new_text}",
                "intgr": False,
                "locations": [],
                "original_text": new_text,
                "page": '',
                "reason_type": item.get("stocks", "")
            })

    return corrections


# 509 debug
def convert_format(filtered_items):
    checkResults = {}

    for correction in filtered_items.get("result", {}).get("corrections", []):
        page = correction["page"] + 1  # 페이지 번호를 1부터 시작하도록 변환
        position = {}

        # 수정 내역 변환
        change = {
            "before": correction["original_text"],
            "after": correction["comment"].split("→")[-1].strip(),
        }
        if correction["intgr"]:
            name = "不一致"
        else:
            name = ""

        # 위치 정보가 있는 경우 변환
        if correction["locations"]:
            for idx, loc in enumerate(correction["locations"]):  # 모든 위치 정보를 처리
                pdf_height = loc.get("pdf_height", 792)  # PDF 높이 (기본값: A4 크기, 792pt)

                # 첫 번째 위치만 x 좌표를 조정
                #debug
                # x = loc["x0"] - 22 if idx == 0 else loc["x0"]
                position = {
                    "x": loc["x0"],  # x 좌표는 그대로 사용
                    "y": pdf_height - loc["y1"] + 50,  # y 좌표를 PDF 높이를 기준으로 변환
                    "width": loc["x1"] - loc["x0"],  # 너비 계산
                    "height": loc["y1"] - loc["y0"],  # 높이 계산
                }

                # checkResults에 페이지별 그룹화
                if page not in checkResults:
                    checkResults[page] = [{"title": filtered_items["fileName"], "items": []}]

                # 중복 체크
                existing_item = next((item for item in checkResults[page][0]["items"] if item["name"] == name and item["changes"] == [change]), None)
                if not existing_item:
                    checkResults[page][0]["items"].append({
                        "name": name,
                        "color": "rgba(255, 255, 0, 0.5)", # green background rgba(0, 255, 0, 0.5)
                        "page": page,
                        "position": position,
                        "changes": [change],
                        "reason_type":correction["reason_type"], # for intergarteion reasion
                        "check_point":correction["check_point"], # 66 debug, add the check point
                    })

    return {'data': checkResults, 'code': 200}

# public_Fund and check-results
@app.route('/api/check_results', methods=['POST'])
def handle_check_results():
    # 파라미터 유효성 검사
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400
    
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400
    
    pageNumber = request.json.get('pageNumber')
    if not pageNumber:
        return jsonify({"error": "pageNumber is required"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # Cosmos DB에서 특정 id로 데이터를 쿼리하여 가져오기
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))
        
        if not items:
            return jsonify({"error": "Item not found"}), 404

        # 데이터 변환 (트리 구조)
        converted_data = convert_format(items[0])

        return jsonify(converted_data)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# get side bar
@app.route('/api/menu', methods=['POST'])
def handle_menu():
    # 파라미터 유효성 검사
    fund_type = request.json.get('type')
    page = int(request.json.get('page', 1))
    page_size = int(request.json.get('page_size', 10))
    # user_name = request.json.get('user_name')

    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"

    
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # 쿼리 실행
        query = "SELECT * FROM c WHERE CONTAINS(c.id, '.pdf') OR c.upload_type='参照ファイル'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # 결과 필터링
        filtered_items = [item for item in items if item and item.get('id')]

        # 페이지네이션 적용
        total = len(filtered_items)
        start = (page - 1) * page_size
        end = start + page_size
        paged_items = filtered_items[start:end]

        response = {
            "code": 200,
            "data": paged_items,
            "total": total
        }

        return jsonify(response)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/menu_all', methods=['POST'])
def handle_menu_all():
    # 파라미터 유효성 검사
    fund_type = request.json.get('type')

    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # 쿼리 실행
        query = "SELECT * FROM c WHERE CONTAINS(c.id, '.pdf') OR c.upload_type='参照ファイル'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # 결과 필터링
        filtered_items = [item for item in items if item and item.get('id')]
        response = {
        "code": 200,
        "data": filtered_items
        }

        return jsonify(response)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
# Cosmos DB 상태 확인 엔드포인트
# 모니터링 상태 컨테이너 설정
MONITORING_CONTAINER_NAME = "monitoring-status"

# Cosmos DB 상태 확인 엔드포인트
@app.route('/api/monitoring-status', methods=['GET'])
def get_monitoring_status():
    try:
        # Cosmos DB 연결
        container = get_db_connection(MONITORING_CONTAINER_NAME)
        
        # Cosmos DB에서 데이터를 쿼리하여 가져오기
        query = "SELECT * FROM c"  # SQL 쿼리 (전체 데이터를 가져옴)
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        # 각 아이템의 '_id'는 이미 문자열로 되어 있기 때문에 추가적인 변환이 필요하지 않음
        for item in items:
            item['id'] = item['id']  # Cosmos DB에서 _id는 id로 제공

        return jsonify(items), 200  # JSON 형식으로 반환
        
    except CosmosResourceNotFoundError:
        logging.error("Monitoring status document not found")
        return jsonify({"error": "Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# 상태 업데이트 엔드포인트
@app.route('/api/monitoring-status', methods=['PUT'])
def update_monitoring_status():
    try:
        # Cosmos DB 연결
        container = get_db_connection(MONITORING_CONTAINER_NAME)
        
        new_status = request.json.get('status', 'off')
        
        # 문서 업데이트
        status_item = {
            'id': 'monitoring_status',
            'type': 'control',  # 파티션 키 필드
            'status': new_status,
            "timestamp": datetime.utcnow().isoformat()
        }
        
        container.upsert_item(body=status_item)
        logging.info(f"Monitoring status updated to {new_status}")
        return jsonify({'message': 'Status updated', 'new_status': new_status, 'code': 200}), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Database operation failed"}), 500
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# read/unread status change
# Cosmos DB 상태 확인 엔드포인트
@app.route('/api/update_read_status', methods=['POST'])
def get_read_status():
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400
    
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # Cosmos DB에서 특정 id로 데이터를 쿼리하여 가져오기
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))
        
        if not items:
            return jsonify({"error": "Item not found"}), 404

        # 조회된 항목 반환
        return jsonify(items[0]), 200
        
    except CosmosResourceNotFoundError:
        logging.error("read status document not found")
        return jsonify({"error": "read Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/update_read_status', methods=['PUT'])
def update_read_status():
    # 파라미터 유효성 검사
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400

    mark = request.json.get('mark')
    if mark not in ['read', 'unread']:
        return jsonify({"error": "Invalid mark value"}), 400

    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # 컨테이너 이름 결정
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 연결
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))

        if not items:
            return jsonify({"error": "Item not found"}), 404

        # 조회된 문서 가져오기
        status_item = items[0]

        # readStatus와 timestamp만 업데이트
        status_item['readStatus'] = mark
        status_item['timestamp'] = datetime.utcnow().isoformat()
        
        # 문서 업데이트
        container.upsert_item(body=status_item)
        logging.info(f"readStatus updated to {mark} for item {selected_id}")
        return jsonify({'message': 'Status updated', 'new_status': mark, 'code': 200}), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Item not found"}), 404
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500


@app.route("/api/health")
def health_check():
    return "OK", 200

# 로깅 설정
logging.basicConfig(level=logging.INFO)

def get_storage_container():
    """
    Azure AD RBAC 방식을 사용하여 Azure Blob Storage에 연결하고, ContainerClient 객체를 반환하는 함수.
    :return: ContainerClient 객체
    """
    try:
        # BlobServiceClient 생성
        blob_service_client = BlobServiceClient(account_url=ACCOUNT_URL, credential=credential)
        
        # 컨테이너 클라이언트 가져오기
        container_client = blob_service_client.get_container_client(STORAGE_CONTAINER_NAME)
        
        print("Connected to Azure Blob Storage via Azure AD RBAC")
        logging.info("Connected to Azure Blob Storage via Azure AD RBAC")
        
        return container_client
    except Exception as e:
        logging.error(f"Azure Blob Storage Connection Error: {e}")
        print(f"Azure Blob Storage Connection Error: {e}")
        raise e
    
def allowed_file(filename):
    """
    업로드 가능한 파일 형식을 확인하는 함수.
    
    :param filename: 파일명
    :return: bool
    """
    ALLOWED_EXTENSIONS = {'pdf', 'xlsx','txt','xls','xlsx','xlm','xlmx'}  # PDF와 Excel 파일 허용 
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/test_token', methods=['GET'])
def test_token():
    try:
        token_cache._refresh_token()
        token = token_cache.get_token()
        return jsonify({"access_token": token}), 200
    except Exception as e:
        logging.exception("Token Get Error")
        return jsonify({"message": f"Token Get Error: {str(e)}"}), 500

# uploadpdf,api/brand
def parse_escaped_json(raw_text: str):
    text = raw_text.strip()
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    
    text = text.replace('```json', '')
    text = text.replace('```', '')

    text = text.replace('""', '"')

    # 5) 이제 text는 완전한 JSON 문자열이므로 json.loads로 파싱
    parsed = json.loads(text)
    return parsed

def parse_gpt_response(answer):
    """GPT 응답에서 유효한 JSON 추출"""
    try:
        # JSON 형식 직접 추출 시도
        json_str = re.search(r'\{[\s\S]*?\}', answer).group()
        return json.loads(json_str)
    except (AttributeError, json.JSONDecodeError):
        # Python 딕셔너리 형식 대응
        dict_str = re.search(r'corrected_map\s*=\s*\{[\s\S]*?\}', answer, re.DOTALL)
        if dict_str:
            dict_str = dict_str.group().split('=', 1)[1].strip()
            return ast.literal_eval(dict_str)
        return {}

def detect_corrections(original, corrected):
    """정확한 변경 부분 감지"""
    matcher = SequenceMatcher(None, original, corrected)
    corrections = {}
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'replace':
            orig_part = original[i1:i2].strip()
            corr_part = corrected[j1:j2].strip()
            if orig_part and corr_part:
                corrections[orig_part] = corr_part
    return corrections

def filter_corrected_map(corrected_map):
    # 불필요한 키와 값을 제거
    keys_to_remove = [" ", "  "]
    for key in keys_to_remove:
        if key in corrected_map:
            del corrected_map[key]
    return corrected_map

# 512 debug
def apply_corrections(input_text, corrected_map):
    result = input_text


    for original, corrected in corrected_map.items():

        # original == corrected 면 치환하지 않음
        if result == corrected:
            continue

        if re.search(re.escape(corrected), result):
            # corrected 안에 original이 포함된 경우에는 중복 방지를 위해 skip
            continue

        # 이미 이 용어(original)가 교정된 형식으로 존재하는 경우는 제외
        pattern_already_corrected = re.compile(
            rf"<span style=\"color:red;\">{re.escape(corrected)}</span>\s*"
            rf"\(<span>修正理由: 用語の統一\s*<s style=\"background:yellow;color:red\">{re.escape(original)}</s>\s*→\s*{re.escape(corrected)}</span>\)"
        )
        if pattern_already_corrected.search(result):
            continue

        # original이 존재할 때만 치환
        if re.search(original, result):
            replacement = (
                f'<span style="color:red;">{corrected}</span> '
                f'(<span>修正理由: 用語の統一 '
                f'<s style="background:yellow;color:red">{original}</s> → {corrected}</span>)'
            )
            result = result.replace(original, replacement)

    return result


DICTIONARY_CONTAINER_NAME = "dictionary"
def fetch_and_convert_to_dict():
    """DB에서 데이터를 읽어서 딕셔너리로 변환"""
    try:
        container = get_db_connection(DICTIONARY_CONTAINER_NAME)
        # 🔹 모든 데이터를 가져오기
        query = "SELECT c.original, c.corrected FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        # 🔹 변환된 딕셔너리 생성
        corrected_dict = {item["original"]: item["corrected"] for item in items if "original" in item and "corrected" in item}

        return corrected_dict

    except CosmosHttpResponseError as e:
        print(f"❌ DB error: {e}")
        return {}
    
@app.route('/api/check_upload', methods=['POST'])
def check_upload():
    if 'files' not in request.files:
        return jsonify({"success": False, "message": "No files part"}), 400

    files = request.files.getlist('files')
    file_type = request.form.get("fileType")
    fund_type = request.form.get("fundType")

    for file in files:
        if file.filename == '':
            return jsonify({"success": False, "error": "No selected file"}), 400


        # 파일을 메모리에서 읽기
        file_bytes = file.read()

        if file and allowed_file(file.filename):
            try:
                if file.filename.endswith('.pdf'):  
                    tenbrend_data = check_tenbrend(file.filename,fund_type)
                    # PDF 파일 처리
                    reader = PdfReader(io.BytesIO(file_bytes))
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text()

                    # Encode the PDF bytes to Base64
                    file_base64 = base64.b64encode(file_bytes).decode('utf-8')

                    return jsonify({
                        "success": True,
                        "original_text": extract_text_from_base64_pdf(file_bytes),  # file_bytesPDF text순서  , input = extract_text_from_base64_pdf(pdf_base64)
                        "pdf_bytes": file_base64,  # PDF 파일의 Base64 인코딩
                        "file_name": file.filename,
                        "tenbrend_data":tenbrend_data,
                        # "fund_type": fund_type
                    })
                
                elif file.filename.endswith('.txt'):
                    
                    text = file_bytes.decode('utf-8')  # UTF-8 인코딩 사용

                    return jsonify({
                        "success": True,
                        "prompt_text": text  # Excel에서 읽어온 원본 텍스트
                    })

                elif regcheck.search(r'\.(xls|xlsx|xlm|xlmx)$',file.filename):
                    """
                    엑셀 파일을 처리하여 GPT로 교정 후, 수정된 엑셀을 반환하는 함수
                    :param file_bytes: 업로드된 엑셀 파일 바이너리
                    :return: 수정된 엑셀 바이너리 (Base64 인코딩)
                    """
                    # 🔹 1️⃣ corrected_map 초기화 (에러 방지)
                    corrected_map = fetch_and_convert_to_dict()
                    all_text=[]

                    # 🔹 2️⃣ 압축 해제용 임시 폴더(또는 메모리상 in-memory zip)
                    in_memory_zip = zipfile.ZipFile(io.BytesIO(file_bytes), 'r')

                    # 새 ZIP(수정본)을 만들기 위한 BytesIO
                    output_buffer = io.BytesIO()
                    new_zip = zipfile.ZipFile(output_buffer, 'w', zipfile.ZIP_DEFLATED)

                    # 🔹 3️⃣ 모든 파일 반복                
                    for item in in_memory_zip.infolist():
                        file_data = in_memory_zip.read(item.filename)

                        # 🔹 4️⃣ drawingN.xml인지 체크 (텍스트 박스 포함 가능성 있음)
                        if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
                            try:
                                tree = ET.fromstring(file_data)
                                ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

                                # 모든 <a:t> 태그를 찾고, 텍스트 박스의 위치 정보를 기반으로 정렬
                                text_elements = []
                                for t_element in tree.findall(".//a:t", ns):
                                    original_text = t_element.text
                                    if original_text:
                                        # 텍스트 박스의 위치 정보 추출
                                        parent = t_element.getparent()
                                        if parent is not None:
                                            x = parent.attrib.get('x', 0)
                                            y = parent.attrib.get('y', 0)
                                            text_elements.append((float(y), float(x), original_text.strip()))  # y, x 순서로 저장

                                # 위치 정보를 기준으로 정렬 (상단에서 하단으로, 좌측에서 우측으로)
                                text_elements.sort(key=lambda item: (item[0], item[1]))  # y, x 순으로 정렬

                                # 정렬된 텍스트를 all_text에 추가
                                for _, _, text in text_elements:
                                    all_text.append(text)

                                # 수정된 XML 내용을 다시 직렬화
                                file_data = ET.tostring(tree, encoding='utf-8', standalone=False)

                            except Exception as e:
                                print(f"Warning: Parsing {item.filename} failed - {e}")

                        
                            try:
                                tree = ET.fromstring(file_data)
                                ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                                # 모든 <ss:Row> 태그를 순회하여 각 행의 셀을 읽음
                                for row in tree.findall(".//ss:Row", ns):
                                    for cell in row.findall("ss:Cell", ns):
                                        value_element = cell.find("ss:Data", ns)
                                        if value_element is not None and value_element.text:
                                            all_text.append(value_element.text.strip())

                                        # 합병된 셀 처리
                                        if cell.attrib.get('ss:MergeAcross') is not None:
                                            merged_value = value_element.text.strip() if value_element is not None else ""
                                            for _ in range(int(cell.attrib['ss:MergeAcross'])):
                                                all_text.append(merged_value)  # 합병된 셀의 값을 반복 추가

                            except Exception as e:
                                print(f"Warning: Parsing {item.filename} failed - {e}")
                                
                        new_zip.writestr(item, file_data)

                    # merge all text one string
                    combined_text = ''.join(all_text)
                    
                    # 612 debug
                    # if file_type != "参照ファイル":
                    #     result_map = gpt_correct_text(combined_text)
                    #     corrected_map.update(result_map)  # 결과 맵 병합
                    # else:
                    #     corrected_map = ""


                    # 기존 zip 마무리
                    in_memory_zip.close()
                    # 새 zip 마무리
                    new_zip.close()

                    # 🔹 7️⃣ 수정된 엑셀을 Base64로 인코딩
                    output_buffer.seek(0)
                    # excel_base64 = base64.b64encode(output_buffer.getvalue()).decode('utf-8')
                    excel_base64 = base64.b64encode(file_bytes).decode('utf-8')

                    return jsonify({
                        "success": True,
                        "original_text": combined_text, #corrected_map,
                        "excel_bytes": excel_base64,
                        "combined_text":combined_text,
                        "file_name": file.filename
                    })

            except Exception as e:
                logging.error(f"Error processing file {file.filename}: {str(e)}")
                return jsonify({"success": False, "error": str(e)}), 500

    return jsonify({"success": False, "error": "Invalid file type"}), 400


# 5007 debug
def remove_correction_blocks(html_text):
    # '提示' 또는 '修正理由' 포함된 span 블록 전체를 삭제하는 정규식
    pattern = re.compile(
        r'<span[^>]*?>.*?<\/span>\s*\(<span>提示:<s[^>]*?>.*?<\/s><\/span>\)',
        re.DOTALL
    )
    return pattern.sub('', html_text)

half_to_full_dict = {
    "ｦ": "ヲ", "ｧ": "ァ", "ｨ": "ィ", "ｩ": "ゥ", "ｪ": "ェ", "ｫ": "ォ",
    "ｬ": "ャ", "ｭ": "ュ", "ｮ": "ョ", "ｯ": "ッ", "ｰ": "ー",
    "ｱ": "ア", "ｲ": "イ", "ｳ": "ウ", "ｴ": "エ", "ｵ": "オ",
    "ｶ": "カ", "ｷ": "キ", "ｸ": "ク", "ｹ": "ケ", "ｺ": "コ",
    "ｻ": "サ", "ｼ": "シ", "ｽ": "ス", "ｾ": "セ", "ｿ": "ソ",
    "ﾀ": "タ", "ﾁ": "チ", "ﾂ": "ツ", "ﾃ": "テ", "ﾄ": "ト",
    "ﾅ": "ナ", "ﾆ": "ニ", "ﾇ": "ヌ", "ﾈ": "ネ", "ﾉ": "ノ",
    "ﾊ": "ハ", "ﾋ": "ヒ", "ﾌ": "フ", "ﾍ": "ヘ", "ﾎ": "ホ",
    "ﾏ": "マ", "ﾐ": "ミ", "ﾑ": "ム", "ﾒ": "メ", "ﾓ": "モ",
    "ﾔ": "ヤ", "ﾕ": "ユ", "ﾖ": "ヨ",
    "ﾗ": "ラ", "ﾘ": "リ", "ﾙ": "ル", "ﾚ": "レ", "ﾛ": "ロ",
    "ﾜ": "ワ", "ﾝ": "ン",
    "%": "％", "@": "＠","(":"（",")":"）"
}

full_to_half_dict = {
    '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
    '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
    'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
    'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
    'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
    'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
    'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y', 
    'Ｚ': 'Z','＋':'+','－':'-'
}

king_dict = {
    '中銀': '中央銀行', 
}


def year_half_dict():
    # 전각 문자와 반각 문자의 매핑
    full_half_mapping = {
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
        'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
        'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
        'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
        'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
        'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y',
        'Ｚ': 'Z',
        'ａ': 'a', 'ｂ': 'b', 'ｃ': 'c', 'ｄ': 'd', 'ｅ': 'e',
        'ｆ': 'f', 'ｇ': 'g', 'ｈ': 'h', 'ｉ': 'i', 'ｊ': 'j',
        'ｋ': 'k', 'ｌ': 'l', 'ｍ': 'm', 'ｎ': 'n', 'ｏ': 'o',
        'ｐ': 'p', 'ｑ': 'q', 'ｒ': 'r', 'ｓ': 's', 'ｔ': 't',
        'ｕ': 'u', 'ｖ': 'v', 'ｗ': 'w', 'ｘ': 'x', 'ｙ': 'y',
        'ｚ': 'z',
        '＋': '+', '－': '-', '＝': '=', '（': '(', '）': ')',
        '［': '[', '］': ']', '｛': '{', '｝': '}', '；': ';',
        '：': ':', '‘': "'", '’': "'", '“': '"', '”': '"',
        '、': ',', '。': '.', '・': '・', '！': '!', '？': '?',
        '〜': '~', '￥': '$', '＃': '#', '％': '%', '＠': '@'
    }
    return full_half_mapping


# 半角→,-全角
def half_and_full_process(text, mapping):
    return ''.join(mapping.get(c, c) for c in text)

replace_rules = {
    "AAA": "AAA（全米自動車協会）",
    "ABS": "ABS（資産担保証券、各種資産担保証券）",
    "ADB": "ADB（アジア開発銀行）",
    "ADR": "ADR（米国預託証券）",
    "AI": "AI（人工知能）",
    "AIIB": "AIIB（アジアインフラ投資銀行）",
    "APEC": "APEC（アジア太平洋経済協力会議）",
    "API": "API（全米石油協会）",
    "BIS": "BIS（国際決済銀行）",
    "BOE": "BOE（英中央銀行、イングランド銀行）",
    "BRICS（5ヵ国）": "BRICS（ブラジル、ロシア、インド、中国、南アフリカ）",
    "CDS市場": "CDS（クレジット・デフォルト・スワップ）市場",
    "CFROIC": "CFROIC（投下資本キャッシュフロー利益率）",
    "Chat GPT": "Chat GPT（AIを使った対話型サービス）",
    "CMBS": "CMBS（商業用不動産ローン担保証券）",
    "COP26": "COP26（国連気候変動枠組み条約第26回締約国会議）",
    "CPI": "CPI（消費者物価指数）",
    "CSR": "CSR（企業の社会的責任）",
    "DR": "DR（預託証書）",
    "DRAM": "DRAM（半導体素子を利用した記憶装置のひとつ）",
    "DX": "DX（デジタルトランスフォーメーション）",
    "EC": "EC（電子商取引）",
    "ECB": "ECB（欧州中央銀行）",
    "EIA": "EIA（米エネルギー省エネルギー情報局）",
    "EMEA": "EMEA（欧州・中東・アフリカ）",
    "EPA": "EPA（米環境保護局）",
    "EPS": "EPS（一株当たり利益）",
    "ESM": "ESM（欧州安定メカニズム）",
    "ESG": "ESG（環境・社会・企業統治）",
    "EU": "EU（欧州連合）",
    "EV": "EV（電気自動車）",
    "EVA": "EVA（経済的付加価値）",
    "FASB": "FASB（米財務会計基準審議会）",
    "FDA": "FDA（米国食品医薬品局）",
    "FFレート（米国の場合）": "政策金利（FFレート）",
    "FOMC": "FOMC（米連邦公開市場委員会）",
    "FRB": "FRB（米連邦準備制度理事会）",
    "FTA": "FTA（自由貿易協定）",
    "G7": "G7（主要7ヵ国会議）",
    "G8": "G8（主要8ヵ国首脳会議）",
    "G20": "G20（20ヵ国・地域）財務相・中央銀行総裁会議、首脳会議",
    "GDP": "GDP（国内総生産）",
    "GPIF": "年金積立金管理運用独立行政法人（GPIF）",
    "GNP": "GNP（国民総生産）",
    "GST　※インドの場合": "GST（物品・サービス税）",
    "IEA": "IEA（国際エネルギー機関）",
    "IMF": "IMF（国際通貨基金）",
    "IoT": "IoT（モノのインターネット）",
    "IPEF": "IPEF（インド太平洋経済枠組み）",
    "IPO": "IPO（新規株式公開）",
    # "ISM": "ISM（全米供給管理協会）",
    "ISM非製造業景況": "ISM非製造業景況指数",
    "IT": "IT（情報技術）",
    "LBO": "LBO（レバレッジド・バイアウト：対象企業の資産を担保に資金調達する買収）",
    "LED": "LED（発光ダイオード）",
    "LME": "LME（ロンドン金属取引所）",
    "LNG": "LNG（液化天然ガス）",
    "M&A": "M&A（企業の合併・買収）",
    "MAS": "MAS（シンガポール金融通貨庁）",
    "MBA": "MBA（全米抵当貸付銀行協会）",
    "MBO": "MBO（経営陣による買収）",
    "MBS": "MBS（住宅ローン担保証券）",
    "NAFTA": "NAFTA（北米自由貿易協定）",
    "NAHB": "NAHB（全米住宅建設業者協会）",
    "NAIC": "NAIC（全米保険監督官協会）",
    "NAR": "NAR（全米不動産業者協会）",
    "NDF": "NDF（為替先渡取引のひとつ）",
    "NISA": "NISA（少額投資非課税制度）",
    "OECD": "OECD（経済協力開発機構）",
    "OEM": "OEM（相手先ブランドによる生産）",
    "OPEC": "OPEC（石油輸出国機構）",
    "OPECプラス": "OPECプラス（OPEC（石油輸出国機構）と非加盟産油国で構成するOPECプラス）",
    "PBR": "PBR（株価純資産倍率）",
    "PCE": "PCE（個人消費支出）",
    "PCFR": "PCFR（株価キャッシュフロー倍率）",
    "PER": "PER（株価収益率）",
    "PMI": "PMI（購買担当者景気指数）",
    "PPI": "PPI（生産者物価指数）",
    "QE": "QE（量的金融緩和）",
    "QT": "QT（量的引き締め）",
    "Quad": "Quad（日米豪印戦略対話）",
    "RBA": "RBA（豪州準備銀行）",
    "RCEP": "RCEP（地域的な包括的経済連携協定）",
    "RBI": "RBI（インド準備銀行）",
    "ROA": "ROA（総資産利益率）",
    "ROE": "ROE（自己資本利益率）",
    "S&L": "S&L（貯蓄貸付組合）",
    "SDGs": "SDGs（持続可能な開発目標）",
    "SEC": "SEC（米証券取引委員会）",
    "SQ": "SQ（特別清算指数）",
    "SRI": "SRI（社会的責任投資）",
    "SUV": "SUV（スポーツ用多目的車）",
    "TALF": "TALF（ターム物資産担保証券貸出制度）",
    "TOB": "TOB（株式公開買付け）",
    "TTM": "仲値",
    "TPP": "TPP（環太平洋経済連携協定）",
    "UAE": "UAE（アラブ首長国連邦）",
    "UAW": "UAW（全米自動車労働組合）",
    "USDA": "USDA（米国農務省）",
    "USMCA": "USMCA（米国・メキシコ・カナダ協定）",
    "USTR": "USTR（米通商代表部）",
    "VAT": "VAT（付加価値税）",
    "WTI": "WTI（ウエスト・テキサス・インターミディエート）",
    "WTO": "WTO（世界貿易機関）",
    # "独ZEW景気期待指数（または独ZEW景況感指数）": "ZEW単独使用の場合は括弧書き付き、または欧州経済研究センターのみとしZEWを使わない。",
    # "": "※通貨略称：コメント本文ではUSD、GBPなどの通貨略称は原則使用せず、米ドル、英ポンド等と表記する。",
    "アセットアロケーション": "アセットアロケーション（資産配分）",
    "アンダーウェイト": "アンダーウェイト（ベンチマーク※に比べ低めの投資比率）※参考指数を用いているものはベンチマークの代わりに参考指数を使うこと",
    "E-コマース": "Eコマース（電子商取引）",
    "EC": "EC（電子商取引）",
    "e-コマース": "eコマース（電子商取引）",
    "イールドカーブ": "イールドカーブ（利回り曲線）",        # ※欄外に注記する場合イールドカーブ（利回り曲線）とは、横軸に残存年数、縦軸に利回りをとった座標に、債券利回りを点描して結んだ曲線のことを指します。
    "イールドカーブ・コントロール": "イールドカーブ・コントロール（長短金利操作）",
    "イールドカーブのスティープ化": "イールドカーブのスティープ化（長・短金利格差の拡大）",
    "イールドカーブのフラット化": "イールドカーブのフラット化（長・短金利格差の縮小）",
    "インカムゲイン": "インカムゲイン（利子収入）",
    "インタラクティブ": "インタラクティブ（双方向性）",
    # "インバウンド": "インバウンド（訪日外国人）　＊～消費として使用",
    "エクイティ・ファイナンス": "エクイティ・ファイナンス（新株発行等による資金調達）",
    # "エクスポージャー": "＊積極的に使用しない。　（価格変動リスク資産の配分比率、割合）",
    # "円キャリートレード": "円キャリートレード（円などの低金利の通貨を売り、高金利の通貨を買う取引）　　〃　　　　（低金利の円で資金を調達して、高金利の通貨で運用する取引）",
    "オバマケア": "オバマケア（医療保険制度改革法）",
    # "オーバーウェイト": "オーバーウェイト（ベンチマーク※に比べ高めの投資比率）※参考指数を用いているものはベンチマークの代わりに参考指数を使うこと",
    "オンデマンド": "オンデマンド（注文生産）",
    "カントリー･アロケーション": "カントリー･アロケーション（国別資産配分）",
    "逆イールド": "逆イールド（短期債券の利回りが長期債券の利回りを上回っている状態）",
    "キャッシュフロー": "キャッシュフロー（現金収支）",
    "キャピタルゲイン": "キャピタルゲイン（値上がり益）",
    "キャリートレード": "キャリートレード（低金利の資金を調達して、高金利の資産で運用する取引）",
    # "クレジット（信用）市場": "信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称。　企業の信用リスクを取引する市場。",
    "クレジットスプレッド": "クレジットスプレッド（企業の信用力の差による利回りの差）",
    "グローバリゼーション": "グローバリゼーション（地球規模化）",
    "コジェネレーション": "コジェネレーション（熱電供給システム）",
    "コーポレート・ガバナンス": "コーポレート・ガバナンス（企業統治）",
    "コングロマリット": "コングロマリット（複合企業）",
    "コンソーシアム": "コンソーシアム（共同事業）",
    "サーベイランス": "サーベイランス（調査監視）",
    "サステナビリティ": "サステナビリティ（持続可能性）",
    "サブプライムローン": "サブプライムローン（信用度の低い個人向け住宅融資）",
    "サプライチェーン": "サプライチェーン（供給網）",
    "ジェネリック医薬品": "ジェネリック医薬品（後発薬）",
    "シャリア": "シャリーア",
    "シクリカル": "シクリカル（景気敏感）",
    # "システミック・リスク": "個別の金融機関の支払不能等や、特定の市場または決済システム等の機能不全が、他の金融機関、他の市場、または金融システム全体に波及するリスク",
    "シャドーバンキング": "シャドーバンキング（影の銀行）",
    "ショートポジション": "ショートポジション（売り持ち）",
    "信用市場": "企業の信用リスクを取引する市場",
    # "スタグフレーション": "（景気後退下のインフレ）、（物価上昇と景気悪化が同時に進むこと）*スタグネーション（景気後退）とインフレーション（物価上昇）の合成語",  aerzhsdtgtg
    "スティープ化": "スティープ化（長短金利格差の拡大）",
    "ストレステスト": "ストレステスト（健全性審査）",
    "スプレッド": "スプレッド（利回り格差）",
    "スマートシティー": "スマートシティー（ITを活用した次世代型の都市）",
    "スマートモビリティ": "スマートモビリティ（従来の交通・移動を変える新たなテクノロジー）",
    "セーフティネット": "セーフティネット（安全網）",
    "全人代": "全人代（全国人民代表大会）",
    # "センチメント": "センチメント（雰囲気、市場に対する見方、市場心理）",
    "ソフトランディング": "ソフトランディング（軟着陸）",
    # "ソブリンリスク": "国家のリスク",
    "ダイバーシティ": "ダイバーシティ（多様性）",
    "ターミナルレート": "ターミナルレート（政策金利の最終到達水準）",
    "テーパリング": "テーパリング（量的金融緩和の縮小）",
    # "ディストレス債券": "※欄外に注記信用事由などにより、価格が著しく下落した債券",    aerzhsdtgtg
    "ディフェンシブ": "ディフェンシブ（景気に左右されにくい）",
    # "テクニカル": "テクニカル（過去の株価の動きから判断すること）",
    "デフォルト": "デフォルト（債務不履行）",
    # "デフォルト債": "※欄外に注記デフォルトとは一般的には債券の利払いおよび元本返済の不履行、もしくは遅延などをいい、このような状態にある債券をデフォルト債といいます。",  aerzhsdtgtg
    "デュレーション": "デュレーション（金利感応度）",        # ※欄外に注記デュレーションとは、金利がある一定の割合で変動した場合、債券の価格がどの程度変化するかを示す指標です。すなわち、この値が大きいほど金利変動に対する債券価格の変動率が大きくなります。（目論見書の表記、他パターン有）
    "デリバティブ": "デリバティブ（金融派生商品）",
    # "投資適格債": "※欄外に注記格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債",   aerzhsdtgtg
    "ドルペッグ制": "ドルペッグ（連動）制",
    # "ニュートラル": "ニュートラル（ベンチマーク※並みの投資比率）",
    "バイオシミラー": "バイオシミラー（後続薬）",
    "バイオマス": "バイオマス（生物を利用して物質やエネルギーを得ること）",
    "バーチャル": "バーチャル（仮想）",
    "バリュエーション": "バリュエーション（投資価値評価）",
    "バリュー": "バリュー（割安）",
    "5G": "5G（第5世代移動通信システム）",
    # "ファンダメンタルズ": "ファンダメンタルズ（基礎的条件）",    # ファンダメンタルズ（賃料や空室率、需給関係などの基礎的条件）※REITファンドで使用する
    "フィンテック": "フィンテック（金融と技術の融合）",       # ＊「Fｉｎａｎｃｅ 」+「Tｅｃｈｎｏｌｏｇｙ」を組み合わせた造語
    "フェアバリュー": "フェアバリュー（適正価格）",
    "フェーズ2": "フェーズ2（臨床試験の中間段階）",
    "フェーズ3": "フェーズ3（臨床試験の最終段階）",
    "フードデリバリー": "フードデリバリー（料理等の宅配サービス）",
    "フルインベストメント": "フルインベストメント（高位組入）",
    "ブロードバンド": "ブロードバンド（大容量･高速通信）",
    "ポテンシャル": "ポテンシャル（潜在力）",
    "ポピュリズム": "ポピュリズム（大衆迎合主義）",
    # "ボラティリティー": "ボラティリティー（価格変動性）",
    "モーゲージ": "モーゲージ（不動産担保ローン）",
    "モーゲージ債": "モーゲージ債（不動産ローン担保債券）",
    "モラルハザード": "モラルハザード（倫理崩壊）",
    "リエンジニアリング": "リエンジニアリング（業務の抜本的革新）",
    "リオープン": "リオープニング（経済活動再開）",
    "リセッション": "リセッション（景気後退）",
    "リターンリバーサル": "リターンリバーサル（過剰反応効果）",
    "リバウンド": "リバウンド（反発）",
    "リバランス": "リバランス（投資比率の再調整）",
    "レパトリ減税": "レパトリ（海外収益の本国還流）減税",
    "レバレッジドローン": "低格付け等の借り手向け融資",
    "レラティブ・バリュー": "レラティブ・バリュー（相対価値）",
    "ロックダウン": "ロックダウン（都市封鎖）",
    "ロングポジション": "ロングポジション（買い持ち）",
    "殆ど": "ほとんど",
    "真似": "まね",
    "亘る": "わたる",
    "但し": "ただし",
    "牽制": "けん制",
    "牽引": "けん引",
    "終焉": "終えん",
    "収斂": "収れん",
    "逼迫": "ひっ迫",
    "横這い": "横ばい",
    "ヶ月": "ヵ月",
    "入替え": "入れ替え",
    "入替": "入れ替",
    "行われる": "行なわれる",
    "売付":"売り付け",
    "売付け":"売り付け",
    "格付": "格付け", 
    "買建て": "買い建て",
    "売建て": "売り建て",
    "切上げ": "切り上げ",
    "切捨て": "切り捨て",
    "組入れ": "組み入れ", 
    "繰上げ償還": "繰上償還",
    "先き行き": "先行き",
    "下支える": "下支えする",
    "取り引き": "取引",
    "引上げ": "引き上げ",
    "引下げ": "引き下げ",
    "引続き": "引き続き",
    "引締め": "引き締め",
    "薄商い": "取引量が少なく",
    "コア銘柄": "中核銘柄、コア（中核）銘柄",
    "トリガー": "きっかけ",
    "ブルーチップ企業": "優良企業",
    # "来い": "下落（上昇）",
    "まちまち": "異なる動き",
    "ハト派／タカ派": "金融緩和に前向き／金融引き締め重視",
    # "織り込む": "反映され",
    "相場": "市場",
    "連れ高": "影響を受けて上昇",
    "伝播": "広がる",
    "でんぱ": "広がる",
    "トレンド": "傾向",
    "レンジ": "範囲",
    "回金": "円転",
    "ローン": "貸し付け",  # "住宅ローン" iudrhasi
    "所謂": "いわゆる",
    "暫く": "しばらく",
    "留まる": "とどまる",
    "止まる": "とどまる",
    "尚": "なお",
    "筈": "はず",
    "蓋然性": "可能性",
    "商い": "出来高",
    "後倒し": "延期",
    "経済正常化": "経済活動正常化",
    "金融正常化": "金融政策正常化",
    "日本銀行": "日銀",
    "政治的リスク": "政治リスク",
    "地政学リスク": "地政学的リスク",
    "への組み入れ": "の組み入れ",
    "利回りは上昇": "利回りは上昇（価格は下落）",
    "利回りは低下": "利回りは低下（価格は上昇）",
    "マイナスに寄与": "マイナスに影響",
    "米国国債": "米国債",
    "新型コロナ": "新型コロナウイルス",
    "コロナウイルス": "新型コロナウイルス",
    "立ち後れ": "立ち遅れ",
    "伸張": "伸長",
    # "資金流出入": "外国人投資家の資金流出⇔外国人投資家からの資金流入",
    "ダウ平均": "ダウ平均株価",
    "NYダウ": "ダウ平均株価",
    "買い付けました": "買い付けしました",
    '中銀': '中央銀行', #623
    '行われて': '行なわれて', #623
    '行い': '行な', #623
    '行う': '行なう', #623
    '行って': '行なって', #623
    '買い付け': '買い付けし', #64977
    'EDAツール': 'EDAツール（電子設計自動化ツール）', #623
    '前年比': '前年同期比', #623
    "買付":"買い付け",
    "買付け":"買い付けし", #63207
    "TOPIX":"TOPIX（東証株価指数）", #63207
}

def merge_brackets(content: str) -> str:
    """
    괄호 내부의 줄바꿈을 제거합니다. 예: 'CPI（消費者物\n価指数）' -> 'CPI（消費者物価指数）'
    """
    return re.sub(r'（[^）\n\r]*[\n\r]+[^）]*）', lambda m: m.group(0).replace("\n", "").replace("\r", ""), content)

def insert_year_by_regex(text: str) -> str:
    """
    '月日' 날짜가 나오고, 앞에 연도가 없으면, 가장 가까운 앞쪽의 'XXXX年'을 삽입한다.
    괄호 여부에 관계없이 처리.
    """

    # 날짜 표현 중 연도 없는 경우만 추출
    pattern = r'(?<!\d{4}年)(\d{1,2}月\d{1,2}日)'

    def replacer(match):
        date_pos = match.start()

        # date 앞에서 가장 가까운 연도 추출
        year_matches = list(regcheck.finditer(r'(\d{4})年', text[:date_pos]))
        if year_matches:
            last_year = year_matches[-1].group(1)
            return f'{last_year}年{match.group(1)}'
        else:
            return match.group(1)  # 연도가 없으면 그대로 반환

    return re.sub(pattern, replacer, text)


def get_Years_regex(text: str) -> str:
    mapping = year_half_dict()  # 딕셔너리 호출
    return ''.join(mapping.get(c, c) for c in text)

def opt_check_eng(content, rules):
    content = merge_brackets(content)  # 1️⃣ 괄호 내 줄바꿈 제거
    result = []
    content = content.replace("(", "（").replace(")", "）")
    for k, v in rules.items():
        raw_key = k.replace("(", "（").replace(")", "）")
        full_key = v.replace("(", "（").replace(")", "）")

        escaped_k = regcheck.escape(raw_key)
        escaped_v = regcheck.escape(full_key)

        new_k = escaped_k
        if raw_key.isalpha() or raw_key in ["S&L", "M&A"]:
            if raw_key == "OPEC":
                new_k = f"(?<![a-zA-Z]){escaped_k}(?!プラス|[a-zA-Z])"
            elif raw_key == "スティープ化":
                new_k = f"(?<!イールドカーブの){escaped_k}"
            elif raw_key == "イールドカーブ":
                new_k = f"{escaped_k}(?!・コントロール|のスティープ化|のフラット化)"
            elif raw_key == "キャッシュフロー":
                new_k = f"(?<!フリー){escaped_k}"
            elif raw_key == "キャリートレード":
                new_k = f"(?<!円){escaped_k}"
            elif raw_key == "スプレッド":
                new_k = f"(?<!クレジット){escaped_k}"
            elif raw_key == "バリュー":
                new_k = f"(?<!レラティブ・|フェア){escaped_k}"
            elif raw_key == "モーゲージ":
                new_k = f"{escaped_k}(?!債)"
            elif raw_key == "商い":
                new_k = f"(?<!薄){escaped_k}"
            else:
                new_k = f"(?<![a-zA-Z]){escaped_k}(?![a-zA-Z])"
        # 예외 처리: 中銀
        elif raw_key == "中銀":
            # '中央銀行' 앞에 수식어 포함 여부 확인
            matches = re.finditer(escaped_v, content)
            exclude = False
            for m in matches:
                # 예: '欧州中央銀行' => m.start() - 2 >= 0, 앞 2글자 포함 확인
                prefix = content[max(0, m.start() - 2): m.start()]
                if prefix and not re.match(r"[ \t\n\r]", prefix):
                    exclude = True
                    break
            if exclude:
                new_k = escaped_k  # full_key는 건너뛰고, raw_key만 검사
                full_match = None
            else:
                full_match = re.search(escaped_v, content)
                

        raw_match = regcheck.search(new_k, content)
        full_match = regcheck.search(escaped_v, content)

        # 일반 조건: full_key가 먼저 등장하면 제외
        if raw_key != "中銀":  # 중銀 예외 상황 제외
            if full_match and raw_match and full_match.start() <= raw_match.start():
                continue
            elif full_match and not raw_match:
                continue

        if raw_match:
            result.append({raw_key: full_key})
        
    return result

# 0501 debug
def find_corrections(corrected_text,input_text,pageNumber):
    """
    GPT-4의 응답 결과(corrected_text)를 분석하여 틀린 부분을 찾습니다.
    """
    corrections = []
    # 정규 표현식을 사용하여 틀린 부분과 수정 이유 또는提示을 추출
    pattern = r'<span\s+style="color:red;">([\s\S]*?)<\/span>\s*\(<span>\s*修正理由[::]\s*([\s\S]*?)\s*<s[^>]*>([\s\S]*?)<\/s>\s*→\s*([\s\S]*?)<\/span>\)'
    matches = re.findall(pattern, corrected_text)

    # 디버깅: matches 출력
    print("Matches found:", matches)
    # <span style="color:red;">上午12时00分</span> (<span>修正理由: 不要な中国語表記 <s style="background:yellow;color:red">上午12时00分</s> → （削除）</span>)

    for match in matches:
        if len(match) == 4:
            corrected_text_re = match[0]  #610 debug
            reason_type = match[1].strip()
            original_text = match[2].strip()
            target_text = match[3].strip()

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": corrected_text_re,  # 전체 입력
                "comment": comment,
                "reason_type":reason_type,
                "check_point": input_text.strip(),
                "locations": [],
                "intgr": False, # for debug 62
            })
    
    return corrections

# 0623 debug
def find_corrections_wording(input_text,pageNumber,tenbrend,fund_type):
    """
    GPT-4의 응답 결과(corrected_text)를 분석하여 틀린 부분을 찾습니다.
    """
    corrections = []

#-------------------
    #常用外汉字
    common_par = r"((啞|蛙|鴉|埃|挨|曖|靄|軋|斡|按|庵|鞍|闇|已|夷|畏|韋|帷|萎|椅|葦|彙|飴|謂|閾|溢|鰯|尹|咽|殷|淫|隕|蔭|于|迂|盂|烏|鬱|云|暈|穢|曳|洩|裔|穎|嬰|翳|腋|曰|奄|宛|怨|俺|冤|袁|婉|焉|堰|淵|焰|筵|厭|鳶|燕|閻|嚥|嗚|凰|嘔|鴨|甕|襖|謳|鶯|鷗|鸚|臆|俤|瓜|呵|苛|珂|迦|訛|訶|跏|嘩|瑕|榎|窩|蝦|蝸|鍋|顆|牙|瓦|臥|俄|峨|訝|蛾|衙|駕|芥|乖|廻|徊|恢|晦|堺|潰|鞋|諧|檜|蟹|咳|崖|蓋|漑|骸|鎧|喀|廓|摑|攪|愕|萼|諤|顎|鰐|樫|絣|筈|葛|闊|鰹|萱|奸|串|旱|函|咸|姦|宦|柑|竿|悍|桓|涵|菅|嵌|鉗|澗|翰|諫|瞰|韓|檻|灌|玩|雁|翫|頷|癌|贋|几|卉|其|祁|耆|埼|悸|揆|毀|箕|畿|窺|諱|徽|櫃|妓|祇|魏|蟻|掬|麴|吃|屹|拮|謔|仇|臼|汲|灸|咎|邱|柩|笈|躬|厩|嗅|舅|炬|渠|裾|噓|墟|鋸|遽|欅|匈|怯|俠|脇|莢|竟|卿|僑|嬌|蕎|鋏|頰|橿|疆|饗|棘|髷|巾|僅|禽|饉|狗|惧|軀|懼|俱|喰|寓|窟|粂|偈|荊|珪|畦|脛|頃|痙|詣|禊|閨|稽|頸|髻|蹊|鮭|繫|睨|戟|隙|抉|頁|訣|蕨|姸|倦|虔|捲|牽|喧|硯|腱|鍵|瞼|鹼|呟|眩|舷|諺|乎|姑|狐|股|涸|菰|袴|壺|跨|糊|醐|齬|亢|勾|叩|尻|吼|肛|岡|庚|杭|肴|咬|垢|巷|恍|恰|狡|桁|胱|崗|梗|喉|腔|蛤|幌|煌|鉤|敲|睾|膏|閤|膠|篝|縞|薨|糠|藁|鮫|壙|曠|劫|毫|傲|壕|濠|嚙|轟|剋|哭|鵠|乞|忽|惚|昏|痕|渾|褌|叉|些|嗟|蓑|磋|坐|挫|晒|柴|砦|犀|賽|鰓|榊|柵|炸|窄|簀|刹|拶|紮|撒|薩|珊|餐|纂|霰|攢|讃|斬|懺|仔|弛|此|址|祀|屍|屎|柿|茨|恣|砥|祠|翅|舐|疵|趾|斯|覗|嗜|滓|獅|幟|摯|嘴|熾|髭|贄|而|峙|痔|餌|竺|雫|𠮟|悉|蛭|嫉|膝|櫛|柘|洒|娑|這|奢|闍|杓|灼|綽|錫|雀|惹|娶|腫|諏|鬚|呪|竪|綬|聚|濡|襦|帚|酋|袖|羞|葺|蒐|箒|皺|輯|鍬|繡|蹴|讐|鷲|廿|揉|絨|粥|戌|閏|楯|馴|杵|薯|藷|汝|抒|鋤|妾|哨|秤|娼|逍|廂|椒|湘|竦|鈔|睫|蛸|鉦|摺|蔣|裳|誦|漿|蕭|踵|鞘|篠|聳|鍾|醬|囁|杖|茸|嘗|擾|攘|饒|拭|埴|蜀|蝕|燭|褥|沁|芯|呻|宸|疹|蜃|滲|賑|鍼|壬|訊|腎|靱|塵|儘|笥|祟|膵|誰|錐|雖|隋|隧|芻|趨|鮨|丼|凄|栖|棲|甥|貰|蜻|醒|錆|臍|瀞|鯖|脆弱?|贅|脊|戚|晰|蹟|泄|屑|浙|啜|楔|截|尖|苫|穿|閃|陝|釧|揃|煎|羨|腺|詮|煽|箋|撰|箭|賤|蟬|癬|喘|膳|狙|疽|疏|甦|楚|鼠|遡|蘇|齟|爪|宋|炒|叟|蚤|曾|湊|葱|搔|槍|漕|箏|噌|瘡|瘦|踪|艘|薔|甑|叢|藪|躁|囃|竈|鰺|仄|捉|塞|粟|杣|遜|噂|樽|鱒|侘|咤|詫|陀|拿|荼|唾|舵|楕|驒|苔|殆|堆|碓|腿|頽|戴|醍|托|鐸|凧|襷|燵|坦|疸|耽|啖|蛋|毯|湛|痰|綻|憚|歎|簞|譚|灘|雉|馳|蜘|緻|筑|膣|肘|冑|紐|酎|厨|蛛|註|誅|疇|躊|佇|楮|箸|儲|瀦|躇|吊|帖|喋|貼|牒|趙|銚|嘲|諜|寵|捗|枕|槌|鎚|辻|剃|挺|釘|掟|梯|逞|啼|碇|鼎|綴|鄭|薙|諦|蹄|鵜|荻|擢|溺|姪|轍|辿|唸|塡|篆|顚|囀|纏|佃|淀|澱|臀|兎|妬|兜|堵|屠|賭|宕|沓|套|疼|桶|淘|萄|逗|棹|樋|蕩|鄧|橙|濤|檮|櫂|禱|撞|禿|瀆|栃|咄|沌|遁|頓|吞|貪|邇|匂|韮|涅|禰|捏|捻|撚|膿|囊|杷|爬|琶|頗|播|芭|罵|蟇|胚|徘|牌|稗|狽|煤|帛|柏|剝|粕|箔|莫|駁|瀑|曝|畠|捌|撥|潑|醱|筏|跋|噺|氾|汎|阪|叛|袢|絆|斑|槃|幡|攀|挽|磐|蕃|屁|庇|砒|脾|痺|鄙|誹|臂|枇|毘|梶|媚|琵|薇|靡|疋|畢|逼|謬|豹|憑|瓢|屛|廟|牝|瀕|憫|鬢|斧|阜|訃|俯|釜|腑|孵|鮒|巫|葡|撫|蕪|諷|祓|吻|扮|焚|糞|幷|聘|蔽|餅|斃|袂|僻|璧|襞|蔑|瞥|扁|篇|騙|娩|鞭|哺|圃|蒲|戊|牡|姥|菩|呆|彷|庖|苞|疱|捧|逢|蜂|蓬|鞄|鋒|牟|芒|茫|虻|榜|膀|貌|鉾|謗|吠|卜|勃|梵|昧|邁|枡|俣|沫|迄|曼|蔓|瞞|饅|鬘|鰻|蜜|鵡|冥|瞑|謎|麵|蒙|朦|勿|籾|悶|揶|爺|鑓|喩|揄|愈|楡|尤|釉|楢|猷|飫|輿|孕|妖|拗|涌|痒|傭|熔|瘍|蠅|沃|螺|萊|蕾|洛|埒|拉|辣|瀾|爛|鸞|狸|裡|罹|籬|戮|慄|掠|笠|溜|榴|劉|瘤|侶|梁|聊|菱|寥|蓼|淋|燐|鱗|屢|蛉|蠣|櫟|礫|轢|煉|漣|憐|簾|鰊|攣|賂|魯|濾|廬|櫓|蘆|鷺|弄|牢|狼|榔|瘻|﨟|臘|朧|蠟|籠|聾|肋|勒|漉|麓|窪|歪|猥|隈|或|罠|椀|碗|彎|一旦).{,5})"
    common_list = regcheck.findall(common_par, input_text)
    
    for word in common_list:
        reason_type = "常用外漢字の使用"
        corrections.append({
            "page": pageNumber,
            "original_text": word[0],  # original_text,
            "comment": word[0],
            "reason_type": reason_type,
            "check_point": word[0],  # 필요에 따라 입력
            "locations": [],  # 필요에 따라 입력
            "intgr": False,  # for debug 62
        })
#-------------------
    if fund_type == 'public':
        # （全角→半角） -0.09% → -0.09％
        pattern_half_width_katakana = r"[ｦ-ﾝ%＠()]+"
        half_width_katakana_matches = regcheck.findall(pattern_half_width_katakana, input_text)

        for match in half_width_katakana_matches:
            corrected_text_re = half_and_full_process(match,half_to_full_dict)  # 반각 카타카나를 전각으로 변환
            reason_type = "半角カタカナを全角カタカナに統一"  # 수정 이유
            original_text = match  # 원본 텍스트
            target_text = corrected_text_re  # 전각으로 변환된 텍스트
            # 「％」表記の統一（全角→半角） -0.09% → -0.09％
            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,#corrected_text_re
                "comment": comment,
                "reason_type": reason_type,
                "check_point": input_text.strip(),
                "locations": [],
                "intgr": False, # for debug 62
            })

        # 半角→全角
        pattern_full_width_numbers_and_letters = r"[０-９Ａ-Ｚ＋－]+"
        full_width_matches = regcheck.findall(pattern_full_width_numbers_and_letters, input_text)

        for match in full_width_matches:
            corrected_text_re = half_and_full_process(match,full_to_half_dict)  # 전각 숫자 및 알파벳을 반각으로 변환
            reason_type = "全角数字・アルファベットを半角数字・アルファベットに統一"  # 수정 이유
            original_text = match  # 원본 텍스트
            target_text = corrected_text_re  # 반각으로 변환된 텍스트

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": input_text.strip(),
                "locations": [],
                "intgr": False, # for debug 62
            })
#-------------------
    # 年度
    if fund_type == 'public':
        cleaned_text = re.sub(r'\n\s*', '', input_text)
        current_year = insert_year_by_regex(cleaned_text)
        date_pattern = r'(?<!\d年)(\d{1,2}月\d{1,2})|(\d{4}年\d{1,2}月\d{1,2}日)'
        date_matches = regcheck.findall(date_pattern, cleaned_text)

        for match in date_matches:
            date_str = match[0] if match[0] else match[1]
            
            if match[1]:  # 연도가 있는 경우
                continue

            # 연도가 없는 경우 현재 연도를 추가
            date_with_year = f"{current_year}"

            pattern = f".{{0,8}}{date_pattern}"
            _original_re = regcheck.search(pattern, cleaned_text, re.DOTALL)


            if _original_re:
                _original_text = _original_re.group()
            else:
                _original_text = date_str

            # 전각 숫자 및 알파벳을 반각으로 변환
            corrected_text_re = get_Years_regex(date_with_year)  # 전각 숫자 및 알파벳을 반각으로 변환
            reason_type = "表記の統一"  # 수정 이유
            original_text = _original_text  # 원본 텍스트
            target_text = corrected_text_re  # 반각으로 변환된 텍스트

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": input_text.strip(),
                "locations": [],  # 위치 정보는 필요에 따라 추가
                "intgr": False,  # for debug
            })
#-------------------
    # 英略词
    if fund_type == 'public':
        results = opt_check_eng(input_text, replace_rules)
        for item in results:
            for k, v in item.items():
                original_text = k  # 키 값을 original_text에 저장 AI
                corrected_text_re = v  # 값(v)을 corrected_text_re에 저장 AI（人工知能）
                reason_type = "用語の統一"  # 수정 이유

                comment = f"{reason_type} {original_text} → {corrected_text_re}"

            corrections.append({
                "page": pageNumber,
                "original_text": extract_text(input_text, original_text),# original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": input_text.strip(),  # 필요에 따라 입력
                "locations": [],  # 필요에 따라 입력
                "intgr": False,  # for debug 62
            })
#-------------------
    # tenbrend
    for item in tenbrend:
        old_text = item.get("元組入銘柄解説", "").strip()
        new_text = item.get("新組入銘柄解説", "").strip()

        corrections.append({
            "check_point": "組入銘柄解説",
            "comment": f"{old_text} → {new_text}",
            "intgr": False,
            "locations": [],
            "original_text": new_text,
            "page": pageNumber,
            "reason_type": item.get("stocks", "")
        })

    return corrections

def extract_text(input_text, original_text):
    # 정규 표현식 패턴: target 뒤에 （가 있는 경우와 없는 경우를 처리
    pattern = rf"{original_text}（[^）]*）|{original_text}"  # target 뒤에 （가 있고, 그 뒤에 어떤 문자(닫는 괄호 제외)가 올 수 있으며, 마지막에 ）가 오는 경우를 매칭

    # 정규 표현식으로 매칭
    match = regcheck.search(pattern, input_text)
    
    if match:
        return match.group(0)  # 매칭된 텍스트 반환
    else:
        return None  # 매칭되지 않는 경우 None 반환
                
def extract_corrections(corrected_text, input_text):
    corrections = []
    
    # 여러 개의 correction span을 처리하는 정규식
    pattern_alt = re.compile(
        r'<span.*?>(.*?)<\/span>\s*'
        r'\(<span>提示:\s*(.*?)\s*<s.*?>(.*?)<\/s>\s*→\s*(.*?)<\/span>\)',
        re.DOTALL
    )

    matches = pattern_alt.findall(corrected_text)

    for match in matches:
        original = match[0].strip()
        reason = match[1].strip()
        reason_type = match[2].strip()
        corrected = match[3].strip()

        comment = f"{reason} → {corrected}" if corrected else reason

        corrections.append({
            "page": 0,
            "original_text": input_text.strip(),  # 전체 입력값
            "comment": comment,
            "reason_type": reason_type,
            "check_point": input_text.strip(),
            "locations": [],
            "intgr": True,
        })

    return corrections

    
def add_comments_to_pdf(pdf_bytes, corrections):
    """
    PDF 파일에서 틀린 부분을 찾아 코멘트를 추가합니다.

    :param pdf_bytes: PDF 파일의 바이트 데이터
    :param corrections: 수정 사항 리스트 (각 항목은 page, original_text, comment를 포함)
    :return: 수정된 PDF 파일의 BytesIO 객체
    """
    # 입력 유효성 검사
    if not isinstance(pdf_bytes, bytes):
        raise ValueError("pdf_bytes must be a bytes object.")
    if not isinstance(corrections, list):
        raise ValueError("corrections must be a list of dictionaries.")
    for correction in corrections:
        if not all(key in correction for key in ["page", "original_text", "comment"]):
            raise ValueError("Each correction must contain 'page', 'original_text', and 'comment' keys.")

    try:
        # PDF 파일 열기 (BytesIO에서 직접 열기)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        raise ValueError(f"Invalid PDF file: {str(e)}")

    for  idx,correction in enumerate(corrections):
        page_num = correction["page"]
        original_text = correction["original_text"]
        comment = correction["comment"]

        # 페이지 번호 유효성 검사
        if page_num < 0 or page_num >= len(doc):
            raise ValueError(f"Invalid page number: {page_num}")

        page = doc[page_num]
        text_instances = page.search_for(original_text)

        # 틀린 위치(좌표) 저장을 위해, 일단 빈 리스트로 초기화
        found_locations = []

        
        for inst in text_instances:
            rect = fitz.Rect(inst)  # 틀린 부분의 위치
            # annot = page.add_text_annot(rect.top_left, comment)  # 코멘트 추가
            annot = page.add_highlight_annot(rect)  # 텍스트 하이라이트 추가
            # 코멘트 색상 및 투명도 설정
            annot.set_colors(stroke=(0.5, 0.5, 0.5))  # 테두리 색상 (빨간색)
            annot.set_colors(fill=(1, 0, 0.8))  # 배경색 (연한 노란색)
            annot.set_opacity(0.5)  # 투명도 설정 (0: 완전 투명, 1: 완전 불투명),50%

            # 코멘트 폰트 크기 조정 (옵션)
            annot.set_info(title="コメントチェック", content=comment)

            annot.update()  # 주석 업데이트

            # === 틀린 위치(좌표) 정보 저장 ===
            found_locations.append({
                "x0": rect.x0,
                "y0": rect.y0,
                "x1": rect.x1,
                "y1": rect.y1
            })

        # corrections 내 해당 항목에 위치 정보 저장
        corrections[idx]["locations"] = found_locations

    # 수정된 PDF를 BytesIO에 저장
    output = io.BytesIO()
    doc.save(output)
    output.seek(0)
    doc.close()

    return output

def add_comments_to_excel(excel_bytes, corrections):
    """
    엑셀 파일에서 틀린 부분을 찾아 코멘트를 추가합니다.

    :param excel_bytes: 엑셀 파일의 바이트 데이터
    :param corrections: 수정 사항 리스트 (각 항목은 sheet, cell, original_text, comment를 포함)
    :return: 수정된 엑셀 파일의 BytesIO 객체
    """
    # 엑셀 파일 열기
    excel_file = io.BytesIO(excel_bytes)
    workbook = load_workbook(excel_file)  # openpyxl로 엑셀 파일 로드

    # 각 시트에 대해 처리
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for correction in corrections:
            if correction["sheet"] == sheet_name:
                cell = correction["cell"]  # 예: "A1", "B2"
                original_text = correction["original_text"]
                comment = correction["comment"]

                # 셀 값 확인
                if sheet[cell].value and original_text in str(sheet[cell].value):
                    # 빨간색으로 하이라이트
                    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    sheet[cell].fill = fill

                    # 코멘트 추가
                    sheet[cell].comment = Comment(comment, "Author")

    # 수정된 엑셀 파일 저장
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# pre processing
def normalize_text_for_search(text: str) -> str:
    import re
    replacements = {
        "（": "(", "）": ")", "【": "[", "】": "]",
        "「": "\"", "」": "\"", "『": "\"", "』": "\"",
        "　": " ", "○": "〇", "・": "･", 
        "–": "-", "―": "-", "−": "-", "ー": "-"
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[\u200b\u200c\u200d\u00a0]", "", text)
    return re.sub(r"\s+", " ", text).strip()


#0617 debug
def find_locations_in_pdf(pdf_bytes, corrections):
    """
    PDF에서 'original_text'를 검색하여 위치 정보를 corrections에 추가한다.
    주석(하이라이트)은 생성하지 않고, 단순히 위치 정보만 찾아서 저장.
    """
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        raise ValueError(f"Invalid PDF file: {str(e)}")

    for idx, correction in enumerate(corrections):
        page_num = correction.get("page", 0)
        original_text = correction["original_text"]

        if page_num < 0 or page_num >= len(doc):
            print(f"Warning: Invalid page number: {page_num}")
            continue

        page = doc[page_num]
        found_locations = []
        # pre processing
        text_instances = page.search_for(original_text)

        if not text_instances:
            print(f"Warning: Text '{original_text}' not found on page {page_num}.")
            found_locations.append({
                "x0": 0,
                "y0": 0,
                "x1": 0,
                "y1": 0
            })
            
        else:
            for inst in text_instances:
                rect = fitz.Rect(inst)
                found_locations.append({
                    "x0": rect.x0,
                    "y0": rect.y0,
                    "x1": rect.x1,
                    "y1": rect.y1
                })

        # corrections[idx]에 locations 필드가 없으면 추가
        if "locations" not in corrections[idx]:
            corrections[idx]["locations"] = []
        corrections[idx]["locations"].extend(found_locations)

    doc.close()
    return corrections


@app.route('/api/write_upload', methods=['POST'])
def write_upload():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")

        data = request.json
        prompt = data.get("input", "")
        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        if not prompt:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        # db get map
        corrected_map = fetch_and_convert_to_dict()

#424 -update new pmpt
        prompt_result = f"""
        You are a professional Japanese text proofreading assistant. 
        Please carefully proofread the content of a Japanese report following the rules below. 
        This includes not only Japanese text but also English abbreviations (英略語), foreign terms (外来語),
        and specialized terminology (専門用語). Ensure that all language elements are reviewed according to the guidelines and corrected where necessary.:

        **Report Content to Proofread:**
        {prompt}

        **Proofreading Requirements:**
        1. **Check for typos and missing characters (誤字脱字がないこと):**
        - Ensure there are no **spelling errors** or **missing characters** in the report. 
        - あなたの役割は、日本語の誤字・脱字・表記ミスを修正し、不完全な単語や文章に適切な語を補完することです。  
        以下のルールに従い、入力されたテキストを校正してください。
            **Common Mistakes Examples (誤字・脱字の例)**:
            Example:
            - `リテール投家` → `リテール投資家` (誤字: 家 → 資)
            - `長国債` → `長期国債` (脱字: 期を追加)
            - `識された` → `意識された` (表記の統一)
            - `金緩和期待` → `金融緩和期待` (誤字: 金 → 金融)
            - `見方が動し` → `見方が変動し` (誤字: 動し → 動き)
            - `視する` → `重視する`  
            - `経成長` → `経済成長`  
            - `送配電備` → `送配電設備`  
            - `業見通し` → `業績見通し`  
            - `常増益` → `経常増益`  
            - `財政策` → `財政政策`  
            - `方` → `方針`  
            - `手Eコマース` → `大手Eコマース`  
            - `響しました` → `影響しました`  
            - `施され` → `実施された`  
            - `企業の合併・収` → `企業の合併・回収`  
            - `本とします` → `基本とします`  
            - `務状況` → `財務状況`
            - `内投資信託` → `国内投資信託`  
            - `持しました` → `維持しました`  
            - `マイナス因` → `マイナス要因`  
            - `反される` → `反映される`  
            - `替ヘッジ` → `為替ヘッジ`  
            - `比は` → `比率は`
            - `規緩和` → `規制緩和`
            - `景済指標` → `経済指標`
            - `剤` → `経済`
            - `昇するなどまちまちでした。` → `異なる動きとなりました。` (Ensure that the original text is not directly modified but follows this guideline.)
            - `積極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `消極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `（割安に）放置` → `割安感のある`
            - `限定的` → `他の適切な表現に修正 （修正理由: 効果や影響がプラスかマイナスか不明瞭なため）`
            - `利益確定の売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `利食い売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `必ず～` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～になる` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～である` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
  
            **Disambiguation Rule**:
            - 「沈静」＝自然に落ち着く (natural calming down; happens over time)
            - 「鎮静」＝人為的におさめる (intentional suppression; medically or artificially done)

            **Correction Policy**:
            1. Detect whether the context implies a natural or artificial calming.
            2. If the usage does not match the context, correct it using the appropriate word.
            3. Highlight the correction using the format below:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`
            4. Do **not** modify the original sentence structure or paragraph formatting.
            5. Only apply the correction when the term is clearly misused.
            6. If the current usage is correct, do not change or annotate it.

            **Example**:
            - Input: 市場は徐々に鎮静していった。
            - Output: 市場は徐々に <span style="color:red;">沈静</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">鎮静</s> → 沈静</span>) していった。


        - 表現の使用制限:
            Expression Usage Restrictions:
            Restricted Expressions:

            - 魅力的な
            - 投資妙味
            - 割高感
            - 割安感

            Usage Conditions:
            The above expressions can be used if evidence is provided.

            However, these expressions should not be used in contexts where the word "fund" (ファンド) or any related reference is mentioned. In any sentence or context where "fund" or "ファンド" appears, these expressions should be avoided.

            使用例:
            魅力的な: 根拠に基づいて使用することは可能ですが、ファンドについては使用しないようにしてください。
            投資妙味: 投資妙味があることを示す場合でも、ファンドに対する言及は避け、他の投資対象に適用するようにしてください。
            割高感: 割高感について述べる場合、ファンド以外の投資対象に対して適用してください。
            割安感: 割安感について言及する場合も、ファンドに対して使用することは不可です。

            ✅ 出力フォーマット:
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)
            ✅ Exsample1:
            Input:ファンドは魅力的な投資先として紹介された。

            Output:
            ファンドは
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対する使用は不可</span>)投資先として紹介された。
            ✅ Exsample2:
            Input:
            このファンド銘柄には投資妙味がある。
            
            Output:
            この銘柄には
            <span style="color:red;">投資妙味</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)がある。


        - 数字やパーセント（％）を含む文章の誤りをチェックする
            文脈を理解し、数値・割合の前後の語句が適切か確認すること。
            例:
            月末時点（20日判定）での為替ヘッジのターゲット比は48％です。
            → 正しい表記: 月末時点（20日判定）での為替ヘッジのターゲット比率は48％です。
            市場の成長率は10%の見込みです。 ✅ (問題なし)
            インフレ率は2上昇しました。 ❌ (誤り: 2%上昇しました。 に修正)
            販売シェアは15の拡大が予想されます。 ❌ (誤り: 販売シェアは15%の拡大が予想されます。)
        - 校正ルール
            1. **「行って来い」の適切な置き換え**(Ensure that the original text is not directly modified but follows this guideline.)
            - 文章全体を分析し、「行って来い」が何を指しているのかを判断してください。
            - **価格・指数・レートなどが上昇した意味の場合** → 「行って来い」を「上昇した」に変換
            - **価格・指数・レートなどが下落した意味の場合** → 「行って来い」を「下落した」に変換

            2. **文脈を考慮した校正**
            - 修正の際、周辺の文脈を理解し、自然な形に調整してください。
            
            3. **「横ばい」の適切な置き換え**
            - 文章全体を分析し、「横ばい」の前後の文脈を考慮してください。
            - **期中の変動幅が小さい場合** → 「横ばい」を維持
            - **変動幅が大きく、結果的に同程度となった場合** → 「ほぼ変わらず」または「同程度となる」に変更
            - 周囲の文章に合わせて、より適切な表現を使用してください。

            4. **「出遅れ感」の適切な修正**
            - 対応方針:

            「出遅れ感」 は主観的な相場観が含まれるため、必ず「…と考えます。」に修正してください。

            修正方法: 文脈に応じて自然に「〜と考えます」形に修正します。

            修正後の表現: 文の流れに合わせて自然に表現を調整し、読みやすさを考慮します。

            修正理由: 相場観が含まれているため、主観的な表現を客観的な表現に変えることで文章の信頼性を向上させます。

            出力フォーマット:
            <span style="color:red;">出遅れ感</span>
            (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 「〜と考えます」に修正</span>)

            Exsample:
                Input: この銘柄には出遅れ感がある
                Output: この銘柄には
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある

                Input: 一部のセクターには出遅れ感があると感じられる
                Output: 一部のセクターには
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある


            5. **「上昇要因」・「下落要因」の適切な説明追加**
            - **文脈を分析し、具体的な要因を追加してください**。
            - **「上昇要因」がある場合** → 上昇の理由（例: 企業決算の改善、政策の発表、需給バランスの変化など）を補足。
            - **「下落要因」がある場合** → 下落の理由（例: 景気後退懸念、金融引き締め、地政学リスクなど）を補足。
            - **修正後も文章の流れがスムーズになるように調整してください。

            6. **「予想」「心理」の適切な修正**
            - **「予想」 がある場合:**  
                - **誰の予想か明確でない場合** → 「市場予想」に修正  
            - **「心理」 がある場合:**  
                - **主語が曖昧な場合** → 「市場心理」に修正

        2. **Follow the "Fund Manager Comment Terminology Guide" (ファンドマネージャコメント用語集に沿った記載となっていること):**
        - **Consistent Terminology (表記の統一):**
            - Ensure the **writing format** of financial terms is **consistent throughout the report**.
            Example:
            - `相対に低かった` → `相対的に低かった` (文法修正)
            - `東証33業種分では` → `東証33業種分類では` (表記の統一)
        - **Common Mistakes and Corrections (誤記と修正例)**:
            Example:
            - `政府支の` → `政府支出の遅れ、1回はのを出、2回はのを削除` (誤字: の → 出)
            - `投資比率を維する` → `投資比率を維持する` (一致性不足 動産 → 不動産)
            - `（配当こみ）` → `（配当込み）` (表記の統一: こみ → 込み)
            - `いっぽうで` → `一方で` (表記の統一: いっぽう → 一方)
            - `ウクライナとロシアをめぐる` → `ウクライナとロシアを巡る` (表記の統一: めぐる → 巡る)
            - `東証33業種でみると` → `東証33業種で見ると` (文法修正: みる → 見る)
            - `ひき続き` → `引き続き` (表記の統一: ひき → 引き)
            - `電気機器、銀行業、保険業等` → `電気機器、銀行業、保険業など` (表記の統一: 等 → など)
            - `じき` → `次期`
            - `底いれ後` → `底入れ後`
            - `なかから` → `中から`
            - `おもな` → `主要な`
            - `はやく` → `早急に`
            - `かいつけした` → `買い付けした`
            - `など` → `等`
            - `のぞく` → `除く`
            - `くみいれ` → `組み入れ`
            - `じょうき` → `上記`
            - `とうファンド` → `当ファンド`

        - **Prohibited Words and Phrases (禁止（NG）ワード及び文章の注意事項):**
            - Check if any prohibited words or phrases are used in the report and correct them as per the guidelines.
        - **Replaceable and Recommended Terms/Expressions (置き換えが必要な用語/表現、置き換えを推奨する用語/表現):**
            - If you find terms or expressions that need to be replaced, revise them according to the provided rules.
            - ハト派／タカ派の表記（金融政策に関する）:
                -Exsample:
                - 金融緩和重視  → 金融引き締め重視
                - 金融緩和に前向き  → 金融引き締めに積極的
            - Exsample:
             - 織り込む  → 反映され
             - 相場  → 市場/価格
             - 連れ高  → 影響を受けて上昇
             - 伝播  → 広がる
             - トレンド  → 傾向
             - レンジ  → 範囲
        - **〇％を上回る（下回る）マイナスの表記:**
             - 〇％を上回る  → 〇％を超える
             - 〇％を下回る  → 縮小
             - 〇％を上回る  → 下回るマイナス幅
             - 〇％を下回る  → 縮小


        - **Use of Hiragana (ひらがなを表記するもの):**
            - Ensure the report follows the rules for hiragana notation, replacing content that does not conform to commonly used kanji.
        - **Kana Notation for Non-Standard Kanji (一部かな書き等で表記するもの):**
            - Ensure non-standard kanji are replaced with kana as the standard writing format.
        - **Correct Usage of Okurigana (一般的な送り仮名など):**
            - Ensure the correct usage of okurigana is applied.
        - **English Abbreviations, Loanwords, and Technical Terms (英略語、外来語、専門用語など):**
            - Check if English abbreviations, loanwords, and technical terms are expressed correctly.
        - **Identify and mark any 常用外漢字 (Hyōgai kanji):**
        - **Identify and mark any **常用外漢字 (Hyōgai kanji)** in the following text**
        - **常用外漢字** refers to characters **not included** in the [常用漢字表 (Jōyō Kanji List)](https://ja.wikipedia.org/wiki/常用漢字), which is Japan’s official list of commonly used kanji.
        - Refer to the [Wikipedia list of Hyōgai kanji](https://ja.wikipedia.org/wiki/常用漢字) to determine if a character falls into this category.
        - **For any detected 常用外漢字**, apply the following formatting:
        - **Highlight the incorrect character in red** (`<span style="color:red;">`).
        - **Strike through the incorrect character and provide the reason in yellow highlight.**

        ---

        ### **💡 Output Format (出力フォーマット)**
        - **Incorrect characters should be displayed in red (`<span style="color:red;">`)**.
        - **Corrected text should be marked with a strikethrough (`<s>`) and highlighted in yellow (`background:yellow`)** to show the correction.
        - Use the following structure:
            ```html
            <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **Example Correction**:
            ```plaintext
            鴉 → <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **For multiple Hyōgai kanji**, apply the same structure to each character.

        ---

        ### **✅ Example Input:**
        ```plaintext
        彼は鴉が空を飛ぶのを見た。

        ### **✅ Example output:**
        ```plaintext
        彼は <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>) が空を飛ぶのを見た。

        - **Foreign Exchange Market Trend Analysis**
            In the foreign exchange market (`為替市場`), determine whether `"円だか"` should be revised to `"円高"` (Yen Appreciation) or `"円安"` (Yen Depreciation) based on the **context**.

            #### **** Criteria for Yen Appreciation (円高)**
            - **Yen appreciation (`円高`) occurs when the value of the yen increases relative to other currencies.**  
            The following situations indicate yen appreciation:
            1. **"多くの通貨が対円で下落した"** (Many currencies declined against the yen) → Change `円だか` to **円高**.
            2. **"ドル円が下落した"** (USD/JPY exchange rate declined) → Change `円だか` to **円高**.
            3. **"対米ドルで円の価値が上昇した"** (The yen appreciated against the US dollar) → Change `円だか` to **円高**.

            #### **** Criteria for Yen Depreciation (円安)**
            - **Yen depreciation (`円安`) occurs when the value of the yen declines relative to other currencies.**  
            The following situations indicate yen depreciation:
            1. **"多くの通貨が対円で上昇した"** (Many currencies rose against the yen) → Change `円だか` to **円安**.
            2. **"ドル円が上昇した"** (USD/JPY exchange rate increased) → Change `円だか` to **円安**.
            3. **"対米ドルで円の価値が下落した"** (The yen depreciated against the US dollar) → Change `円だか` to **円安**.


        3. **Replaceable and Recommended Terms/Expressions (推奨される表現の修正)**
        - Use the correct **kanji, hiragana, and katakana** combinations based on standard Japanese financial terms.
            Example:
            - `が好された輸送用機器など` → `が好感された輸送用機器など` (修正理由: 適切な表現)

        - **Task**: Header Date Format Validation & Correction  
        - **Target Area**: Date notation in parentheses following "今後運用方針 (Future Policy Decision Basis)"  
        ---
        ### Validation Requirements  
        1. **Full Format Compliance Check**:  
        - Must follow "YYYY年MM月DD日現在" (Year-Month-Day as of)  
        - **Year**: 4-digit number (e.g., 2024)  
        - **Month**: 2-digit (01-12, e.g., April → 4)  
        - **Day**: 2-digit (01-31, e.g., 5th → 5)  
        - **Suffix**: Must end with "現在" (as of)  

        2. **Common Error Pattern Detection**:  
        ❌ "1月0日" → Missing month leading zero + invalid day 0  
        ❌ "2024年4月1日" → 2024年4月1日
        ❌ "2024年12月" → Missing day value  
        ❌ "2024-04-05現在" → Incorrect separator usage (hyphen/slash)  
        ---
        ### Correction Protocol  
        1. **Leading Zero Enforcement**  
        - Add leading zeros to single-digit months/days (4月 → 4月, 5日 → 5日)  

        2. **Day 0 Handling**  
        - Replace day 0 with YYYYMMDD Date Format  
        - Example: 2024年4月0日 → 2024年4月00日

        3. **Separator Standardization**  
        - Convert hyphens/slashes to CJK characters:  
            `2024/04/05` → `2024年4月5日`  

        ---
        ### Output Format Specification  
        ```html
        <Correction Example>
        <span style="color:red;">（2024年4月0日現在）</span> 
        → 
        <span style="color:green;">（2024年04月00日現在）</span>
        修正理由:
        ①日付0をYYYYMMDD日付フォーマットに置換
        ---

        3. **Consistency with Report Data Section (レポートのデータ部との整合性確認):**
        - Ensure the textual description in the report is completely consistent with the data section, without any logical or content-related discrepancies.

        4. **Eliminate language fluency(単語間の不要なスペース削除):**
        - Ensure that there are no extra spaces.
            -Example:
            input:景気浮揚が意 識されたことで
            output:景気浮揚が意識されたことで
        
        5.  **Layout and Formatting Rules (レイアウトに関する統一):**
            - **文頭の「○」印と一文字目の間隔を統一:**
   
                - When a sentence starts with the symbol ○, make sure there is no space (half-width or full-width) between it and the first character. That is, use ○文字 instead of ○ 文字 or ○　文字.
                    - Any whitespace (half-width or full-width spaces) after ○ must be removed.
                    - This spacing rule must be applied consistently throughout the document.

            半角括弧を全角括弧に統一:
                - Convert all half-width parentheses () to full-width parentheses （）.
                - Example: (注) → （注）
                - Example input: 
                    ○ 世界の高配当株式指数(注)は月間では上昇しました。
                - Exsample output: 
                    <span style="color:red;">○世界</span> (<span>修正理由: 文頭の「○」印と一文字目の間隔を統一 <s style="background:yellow;color:red">○ 世界</s> → ○世界</span>)
                    <span style="color:red;">（注）</span> (<span>修正理由: 半角括弧を全角括弧に統一 <s style="background:yellow;color:red">(注)</s> → （注）</span>)


            - **文章の間隔の統一:**
                - If a sentence begins with "○", ensure that the spacing within the frame remains consistent.
            - **上位10銘柄 コメント欄について、枠内に適切に収まっているかチェック:**
                - If the stock commentary contains a large amount of text, confirm whether it fits within the designated frame. 
                - If the ranking changes in the following month, adjust the frame accordingly.
                - **Check point**
                    1. **文字数制限内に収まっているか？**
                    - 1枠あたりの最大文字数を超えていないか？
                    - 適切な行数で収まっているか？

                    2. **次月の順位変動に伴う枠調整の必要性**
                    - 順位が変更されると枠調整が必要なため、調整が必要な箇所を特定

                    3. **枠内に収まらない場合の修正提案**
                    - 必要に応じて、短縮表現や不要な情報の削除を提案
                    - 重要な情報を損なわずに適切にリライト

                    output Format:
                    - **コメントの枠超過チェック**
                    - (枠超過しているか: はい / いいえ)
                    - (超過している場合、オーバーした文字数)

                    - **順位変動による枠調整の必要性**
                    - (調整が必要なコメントリスト)

                    - **修正提案**
                    - (枠内に収めるための修正後のコメント)

            **Standardized Notation (表記の統一):**
            - **基準価額の騰落率:**
            When there are three decimal places, round off using the round-half-up method to the second decimal place. If there are only two decimal places, keep the value unchanged.
                Make modifications directly in this article and explain the reasons for the modifications.

                exsample:
                0.546％（×） → 0.55％（○）
                修正理由: 小数点以下の桁数の丸め（0.546％ → 0.55％）
                If the value is 0.00％, replace it with "前月末から変わらず" or "前月末と同程度" instead of stating "騰落率は変わらず".
                修正理由: 「騰落率は変わらず」という表記はNG。代わりに「基準価額（分配金再投資）は前月末から変わらず」や「前月末と同程度」と記載します。

                exsample:
                0.00％となり（×） → 前月末から変わらず（○）

                騰落率は変わらず（×） → 基準価額（分配金再投資）は前月末から変わらず（○）

                When comparing the performance of the fund with the benchmark (or reference index), the comparison must be made using rounded numbers.

                修正理由: 比較は丸めた数字で行うこと。
                If the fund and benchmark (or reference index) have the same rate of return, use the phrase "騰落率は同程度となりました" instead of saying "騰落率は同じでした".
                修正理由: 同じという表現は避け、代わりに「同程度」と記載すること。

                exsample:
                「騰落率は同じでした」（×） → 「騰落率は同程度となりました」（○）
                If the fund's rate of return is greater than the benchmark's, use the phrase "上回りました" to indicate the fund outperformed the benchmark.
                修正理由: 上回った場合、「上回りました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-2.221％（ベンチマーク）の場合、値の差は0.79％となるため、「上回りました」と記載します。

                If the fund's rate of return is lower than the benchmark's, use the phrase "下降しました" to indicate the fund underperformed the benchmark.
                修正理由: 下降した場合、「下降しました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-0.221％（ベンチマーク）の場合、基金のパフォーマンスは「下降しました」と記載します。

            - **「今後の運用方針」作成日付のルール:**
                - 前月末（営業日）現在で作成。
                - 翌月初の日付になる場合は、作成した日付を入れる。
                - クライアント・サービス部へ送信する以降の日付（先日付）は入れない。
                - 「参考月」より後であり、「チェック期間」より前の日付のみ使用可。
                    - Example:
                    - OK: 参考月＝2024年2月 → 作成日が2024年2月28日（営業日） or 3月1日（翌月初）
                    - NG: 参考月＝2024年2月 → 作成日が3月5日（クライアント・サービス部送信後の先日付）

            - **％（パーセント）、カタカナ:**
                Make modifications directly in this article and explain the reasons for the modifications.
                - **半角カタカナ → 全角カタカナ**（例:「ｶﾀｶﾅ」→「カタカナ」）
                - **半角記号 → 全角記号**（例:「%」→「％」、「@」→「＠」）
                    Example:
                        input: ﾍﾞﾝﾁﾏｰｸ (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して 
                        output: ベンチマーク (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して
                    Example:
                        input: ｶﾀｶﾅ 
                        output: カタカナ
                    Example:
                        input: %
                        output: ％ 
                    Example:
                        input: @
                        output: ＠ 

            - **数字、アルファベット、「＋」・「－」:**
                - **全角数字・アルファベット → 半角数字・アルファベット**（例:「１２３」→「123」、「ＡＢＣ」→「ABC」）
                - **全角「＋」「－」 → 半角「+」「-」**（例:「＋－」→「+-」
                    Example:
                        input: １２３ ＡＢＣ ｱｲｳ ＋－
                        output: 123 ABC アイウ +-

            - **スペースは変更なし**  

            - **「※」の使用:**
                - 「※」は可能であれば **上付き文字（superscript）※** に変換してください。
                - 出力形式の例:
                - 「重要事項※」 → 「重要事項<sup>※</sup>」

            - **（カッコ書き）:**
                - Parenthetical notes should only be included in their first occurrence in a comment.
                    For the following Japanese text, check if parentheses ("（ ）") are used appropriately.
                    If a parenthetical note appears more than once, remove the parentheses for subsequent occurrences.
                    The first occurrence should retain the parentheses, but any further appearances should have the parentheses removed.
                    Modification reason: Parentheses are redundant after the first mention, so the text is cleaned up for consistency and readability.

                **Check point**
                    1. **カッコ書きは、コメントの初出のみに記載されているか？**
                    - 同じカッコ書きが2回以上登場していないか？
                    - 初出ページ以降のコメントにカッコ書きが重複して記載されていないか？

                    2. **ディスクロのページ番号順に従ってルールを適用**
                    - シートの順番ではなく、実際のページ番号を基準にする。

                    3. **例外処理**
                    - 「一部例外ファンドあり」とあるため、例外的にカッコ書きが複数回登場するケースを考慮する。
                    - 例外として認められるケースを判断し、適切に指摘。

                    output Format:
                    - **カッコ書きの初出リスト**（どのページに最初に登場したか）
                    - **重複チェック結果**（どのページで二重記載されているか）
                    - **修正提案**（どのページのカッコ書きを削除すべきか）
                    - **例外ファンドが適用される場合、補足情報**

            - **会計期間の表記:**
                - The use of "～" is prohibited; always use "-".
                - Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 6～8月期（×） → 6-8月期（○）

                - 暦年を採用している国の年度表記:
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                カッコ書きで暦年の期間を明記する。
                - Example:
                    ブラジルの2021年度予算（×） → ブラジルの2021年度（2021年1月-12月）予算（○）

                - 決算期間は「●-●月期」に統一し、日付は省略する。
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 第1四半期（5月21日～8月20日）（×） → 5-8月期（○）
                    イレギュラーなケースも含め、原則「●-●月期」と表記。

            - **「TOPIX」または「東証株価指数」が含まれている場合、以下のルールを適用:**
                文中で使用する場合: 「TOPIX（東証株価指数）」と表記することを指示。
                「文中では『TOPIX（東証株価指数）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                ベンチマーク（BM）や参考指数として使用する場合: 「東証株価指数（TOPIX）（配当込み）」と表記することを指示。
                「BMや参考指数で使用する場合は、『東証株価指数（TOPIX）（配当込み）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」

            - **年をまたぐディスクロコメントの年度表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - When specifying the fiscal year in disclosure comments that span multiple years, always use "yyyy年度".
                - Similarly, for disclosures based on the January-March period, specify the corresponding year.
                - Example:
                    - For a disclosure with a December-end reference, released in January:
                    - 今年度（×） → 2021年度（○）
                    - 来年度（×） → 2022年度（○）
            - **Benchmark, Index, and Reference Index Name Formatting:(ベンチマーク・インデックス・参考指数の名称の表記)
               - Ensure Consistency in Index Terminology:
                    Read the context and identify terms related to "index" (指数) within the text. Ensure that these terms are unified and consistently referred to using the correct and standardized terminology.
                    It is important to carefully analyze each mention of "index" to make sure the terminology is consistent throughout the text.
                    Do not modify the original text directly. Instead, provide comments that explain the reasoning behind the proposed changes, especially when identifying inconsistencies or clarifications needed.
                Example Formatting Guidelines:

                    Incorrect format (×): "ISM非製造業景況"
                    Correct format (○): "ISM非製造業景況指数"
                    Incorrect format (×): "MSCIインド"
                    Correct format (○): "MSCIインド・インデックス"
                    If multiple terms are used to refer to the same index, they should be unified under the correct term. For example, if "MSCIインド指数" and "MSCIインド・インデックス" are used in different places, they should be unified as "MSCIインド・インデックス" in the final report to maintain consistency.
                Handling Multiple Terms Referring to the Same Index:

                    If it can be clearly determined that different terms refer to the same index (e.g., "MSCIインド指数" and "MSCIインド・インデックス"), do not modify them but mark them accordingly. These terms should be noted as referring to the same index.
                    Example:
                    Original: "MSCIインド指数" and "MSCIインド・インデックス"
                    Comment: These are different ways of referring to the same index, so no change is needed.
                Handling Uncertainty in Index Terminology:
                    
                    If there is uncertainty about whether multiple terms refer to the same index (e.g., it is unclear whether "ISM非製造業景況" and "ISM非製造業景況指数" refer to the same index), mark them without modification. Additionally, note that these terms might refer to the same index, but the exact nature of the index should be verified.
                    Example:
                    Original: "ISM非製造業景況" and "ISM非製造業景況指数"
                    Comment: These terms are potentially referring to the same index but require further clarification. Therefore, no changes are made in this case.
                Key Notes:

                Always ensure that consistency is maintained across the report. Even if different names are used for the same index, it is essential to mark them properly and explain that they are different terms for the same entity.
                Consistency applies not only to the formatting of the terms but also to how the terms are presented across the entire document. All references to a given index must follow the same format from the first mention to the last.


            - **上昇 or 下落に関する要因を明記:**
                文章内に 「上昇」 または 「下落」 という単語が含まれている場合、その要因を特定し、明記してください。
                output:
                「●●は上昇（または下落）しました。(理由: ○○)」
            - **指定用語の表記ルールを提示:**
                独Ifo企業景況感指数 または 独Ifo景況感指数 が含まれている場合、以下のメッセージを表示する。
                独Ifo企業景況感指数 または独Ifo景況感指数 の表記について、月報内での統一ルールを確認してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                
                「独ZEW景気期待指数」または「独ZEW景況感指数」 が含まれている場合、以下のメッセージを表示する。
                「『独ZEW景気期待指数』または『独ZEW景況感指数』の表記を使用してください。ZEW単独使用の場合は括弧書き付き、または『欧州経済研究センター』のみとし、『ZEW』単独使用を避けてください。(Ensure that the original text is not directly modified but follows this guideline.)」
            - **特定の金融用語に対する表記ルールを確認・適用:**

                「リフレーション」 が含まれている場合、以下のメッセージを表示。
                「リフレーションとは、デフレーションから抜けて、まだインフレーションにはなっていない状況を指します。」

                「リスクプレミアム」 が含まれている場合、以下のメッセージを表示。
                「リスクプレミアムとは、あるリスク資産の期待収益率が、同期間の無リスク資産（国債など）の収益率を上回る幅を指します。」

                「モメンタム」 が含まれている場合、以下のメッセージを表示。
                「『モメンタム』の使用は避け、相場の『勢い』や『方向性』などの言葉に置き換えてください。」

                「ベージュブック」 が含まれている場合、以下のメッセージを表示。
                「『ベージュブック』は『FRB（米連邦準備制度理事会）が発表したベージュブック（地区連銀経済報告）』と明記してください。スペースが限られる場合は、『ベージュブック（米地区連銀経済報告）』としてください。」

                「フリーキャッシュフロー」 が含まれている場合、以下のメッセージを表示。
                「フリーキャッシュフローとは、税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたものです。」
                
                「システミック・リスク」 が含まれている場合、以下のメッセージを表示。
                「システミック・リスクとは、個別の金融機関の支払不能や特定の市場・決済システム等の機能不全が、他の金融機関、市場、または金融システム全体に波及するリスクを指します。」

                「クレジット（信用）市場」 が含まれている場合、以下のメッセージを表示。
                「クレジット（信用）市場とは、信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称であり、企業の信用リスクを取引する市場です。」
            - **特定の金融用語に対し、欄外に注記を加える指示を表示:**
                「格付別」 が含まれている場合:
                    格付別 -> 格付別
                「格付機関」 が含まれている場合:
                    格付機関 -> 格付機関
                「組入比率」 が含まれている場合:
                    組入比率 -> 組入比率
                「引締策」 が含まれている場合:
                    引締策 -> 引締策
                「国債買入れオペ」 が含まれている場合:
                    国債買入れオペ -> 国債買入オペ

                「投資適格債」 が含まれている場合:
                「※欄外に注記: 投資適格債とは、格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債を指します。」

                「デュレーション」 が含まれている場合:
                「※欄外に注記: デュレーションとは、金利が一定の割合で変動した場合、債券の価格がどの程度変化するかを示す指標です。この値が大きいほど、金利変動に対する債券価格の変動率が大きくなります。」

                「デフォルト債」 が含まれている場合:
                「※欄外に注記: デフォルトとは、一般的に債券の利払いおよび元本返済の不履行、または遅延などを指し、このような状態にある債券を『デフォルト債』といいます。」

                「ディストレス債券」 が含まれている場合:
                「※欄外に注記: ディストレス債券とは、信用事由などにより価格が著しく下落した債券を指します。」
                
                「イールドカーブ」 が含まれている場合:
                「※欄外に注記: イールドカーブ（利回り曲線）とは、横軸に残存年数、縦軸に利回りをとった座標に、債券利回りを点描して結んだ曲線のことを指します。」

            - **組入上位10銘柄」について記述がある場合、以下のルールを適用:**
                「組入上位10銘柄を超える保有銘柄（個別銘柄の特定が可能な子会社名等を含む）は原則として開示禁止である」ことを明示。
                ただし、社内規程に基づき開示が認められているファンドは例外とすることを伝える。

            - **年度表記:**
                - Use four-digit notation for years.(Ensure that the original text is not directly modified but follows this guideline.)
                - Example: 22年（×） → 2022年（○）

            - **前年比 or 前年同月（同期）比の統一:**
                - 「前年同月（同期）比」に統一。
                - 通年の比較には「前年比」の使用可。
                - Ensure that the original text is not directly modified but follows this guideline.
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example:
                        前年比+3.0%（×） → 前年同月比+3.0%（○）
                        2023年のGDPは前年比+3.0％（○）

            - **年をまたいだ経済指標の記載:**
                - コメント内の初出のみに記載する。(In the case where there is a description of the economic indicator over the year, it is described only in the first comment.)
                    - Example:
                        - 2023年12月のCPIは～（○）
                        - 一方2024年1月のユーロ圏PMIは～ （○）
                        - 10-12月期のGDPは～（○）
            - **経済指標について:**
                    -"加速" の対象を明確に記載すること。文脈を考慮し、"加速" の対象を適切に補う。
                        文脈に応じて、"何が加速したのか" を判断し、適切な単語に修正してください。
                        修正ルール:

                        - 前月から加速しました（×） → 何が加速したのかを明記（○）
                        Exsample: 「前月から上昇が加速しました」 → 「景気回復の加速（○）」
                        - 前月から上昇が加速しました（×） → 具体的な経済活動を明記（○）
                        Exsample: 「景気加速（○）」「消費の回復が加速（○）」「投資の拡大が加速（○）」
                        - 経済（×） → 景気（○）（"経済" ではなく "景気" を用いる）
                        Exsample: 
                        前月から加速しました。（×）-> 企業の設備投資が加速しました。（○）
                        経済加速が見られます。（×）-> 景気加速が見られます。（○）
                        消費が前月から加速しました。（×）-> 個人消費の拡大が加速しました。（○）
                        インフレが前月から加速しました。（×）-> 物価上昇のスピードが加速しました。（○）


                    - 日付・国名の明記:
                        いつのものか特定できる場合、○月のものかを明記する。例: 「10月の製造業PMI（購買担当者景気指数）は～」
                        必要に応じて国名も記載する。（文脈に応じて記載していれば、位置は問わない）
                    - 日付・国名の変換ルール:

                        **「下旬」「上旬」**と記載されている場合、文脈から適切な日付に変更する。
                        **「ユーロ圏」**と記載されている場合、文脈に応じて適切な国名に置き換える。
                        Exsample:

                        修正前: 「下旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月下旬のドイツ総合PMI（購買担当者景気指数）が…」

                        修正前: 「上旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月上旬のフランス総合PMI（購買担当者景気指数）が…」

            - **業界指数の表記:**
                - 必ず対象となる「月」を明記する。
                - 月がない場合、最近3ヶ月以内か、3ヶ月以上前のものかを確認する。
                - 必要に応じて国名も記載する。（文脈に応じて位置は自由）
                    - Example:
                        - 製造業PMI（購買担当者景気指数）（×） → 10月の製造業PMI（○）
                        - 直近3ヶ月以内の指数は明示的に「○月」と記載する。（例:12月のCPI）
                        - 3ヶ月以上前の指数は、比較の文脈を明確にする。（例:2023年10月のGDP成長率）
                        - ユーロ圏の10月PMIは～（○）
            - **カタカナ表記の統一:**
                Katakana representation of foreign words should be unified within the document.
                    Ensure that the Katakana form is consistent throughout the text, and choose one version for the entire document.
                    Modification reason: To maintain consistency in the usage of Katakana for foreign words.
                    Example of text modifications:

                    サステナブル (×) → サスティナブル (○)

                    エンターテイメント (×) → エンターテインメント (○)
             
            - **レンジの表記について表記:**
                - Always append "%" when indicating a range.(Ensure that the original text is not directly modified but follows this guideline.)
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: -1～0.5%（×） → -1%～0.5%（○）
            - **償還に関する記載:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 最終リリースの1ヵ月程前より、償還に関する内容を入れること。
                - 例）当ファンドは、●●月●●日に信託の終了日（償還日or繰上償還日）を迎える予定です。
                - ※（）内は、定時償還の場合には償還日、繰上償還の場合には繰上償還日とする。
            - **個別企業名の表記:**
                - 投資環境等においては、個別企業の名称を使わない表現を心掛ける。
                - 例:スイス金融大手クレディ・スイス（×） → スイスの大手金融グループ（○）
            - **プラスに寄与/影響の表記:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 「プラスに寄与」または「プラスに影響」どちらも可。
                - または、「～プラス要因となる」と表記。
            - **マイナスに影響の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「マイナスに寄与」（×）→「マイナスに影響」（○）
                - マイナスの際は「寄与」は使用しない。
                - または、「～マイナス要因となる」と表記。
            - **利回りの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 利回りは「上昇（価格は下落）」または「低下（価格は上昇）」と表記。
            - **低下と下落の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 債券利回りは「低下（○）」と表記し、「下落（×）」は使用しない。
                - 価格は「下落（○）」と表記し、「低下（×）」は使用しない。
                - 金利の「低下（〇）」と表記し、「下落（×）」は使用しない。

            - **資金流出入の表記:**
                - Do notMake modifications directly in this article and explain the reasons for the modifications.
                - 「外国人投資家の資金流出」を「外国人投資家からの資金流入」と記載。

            - **（金利の）先高感/先高観 の表記統一:**
                - 文中に「先高観」という表記がある場合でも、原文は修正しないでください。
                - その代わり、「先高観」の直後に「修正提案」として、「先高感」への統一理由を提示してください。
                - 表記がすでに「先高感」である場合は、何も追記せずそのままにしてください。

                - 表示フォーマット（例）:
                    先高観<span style="color:red;">（修正理由: 用語の統一 <s style="background:yellow;color:red">先高観</s> → 先高感）</span>

                - 必ず原文の構成と文脈を保持し、構文を壊さず、修正理由は補足的に後ろに追記してください。
                
            - **ポートフォリオの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「●●への組み入れ（×）」ではなく、「●●の組み入れ（○）」と表記。
                - 「への投資比率」は使用可能。
                
            - **構成比の0％の表記:
                - 「0％程度」or「ゼロ％程度」の表記を使用すること
                - 変更前表記: 構成比は0％である
                - 統一後表記: 構成比は0％程度
                - Append a correction reason in the following format:
                        `<span style="color:red;">変更前表記</span> (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: 構成比は0％である
                    Output: 
                    <span style="color:red;">構成比は0％である</span> 
                    (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">構成比は0％である</s> → 構成比は0％程度である。</span>)で売られています。

                
            - **'投資環境の記述:** 
                - Make modifications directly in this article and explain the reasons for the modifications.
                **「先月の投資環境」**の部分で「先月末」の記述が含まれる場合、「前月末」に変更してください。
                - Ensure that the original text is directly modified and follows this guideline.
                Example:
                修正前: 先月末の市場動向を分析すると…
                修正後: 前月末の市場動向を分析すると…

            - **通貨表記の統一:**
                - Standardize currency notation across the document.
                    - The first appearance of any currency symbol (e.g., ドル, $, 円, JPY) will be the standard.
                    - All following occurrences of that currency must match this format.

                    - For example, if "100ドル" appears first, then all future "$100" will be rewritten to "100ドル" for consistency.
                    - If "$100" appears first, then "100ドル" should be rewritten as "$100".

                    - Always apply this rule in the direction of "first-appeared" format.
                    - Append a correction reason in the following format:
                        `<span style="color:red;">統一後表記</span> (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: このバッグは100ドルですが、アメリカでは$100で売られています。
                    Output:
                    このバッグは100ドルですが、アメリカでは
                    <span style="color:red;">100ドル</span>
                    (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">$100</s> → 100ドル</span>)で売られています。


            **Preferred and Recommended Terminology (置き換えが必要な用語/表現):**
            - **第1四半期:**
                - Ensure the period is clearly stated.
                - Example: 18年第4四半期（×） → 2018年10-12月期（○）
            - **約○％程度:**
                - Do not use "約" (approximately) and "程度" (extent) together. Choose either one.
                - Example: 約○％程度（×） → 約○％ or ○％程度（○）
            - **大手企業表記の明確化**  
                **Correction Rule:**
                - If a sentence contains vague expressions like「大手○○」, analyze the context to determine what type of company is being referred to.
                - Rewrite it in the format:「大手○○会社」/「大手○○企業」/「大手○○メーカー」depending on the company’s nature.
                - Use context clues (e.g., product type, industry references) to guess the appropriate company category (e.g., 不動産, 自動車, 電機, 金融).
                - Append a correction reason in this format:
                `<span style="color:red;">Changed Expression</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Expression</s> → Changed Expression</span>)`

                **Example Input:**
                - 大手は業界全体に影響力を持つ。
                - 大手が新しい半導体を発表した。

                **Example Output:**
                - <span style="color:red;">大手不動産会社</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手不動産会社</span>) は業界全体に影響力を持つ。
                - <span style="color:red;">大手半導体メーカー</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手半導体メーカー</span>) が新しい半導体を発表した。

                **Important Notes:**
                - Always preserve the original sentence structure and paragraph formatting.
                - Only make corrections when「○○大手」is ambiguous and can be clarified using contextual information.
                - Do not modify proper nouns or known company names (e.g., トヨタ, ソニー).

            - **入力例:**  
                - 「大手メーカー/会社/企業」  
                - **出力:** 「大手不動産会社、大手半導体メーカー」  
            - **The actual company name must be found and converted in the article
            - **先月/前月の表記:
                - 1ヵ月前について言及する場合は、「前月」を使用。
            前期比○％の表記:

            - **前期比年率○％:**
                - 基本的に、期間比較の伸率は「年率」を記載してください。
                - 主に経済統計等で一般的に前期比で年率換算されているものについては、「前期比年率○％」と表記。
            - **第○四半期の表記:**
                **ルール:
                - If the input contains a format like "18年第4四半期", infer it as:
                    - "18年" → "2018年"
                    - "第1四半期" → "1-3月期"
                    - "第2四半期" → "4-6月期"
                    - "第3四半期" → "7-9月期"
                    - "第4四半期" → "10-12月期"
                - Modify the expression accordingly, converting the year to a 4-digit format and specifying the exact month range.
                - Add a correction reason in this format:
                `<span style="color:red;">修正後</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">修正前</s> → 修正後</span>)`

                ---

                **Example:**
                - Input: 18年第4四半期の売上が好調だった。
                - Output: 
                <span style="color:red;">2018年10-12月期</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">18年第4四半期</s> → 2018年10-12月期</span>) の売上が好調だった。

                ---

                **Additional Notes:**
                - Do not modify any proper names, organizations, or if the date range is already correct.
                - Apply to all similar shorthand expressions like "20年第2四半期", "21年第1四半期" etc.
                - Keep the structure and formatting of the original document.


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).

        **Output Requirements:**
        1. **Highlight the original incorrect text in red and include additional details:**
        - For corrected parts:
            - Highlight the original incorrect text in red using `<span style="color:red;">`.
            - Append the corrected text in parentheses, marked with a strikethrough using `<s>` tags.
            - Provide the reason for the correction and indicate the change using the format `123 → 456`.
            - Example:
            `<span style="color:red;">123</span> (<span>修正理由: 一致性不足 <s style="background:yellow;color:red">123</s> → 456</span>)`
        
        2. **Preserve the original structure and formatting of the document:**
        - Maintain paragraph breaks, headings, and any existing structure in the content.

        3. **Use the uploaded correction rules for reference:**
        - {prompt}

        4. **Do not provide any explanations or descriptions in the output. Only return the corrected HTML content.**

         **Corrected Terminology Map (修正された用語リスト):
            {corrected_map}
        - Replace only when the **original** term in `corrected_map` appears in the input text.
        - Do **not** replace anything if the input already contains the `corrected` term (it is already correct).
        - Do **not** perform any reverse replacements (`corrected → original` は禁止).
        - Modify the original text only when the `original` term is found.

        - If the `corrected` term appears in the input, **do not modify it** (it is already correct).
        - Do **not** reverse substitutions (i.e., never convert `corrected` → `original`).

        - After replacing, add the reason in this format:
        Original Term (修正理由: 用語の統一 Original Term → Corrected Term)
        Example:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`

        Example:
        Input: 中銀
        Output: 
        `<span style="color:red;">中央銀行</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">中銀</s> → 中央銀行</span>)`
        ※ Note: Do **not** convert 中央銀行 → 中銀. All replacements must follow the direction from `original` to `corrected` only.

        Input: 中央銀行  
        Output:  
        中央銀行 ← (No correction shown because it is already the correct term)

        If the input already contains the corrected term, it should remain unchanged.
        For English abbreviations or foreign terms, the rule is the same: replace the original term with the corrected term and format as follows:
        Example:
        Input: BOE
        Output: <span style="color:red;">BOE（英中央銀行、イングランド銀行）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BOE</s> → BOE（英中央銀行、イングランド銀行）</span>)
        Input: AAA
        Output: <span style="color:red;">AAA（全米自動車協会）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AAA</s> → AAA（全米自動車協会）</span>)

        Input: インバウンド
        Output: <span style="color:red;">インバウンド（観光客の受け入れ）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">インバウンド</s> → インバウンド（観光客の受け入れ）</span>)

        
        **Except Original Term
        Input: 等
        Output: 
        `<span style="color:red;">等</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">等</s> → など</span>)`

        Input: ローン
        Output: 
        `<span style="color:red;">ローン</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">ローン</s> → 貸し付け</span>)`
                            
        Input: ％を上回る
        Output: 
        `<span style="color:red;">％を上回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を上回る</s> → ％を超える</span>)`
                            
        Input: ％を下回る
        Output: 
        `<span style="color:red;">％を下回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を下回る</s> → ％を下回るマイナス幅</span>)`
        
        Input: 伝播（でんぱ）
        Output: 
        `<span style="color:red;">伝播（でんぱ）</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播（でんぱ）</s> → 広ります</span>)`
        
        Input: 伝播しています
        Output:
        `<span style="color:red;">伝播しています</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播しています</s> → 広がるしています</span>)`
        
        Input: 連れ高
        Output:
        `<span style="color:red;">連れ高</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">連れ高</s> → 影響を受けて上昇</span>)`
        
        Input: 相場
        Output:
        `<span style="color:red;">相場</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">相場</s> → 市場/価格</span>)`
        
        Input: ハト派
        Output:
        `<span style="color:red;">ハト派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">ハト派</s> → 金融緩和重視、金融緩和に前向き</span>)`
        
        Input: タカ派
        Output:
        `<span style="color:red;">タカ派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">タカ派</s> → 金融引き締め重視、金融引き締めに積極的</span>)`
        
        Input: 織り込む
        Output: 
        `<span style="color:red;">織り込む</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">織り込む</s> → 反映され</span>)`
        
        Input: 積極姿勢とした
        Output: 
        `<span style="color:red;">積極姿勢とした</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">積極姿勢とした</s> → 長めとした</span>)`
        
        Input: 限定的
        Output: 
        `<span style="color:red;">限定的</span> (<span>修正理由: 効果や影響がプラスかマイナスか不明瞭なため <s style="background:yellow;color:red">限定的</s> → 他の適切な表現に修正</span>)`
        
        Input: 利益確定の売り
        Output: 
        `<span style="color:red;">利益確定の売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利益確定の売り</s> → が出たとの見方</span>)`
        
        Input: 利食い売り
        Output: 
        `<span style="color:red;">利食い売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利食い売り</s> → が出たとの見方</span>)`
        
        Input: ABS
        Output: 
        `<span style="color:red;">ABS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ABS</s> → ABS（資産担保証券、各種資産担保証券）</span>)`
        
        Input: AI
        Output: 
        `<span style="color:red;">AI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AI</s> → AI（人工知能</span>)`
        
        Input: BRICS（5ヵ国）
        Output: 
        `<span style="color:red;">BRICS（5ヵ国）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BRICS（5ヵ国）</s> → BRICS（ブラジル、ロシア、インド、中国、南アフリカ）</span>)`
        
        Input: CMBS
        Output: 
        `<span style="color:red;">CMBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">CMBS</s> → CMBS（商業用不動産ローン担保証券）</span>)`
        
        Input: ISM
        Output: 
        `<span style="color:red;">ISM</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ISM</s> → ISM（全米供給管理協会）</span>)`
        
        Input: IT
        Output: 
        `<span style="color:red;">IT</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">IT</s> → IT（情報技術）</span>)`
        
        Input: MBS
        Output: 
        `<span style="color:red;">MBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">MBS</s> → MBS（住宅ローン担保証券）</span>)`
        
        Input: PMI
        Output: 
        `<span style="color:red;">PMI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">PMI</s> → PMI（購買担当者景気指数）</span>)`
        
        Input: S&P
        Output: 
        `<span style="color:red;">S&P</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">S&P</s> → S&P（スタンダード・アンド・プアーズ）社</span>)`
        
        Input: アセットアロケーション
        Output: 
        `<span style="color:red;">アセットアロケーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">アセットアロケーション</s> → アセットアロケーション（資産配分）</span>)`
        
        Input: E-コマース
        Output: 
        `<span style="color:red;">Eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">E-コマース</s> → Eコマース（電子商取引）</span>)`
             
        Input: e-コマース
        Output: 
        `<span style="color:red;">eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">eコマース</s> → eコマース（電子商取引）</span>)`
           
        Input: EC
        Output: 
        `<span style="color:red;">EC（電子商取引）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">EC</s> → EC（電子商取引）</span>)`
        
        Input: イールドカーブ
        Output: 
        `<span style="color:red;">イールドカーブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">イールドカーブ</s> → イールドカーブ（利回り曲線）</span>)`
        
        Input: エクスポージャー
        Output: 
        `<span style="color:red;">エクスポージャー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">エクスポージャー</s> → ＊積極的に使用しない。　（価格変動リスク資産の配分比率、割合）</span>)`
        
        Input: クレジット（信用）市場
        Output: 
        `<span style="color:red;">クレジット（信用）市場</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">クレジット（信用）市場</s> → 信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称。　企業の信用リスクを取引する市場。</span>)`
        
        Input: システミック・リスク
        Output: 
        `<span style="color:red;">システミック・リスク</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">システミック・リスク</s> → 個別の金融機関の支払不能等や、特定の市場または決済システム等の機能不全が、他の金融機関、他の市場、または金融システム全体に波及するリスク</span>)`
        
        Input: ディストレス債券
        Output: 
        `<span style="color:red;">ディストレス債券</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディストレス債券</s> → 信用事由などにより、価格が著しく下落した債券</span>)`
        
        Input: ディフェンシブ
        Output: 
        `<span style="color:red;">ディフェンシブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディフェンシブ</s> → ディフェンシブ（景気に左右されにくい）</span>)`
        
        Input: テクニカル
        Output: 
        `<span style="color:red;">テクニカル</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">テクニカル</s> → テクニカル（過去の株価の動きから判断すること</span>)`
        
        Input: デフォルト
        Output: 
        `<span style="color:red;">デフォルト</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト</s> → デフォルト（債務不履行）</span>)`
        
        Input: デフォルト債
        Output: 
        `<span style="color:red;">デフォルト債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト債</s> → デフォルトとは一般的には債券の利払いおよび元本返済の不履行、もしくは遅延などをいい、このような状態にある債券をデフォルト債といいます。</span>)`
        
        Input: デュレーション
        Output: 
        `<span style="color:red;">デュレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デュレーション</s> → デュレーション（金利感応度）</span>)`
        
        Input: 投資適格債
        Output: 
        `<span style="color:red;">投資適格債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">投資適格債</s> → 格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債</span>)`
        
        Input: ファンダメンタルズ
        Output: 
        `<span style="color:red;">ファンダメンタルズ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ファンダメンタルズ</s> → ファンダメンタルズ（賃料や空室率、需給関係などの基礎的条件）※REITファンドで使用する</span>)`
        
        Input: フリーキャッシュフロー
        Output: 
        `<span style="color:red;">フリーキャッシュフロー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">フリーキャッシュフロー</s> → 税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたもの</span>)`
        
        Input: ベージュブック
        Output: 
        `<span style="color:red;">ベージュブック</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ベージュブック</s> → スペースがない場合は、ベージュブック（米地区連銀経済報告）</span>)`
        
        Input: モメンタム
        Output: 
        `<span style="color:red;">モメンタム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">モメンタム</s> →  相場の勢い)が強く、投資家たちは短期的な利益を狙っています。</span>)`
        
        Input: リオープン
        Output: 
        `<span style="color:red;">リオープン</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リオープン</s> → リオープン/リオープニング（経済活動再開）</span>)`
        
        Input: リスクプレミアム
        Output: 
        `<span style="color:red;">リスクプレミアム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リスクプレミアム</s> → 同じ投資期間内において、あるリスク資産の期待収益率が、無リスク資産（国債など）の収益率を上回る幅のこと。</span>)`
        
        Input: リフレーション
        Output: 
        `<span style="color:red;">リフレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リフレーション</s> → リフレーション**デフレーションから抜けて、まだ、インフレーションにはなっていない状況のこと。</span>)`
        
        **【例外用語 – 修正しないこと】**
        - コロナ禍
        - コロナショック
        - 新型コロナ禍
        - 住宅ローン
        - 引き締め策
        - 引き締め政策
        - 組入比率
        - 格付機関
        - 格付別
        - 国債買入オペ

        **特定表現の言い換えルール（文脈判断を伴う修正）:
        文脈に応じて、具体的な表現に言い換えてください。
        「まちまち」の使用
        「まちまち」という曖昧な表現が出現した場合は、その語をそのまま保持した上で、後に「修正理由: 曖昧表現の明確化」を補足してください。

        変換語「異なる動き」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        <span style="color:red;">まちまち</span>
        (<span>修正理由: 曖昧表現の明確化 <s style="background:yellow;color:red">まちまち</s> → 異なる動き</span>)

        -「行って来い」の表現

        文脈に応じて、「上昇（下落）したのち下落（上昇）」のように明確にしてください。
        Exsample:

        Input: 相場は行って来いの展開となった
        Output: 相場は上昇したのち下落する展開となった
        
        変換語「行って来い」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        
        修正理由: 表現の明確化
        <span style="color:red;">行って来い</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Term</s> → 行って来い</span>)

        -「横ばい」表現の適正使用

        小幅な変動であれば「横ばい」を使用可能。
        大きな変動の末に同水準で終了した場合は、「ほぼ変わらず」「同程度となる」などに修正。
        
        変換語「横ばい」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 用語の適正使用
        <span style="color:red;">横ばい</span> (<span>修正理由: 用語の適正使用 <s style="background:yellow;color:red">横ばい</s> → ほぼ変わらず</span>)

        -「（割安に）放置」表現の修正

        「割安感のある」など、より適切な表現に修正してください。

        Exsample:
        Input: 株価は割安に放置された
        Output: 株価には割安感がある状態が続いた

        変換語「（割安に）放置」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 表現の明確化と客観性の向上
        <span style="color:red;">（割安に）放置</span> (<span>修正理由: 表現の明確化と客観性の向上 <s style="background:yellow;color:red"（割安に）放置</s> → 割安感のある</span>)

        """

        # ChatCompletion Call
        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=32768,
            temperature=0,
            seed=42  # 재현 가능한 결과를 위해 seed 설정
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)

        #-----------------------------------debug start

        # 틀린 부분 찾기
        corrections = find_corrections(re_answer)

        #-----------------------------------debug end
        # 엑셀 처리
        if excel_base64:
            try:
                excel_bytes_decoding = base64.b64decode(excel_base64)
                modified_bytes = correct_text_box_in_excel(excel_bytes_decoding,resutlmap)


                # 3) 수정된 XLSX를 반환(다운로드)
                return send_file(
                    io.BytesIO(modified_bytes),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name="annotated.xlsx"
                )
            except Exception as e:
                return jsonify({
                    "success": False,
                    "error": str(e)
                })

        if pdf_base64:
            pdf_bytes = base64.b64decode(pdf_base64)
            # 위치 정보만 찾아 corrections에 저장
            find_locations_in_pdf(pdf_bytes, corrections)

        # 수정된 텍스트와 코멘트를 JSON으로 반환
        return jsonify({
            "success": True,
            "corrections": corrections  # 틀린 부분과 코멘트
        })

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# db and save blob
PUBLIC_FUND_CONTAINER_NAME = "public_Fund"
PRIVATE_FUND_CONTAINER_NAME = "private_Fund"

public_container = get_db_connection(PUBLIC_FUND_CONTAINER_NAME)
private_container = get_db_connection(PRIVATE_FUND_CONTAINER_NAME)


def upload_to_azure_storage(pdf_bytes, file_name, fund_type):
        """Azure Blob Storage에 PDF 업로드"""
        container_name = PUBLIC_FUND_CONTAINER_NAME if fund_type == 'public' else PRIVATE_FUND_CONTAINER_NAME
        
        # 컨테이너 클라이언트 가져오기
        container_client = get_storage_container()

        try:
            blob_client = container_client.get_blob_client(file_name)
            blob_client.upload_blob(pdf_bytes, overwrite=True)
            logging.info(f"✅ Blob uploaded: {file_name} to {container_name}")
            return blob_client.url
        except Exception as e:
            logging.error(f"❌ Storage Upload error: {e}")
            return None


def save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type='', comment_type='',icon=''):
    """응답 데이터를 Cosmos DB에 저장"""
    # Cosmos DB 연결
    container = public_container if fund_type == 'public' else private_container

    item = {
        'id': file_name,  # 고유 ID로 파일 이름 사용
        'fileName': file_name,
        'result': response_data,  # GPT 응답
        'link': link_url,  # 파일 다운로드 링크 저장
        'updateTime': datetime.utcnow().isoformat(),  # 현재 시간
        'status': "issue", 
        'readStatus': "unread",
        'icon': icon,
    }
    if upload_type:
        item.update(upload_type=upload_type)
    if comment_type:
        item.update(comment_type=comment_type)


    try:
        existing_item = list(container.query_items(
                query="SELECT * FROM c WHERE c.id = @id",
                parameters=[{"name": "@id", "value": file_name}],
                enable_cross_partition_query=True
            ))

        if not existing_item:
                container.create_item(body=item)
                logging.info(f"✅ Cosmos DB は保存されています: {file_name}")
        else:
            # 기존 데이터가 존재하면 업데이트 (Upsert 사용 가능)
            existing_item[0].update(item)
            container.upsert_item(existing_item[0])

            logging.info(f"🔄 Cosmos DB 更新完了: {file_name}")
                
    except CosmosHttpResponseError as e:
        logging.error(f"❌Cosmos DB save error: {e}")


@app.route('/api/write_upload_save', methods=['POST'])
def write_upload_save():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ token upload")

        data = request.json
        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")
        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        icon = data.get("icon", "")

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)


        #-debug
        #---------EXCEL-----------
        if excel_base64:
            try:
                excel_bytes_decoding = base64.b64decode(excel_base64)
                modified_bytes = correct_text_box_in_excel(excel_bytes_decoding,resutlmap)

                response_data = {
                    "success": True,
                    "corrections": []
                }

                # Blob에 업로드
                link_url = upload_to_azure_storage(excel_bytes_decoding, file_name, fund_type)
                if not link_url:
                    return jsonify({"success": False, "error": "Blob upload failed"}), 500

                # Cosmos DB에 저장
                save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)
                if upload_type != "参照ファイル":
                    container = get_db_connection(FILE_MONITOR_ITEM)
                    container.upsert_item({"id": str(uuid.uuid4()), "file_name": file_name, "flag": "wait",
                                            "link": link_url, "fund_type": fund_type})

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

            except Exception as e:
                return jsonify({
                    "success": False,
                    "error": str(e)
                })

            # 3) 수정된 XLSX를 반환(다운로드)
            return jsonify({
                "success": True,
                "corrections": [],
                "code": 200,
            })
        # ---------PDF -----------
        if pdf_base64:
            try:
                pdf_bytes = base64.b64decode(pdf_base64)

                response_data = {
                    "corrections": []
                }

                # Blob에 업로드
                link_url = upload_to_azure_storage(pdf_bytes, file_name, fund_type)
                if not link_url:
                    return jsonify({"success": False, "error": "Blob upload failed"}), 500

                # Cosmos DB에 저장
                save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500


        # 수정된 텍스트와 코멘트를 JSON으로 반환
        return jsonify({
            "success": True,
            "corrections": [],
            "code": 200,
        })

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def apply_manual_corrections(text, correction_map):
    """
    특정 텍스트를 지정된 치환 텍스트로 변경
    :param text: GPT가 교정한 텍스트
    :param correction_map: { 기존 텍스트: 교정 텍스트 } 매핑
    :return: 치환이 완료된 최종 텍스트
    """
    if text in correction_map:
        # GPT 교정 결과로 치환
        result = correction_map[text]
    # for old_text, new_text in correction_map.items():
    #     if old_text in text:
    #         text = text.replace(old_text, new_text)
    return result

def gpt_correct_text(prompt):
    """
    GPT 모델을 사용하여 일본어 보고서 텍스트를 교정하고 corrected_map을 동적으로 업데이트합니다.
    :param prompt: 원본 텍스트
    :return: 교정된 텍스트 및 업데이트된 corrected_map
    """
    token = token_cache.get_token()
    openai.api_key = token
    print("✅ Token Update SUCCESS")
    
    if not prompt.strip():  # 빈 텍스트 방지
        return prompt
    hyogaiKanjiList = []
    
    corrected_text = detect_hyogai_kanji(prompt, hyogaiKanjiList)
    prompt_result = f"""
    You are a professional Japanese text proofreading assistant. Please carefully proofread the following Japanese text and provide corrections in a structured `corrected_map` format.

    **Text to Proofread:**
    - {prompt}

    **Proofreading Requirements:**
    1. **Check for typos and missing characters (誤字脱字がないこと):**
    - Ensure there are no spelling errors or missing characters in the content of the report.
    - If errors are found, add them to the `corrected_map` in the format: "incorrect": "correct".

    2. **Follow the Fund Manager Comment Terminology Guide (ファンドマネージャコメント用語集に沿った記載となっていること):**
    - **Consistent Terminology (表記の統一):**
        - Ensure the writing format of terms is consistent throughout the report.
    - **Prohibited Words and Phrases (禁止（NG）ワード及び文章の注意事項):**
        - Check if any prohibited words or phrases are used in the report and correct them as per the guidelines.
    - **Replaceable and Recommended Terms/Expressions (置き換えが必要な用語/表現、置き換えを推奨する用語/表現):**
        - If you find terms or expressions that need to be replaced, revise them according to the provided rules.
    - **Use of Hiragana (ひらがなを表記するもの):**
        - Ensure the report follows the rules for hiragana notation, replacing content that does not conform to commonly used kanji.
    - **Kana Notation for Non-Standard Kanji (一部かな書き等で表記するもの):**
        - Ensure non-standard kanji are replaced with kana as the standard writing format.
    - **Correct Usage of Okurigana (一般的な送り仮名など):**
        - Ensure the correct usage of okurigana is applied.
    - **English Abbreviations, Loanwords, and Technical Terms (英略語、外来語、専門用語など):**
        - Check if English abbreviations, loanwords, and technical terms are expressed correctly.
    - **Identify and mark any 常用外漢字 (Hyōgai kanji):**
        - Identify and mark any 常用外漢字 (Hyōgai kanji) in the following text.
        - 常用外漢字 refers to Chinese characters that are not included in the [常用漢字表 (Jōyō kanji list)](https://ja.wikipedia.org/wiki/常用漢字), the official list of commonly used kanji in Japan.
        - For any 常用外漢字 identified, mark the character with (常用外漢字) next to it.

        1. 入力された全文（Report Content to Proofread）を **一文字ずつ** 走査してください（単語単位ではなく文字単位の照合です）。
2. 各文字を、指定された hyogaiKanjiList の文字と **完全一致** で比較してください。
3. 一致する文字がある場合、その文字を「常用外漢字」として検出してください。
4. 一致しない文字は、常用漢字として無視してください（誤検出を避けるため）。
5. 検出された常用外漢字は、以下のフォーマットで注釈をつけて表示してください。

---

【注釈フォーマット】

次のように、元の漢字に <s> タグと背景色を付け、読みまたは代替語を赤字で示し、その後に理由を添えてください。

例:
<span style="color:red;">ぜい</span> (<span>修正理由: 常用外漢字の使用 <s style="background:yellow;color:red">脆</s> → ぜい</span>)

---

【Report Content to Proofread】:
{corrected_text}

**1: Typographical Errors (脱字・誤字) Detection**
            -Detect any missing characters (脱字) or misused characters (誤字) that cause unnatural expressions or misinterpretation.

            **Proofreading Requirements**:
            - Detect and correct all genuine missing characters (脱字) or misused characters (誤字) that cause grammatical errors or change the intended meaning.
            - Always detect and correct any incorrect conjugations, misused readings, or wrong kanji/verb usage, even if they superficially look natural.
            - Do not point out stylistic variations, natural auxiliary expressions, or acceptable conjugations unless they are grammatically incorrect.
            - Confirm that each kanji matches the intended meaning precisely.
            - Detect cases where non-verb terms are incorrectly used as if they were verbs.

            - ”と”を脱字しました:
                -Example:
                input: 月間ではほぼ変わらずなりました。
                output:
                <span style="color:red;">月間ではほぼ変わらずなりました。</span> (<span>修正理由: 誤字 <s style="background:yellow;color:red">月間ではほぼ変わらずなりました。</s> → 月間ではほぼ変わらずとなりました。</span>)
            - The kanji '剤' was incorrectly used instead of '済', resulting in a wrong word formation.
                -Example:
                input: 経剤成長
                output:
                <span style="color:red;">経済成長</span> (<span>修正理由: 誤字 <s style="background:yellow;color:red">経剤成長</s> → 経済成長</span>)
            - The verb "遊ぶ" was incorrectly conjugated into a non-existent form "あそぼれる" instead of the correct passive form "あそばれる".
                -Example:
                input: あそぼれますか。
                output:
                <span style="color:red;">あそばれますか。</span> (<span>修正理由: 動詞活用の誤り <s style="background:yellow;color:red">あそぼれますか。</s> → あそばれますか。</span>)
            - "と"を省略したら、「アンダーウェイト」は名詞であり、動詞のように「〜した」と活用するのは文法的に誤りです。
                -Example:
                input: アンダーウェイト（参考指数と比べ低めの投資比率）した
                output:
                <span style="color:red;">アンダーウェイト（参考指数と比べ低めの投資比率）とした</span> (<span>修正理由: 動詞活用の誤り <s style="background:yellow;color:red">アンダーウェイト（参考指数と比べ低めの投資比率）した</s> → アンダーウェイト（参考指数と比べ低めの投資比率）とした</span>)


            **correct Example*:
            - "取り組みし"は自然な連用形表現のため、修正不要'
                -Example:
                input: グローバルで事業を展開する
                output:
                <span style="color:red;">グローバルに事業を展開する</span> (<span>修正理由: 格助詞の誤用 <s style="background:yellow;color:red">グローバルで事業を展開する</s> → グローバルに事業を展開する</span>)

        **2: Punctuation (句読点) Usage Check**
            -Detect missing, excessive, or incorrect use of punctuation marks (、。).

            **Proofreading Requirements**:
            -Ensure sentences correctly end with「。」where appropriate.
            -Avoid redundant commas「、」in unnatural positions.
            -Maintain standard business writing style.

            -Example:
            input: 収益見通しが期待できる企業を中心に投資を行なう方針です
            output:
            <span style="color:red;">収益見通しが期待できる企業を中心に投資を行なう方針です</span> (<span>修正理由: 文末句点の欠如 <s style="background:yellow;color:red"収益見通しが期待できる企業を中心に投資を行なう方針です</s> → 収益見通しが期待できる企業を中心に投資を行なう方針です。</span>)

        **3: Unnatural Spaces (不自然な空白) Detection**
            -Detect unnecessary half-width or full-width spaces within sentences.

            **Proofreading Requirements**:
            -Remove any redundant spaces between words or inside terms.
            -Confirm that spacing follows standard Japanese document conventions.

            -Example:
            input: 送 配電設備
            output:
            <span style="color:red;">送配電設備</span> (<span>修正理由: 不要スペース削除 <s style="background:yellow;color:red">送 配電設備</s> → 送配電設備</span>)

            -Example:
            input: マイクロコントローラーや 関連の複合信号製品
            output:
            <span style="color:red;">マイクロコントローラーや関連の複合信号製品</span> (<span>修正理由: 不要スペース削除 <s style="background:yellow;color:red">マイクロコントローラーや 関連の複合信号製品</s> → マイクロコントローラーや関連の複合信号製品</span>)


        **4: Omission or Misuse of Particles (助詞の省略・誤用) Detection**
            - Detect omissions and misuses of grammatical particles (助詞), especially「の」「を」「に」, that lead to structurally incorrect or unnatural expressions.

            **Proofreading Requirements**:

            - Carefully examine whether all necessary particles—particularly「の」「を」「に」—are correctly used in every sentence.
            - Do not tolerate the omission of any structurally required particle, even if the sentence appears understandable or natural overall.
            - Focus on grammatical correctness, not perceived readability.
            - In long texts, perform sentence-by-sentence proofreading to ensure no required particle is missing at any position.
            - If a particle should be present according to standard Japanese grammar but is omitted, it must be explicitly identified and corrected.

            -Example:
            input: 欧州など市場調査開始して
            output:
            <span style="color:red;">欧州などの市場調査を開始して</span> (<span>修正理由: 連体修飾の助詞省略 <s style="background:yellow;color:red">欧州など市場調査開始して</s> → 欧州などの市場調査を開始して</span>)

            -Example:
            input: ECB（欧州中央銀行）など海外主要中銀による
            output:
            <span style="color:red;">ECB（欧州中央銀行）などの海外主要中銀による</span> (<span>修正理由: 所有格助詞「の」の省略 <s style="background:yellow;color:red">ECB（欧州中央銀行）など海外主要中銀による</s> → ECB（欧州中央銀行）などの海外主要中銀による</span>)

            -Example:
            input: 5000億円
            output:
            <span style="color:red;">5,000億円</span> (<span>修正理由: 金額カンマ区切り <s style="background:yellow;color:red">5000億円</s> → 5,000億円</span>)

        **5: Monetary Unit & Number Format (金額表記・数値フォーマット) Check**

            -Detect mistakes in number formatting, especially monetary values.
            -Proofreading Requirements:
            -Apply comma separator every three digits for numbers over 1,000.
            -Ensure currency units (円、兆円、億円) are correctly used.
            -Standardize half-width characters where needed.

            -Example:
            input: 対応には新たな技術開発や制度改革の必要性が指摘されています。
            output:
            <span style="color:red;">対応は新たな技術開発や制度改革の必要性が指摘されています。</span> (<span>修正理由: 格助詞「には」の誤用 <s style="background:yellow;color:red">対応には新たな技術開発や制度改革の必要性が指摘されています。</s> → 対応は新たな技術開発や制度改革の必要性が指摘されています。</span>)


            **Special Instructions**:
            - Always annotate all detected Hyōgai Kanji.
            - Never replace or modify the character unless explicitly instructed.

        **6: Detection of Misused Enumerative Particle「や」**
            **Proofreading Targets**:
            - Detect inappropriate use of the enumerative particle「や」when it connects elements with different grammatical structures.
            - The particle「や」must only be used to list **nouns or noun phrases** that are grammatically equivalent.
            - If the item following「や」is a **verb phrase**, **adverbial clause**, or a structurally different element, then「や」is incorrect.
            - In such cases, replace「や」with a comma「、」to properly separate clauses or adjust the sentence structure.


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).
        - **Task**: Header Date Format Validation & Correction  
        - **Target Area**: Date notation in parentheses following "今後運用方針 (Future Policy Decision Basis)"  
        ---
        ### Validation Requirements  
        1. **Full Format Compliance Check**:  
        - Must follow "YYYY年MM月DD日現在" (Year-Month-Day as of)  
        - **Year**: 4-digit number (e.g., 2024)  
        - **Month**: 2-digit (01-12, e.g., April → 04)  
        - **Day**: 2-digit (01-31, e.g., 5th → 05)  
        - **Suffix**: Must end with "現在" (as of)  

        2. **Common Error Pattern Detection**:  
        ❌ "1月0日" → Missing month leading zero + invalid day 0  
        ❌ "2024年4月1日" → Missing month leading zero (should be 04)  
        ❌ "2024年12月" → Missing day value  
        ❌ "2024-04-05現在" → Incorrect separator usage (hyphen/slash)  
        ---
        ### Correction Protocol  
        1. **Leading Zero Enforcement**  
        - Add leading zeros to single-digit months/days (4月 → 04月, 5日 → 05日)  

        2. **Day 0 Handling**  
        - Replace day 0 with YYYYMMDD Date Format  
        - Example: 2024年4月0日 → 2024年04月00日

        3. **Separator Standardization**  
        - Convert hyphens/slashes to CJK characters:  
            `2024/04/05` → `2024年04月05日`  


        4. **Consistency with Report Data Section (レポートのデータ部との整合性確認):**
        - Ensure the textual description in the report is completely consistent with the data section, without any logical or content-related discrepancies.

        5. **Eliminate language fluency(単語間の不要なスペース削除):**
        - Ensure that there are no extra spaces.
            -Example:
            input:景気浮揚が意 識されたことで
            output:景気浮揚が意識されたことで
        
        6.  **Layout and Formatting Rules (レイアウトに関する統一):**
            - **文頭の「○」印と一文字目の間隔を統一:**
                - When a sentence begins with the "○" symbol, ensure the spacing between the symbol and the first character is consistent across the document.
            - **文章の間隔の統一:**
                - If a sentence begins with "○", ensure that the spacing within the frame remains consistent.
            - **上位10銘柄 コメント欄について、枠内に適切に収まっているかチェック:**
                - If the stock commentary contains a large amount of text, confirm whether it fits within the designated frame. 
                - If the ranking changes in the following month, adjust the frame accordingly.
                - **Check point**
                    1. **文字数制限内に収まっているか？**
                    - 1枠あたりの最大文字数を超えていないか？
                    - 適切な行数で収まっているか？

                    2. **次月の順位変動に伴う枠調整の必要性**
                    - 順位が変更されると枠調整が必要なため、調整が必要な箇所を特定

                    3. **枠内に収まらない場合の修正提案**
                    - 必要に応じて、短縮表現や不要な情報の削除を提案
                    - 重要な情報を損なわずに適切にリライト

                    output Format:
                    - **コメントの枠超過チェック**
                    - (枠超過しているか: はい / いいえ)
                    - (超過している場合、オーバーした文字数)

                    - **順位変動による枠調整の必要性**
                    - (調整が必要なコメントリスト)

                    - **修正提案**
                    - (枠内に収めるための修正後のコメント)

            **Standardized Notation (表記の統一):**
            - **基準価額の騰落率:**
                - When there are three decimal places, round off using the round-half-up method to the second decimal place. If there are only two decimal places, keep the value unchanged.

            - **％（パーセント）、カタカナ:**
                - **半角カタカナ → 全角カタカナ**（例:「ｶﾀｶﾅ」→「カタカナ」）
                - **半角記号 → 全角記号**（例:「%」→「％」、「@」→「＠」）
                    Example:
                        input: % ｶﾀｶﾅ 
                        output: ％ カタカナ

            - **数字、アルファベット、「＋」・「－」:**
                - **全角数字・アルファベット → 半角数字・アルファベット**（例:「１２３」→「123」、「ＡＢＣ」→「ABC」）
                - **全角「＋」「－」 → 半角「+」「-」**（例:「＋－」→「+-」
                    Example:
                        input: １２３ ＡＢＣ ｱｲｳ ＋－
                        output: 123 ABC アイウ +-

            - **スペースは変更なし**  

            - **「※」の使用:**
                - 「※」は可能であれば **上付き文字（superscript）※** に変換してください。
                - 出力形式の例:
                - 「重要事項※」 → 「重要事項<sup>※</sup>」

            - **（カッコ書き）:**
                - Parenthetical notes should only be included in their first occurrence in a comment.
                以下の日本語テキストにおいて、カッコ書き（bracket "（ ）"）が適切に使用されているかをチェックしてください。

                **Check point**
                    1. **カッコ書きは、コメントの初出のみに記載されているか？**
                    - 同じカッコ書きが2回以上登場していないか？
                    - 初出ページ以降のコメントにカッコ書きが重複して記載されていないか？

                    2. **ディスクロのページ番号順に従ってルールを適用**
                    - シートの順番ではなく、実際のページ番号を基準にする。

                    3. **例外処理**
                    - 「一部例外ファンドあり」とあるため、例外的にカッコ書きが複数回登場するケースを考慮する。
                    - 例外として認められるケースを判断し、適切に指摘。

                    output Format:
                    - **カッコ書きの初出リスト**（どのページに最初に登場したか）
                    - **重複チェック結果**（どのページで二重記載されているか）
                    - **修正提案**（どのページのカッコ書きを削除すべきか）
                    - **例外ファンドが適用される場合、補足情報**

            - **会計期間の表記:**
                - The use of "～" is prohibited; always use "-".
                - Example: 6～8月期（×） → 6-8月期（○）
            - **年度表記:**
                - Use four-digit notation for years.
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: 22年（×） → 2022年（○）
            - **レンジの表記:**
                - Always append "%" when indicating a range.
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: -1～0.5%（×） → -1%～0.5%（○）
            - **投資環境の記述:**
                **「先月の投資環境」**の部分で「先月末」の記述が含まれる場合、「前月末」に変更してください。
                Example:
                修正前: 先月末の市場動向を分析すると…
                修正後: 前月末の市場動向を分析すると…

            - **通貨表記の統一:**
                - Standardize currency notation across the document.
                - 日本円は「JPY」または「¥」で統一 円.
                - Exsample: 
                input: ¥100 or JPY 100
                output: 100 円

            **Preferred and Recommended Terminology (置き換えが必要な用語/表現):**
            - **第1四半期:**
                - Ensure the period is clearly stated.
                - Example: 18年第4四半期（×） → 2018年10-12月期（○）
            - **約○％程度:**
                - Do not use "約" (approximately) and "程度" (extent) together. Choose either one.
                - Example: 約○％程度（×） → 約○％ or ○％程度（○）
            - **大手企業表記の明確化**  
            - **「○○大手」** が含まれる場合、文中の **会社名を抽出** し、  
                **「大手○○会社/企業」** の形式に修正する。  
            - **入力例:**  
                - 「大手メーカー/会社/企業」  
                - **出力:** 「大手不動産会社、大手半導体メーカー」  
            - **The actual company name must be found and converted in the article


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).
        ---

        ### **Correction Rules:**
        1. **Only output the corrected_map dictionary. No explanations or extra text.**
        2. **Only incorrect words and their corrected versions should be included.**
        3. **Do not include full sentence corrections.**
        4. **Ensure the corrected_map output is in valid Python dictionary format.**
        5. **Return only the following structure:**
        
        **Output Format:**
        {{
            "incorrect1": "corrected1",
            "incorrect2": "corrected2"
        }}

        """

    
    # ChatCompletion Call
    response = openai.ChatCompletion.create(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "You are a professional Japanese text proofreading assistant."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=32768,
        temperature=0,
        seed=42  # 재현 가능한 결과를 위해 seed 설정
    )

    
    # 응답 처리
    try:
        answer = response['choices'][0]['message']['content']
        corrected_map = parse_gpt_response(answer)
        
        # 전체 교정문 생성
        full_corrected = prompt
        for k, v in corrected_map.items():
            full_corrected = full_corrected.replace(k, v)

        # 동적 변경 사항 추가
        dynamic_corrections = detect_corrections(prompt, full_corrected)
        corrected_map.update(dynamic_corrections)

        return {k: v for k, v in corrected_map.items() if k and v and k != v}

    except Exception as e:
        print(f"처리 실패: {e}")
        return {}

def correct_text_box_in_excel(input_bytes,corrected_map):
    """
    :param input_bytes: 업로드된 엑셀(xlsx) 바이너리
    :param corrected_map: { 기존텍스트: 교정텍스트 } 형태의 매핑 (예: GPT 결과)
    :return: 수정된 엑셀 바이너리(bytes)
    """
    # 1) 압축 해제용 임시 폴더(또는 메모리상 in-memory zip)
    in_memory_zip = zipfile.ZipFile(io.BytesIO(input_bytes), 'r')
    
    # 새 ZIP(수정본)을 만들기 위한 BytesIO
    output_buffer = io.BytesIO()
    new_zip = zipfile.ZipFile(output_buffer, 'w', zipfile.ZIP_DEFLATED)

    # 2) 모든 파일 반복
    for item in in_memory_zip.infolist():
        file_data = in_memory_zip.read(item.filename)

        # 3) drawingN.xml인지 체크
        if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
            # 도형/텍스트박스 XML일 가능성이 있음
            try:
                tree = ET.fromstring(file_data)
                # 네임스페이스가 있는 경우 추출
                ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

                # 4) 모든 <a:t> 태그 찾기
                for t_element in tree.findall(".//a:t", ns):
                    original_text = t_element.text
                #----------------------------------------------------------------
                    # if original_text:  # None 체크
                    #     original_text_gpt = gpt_correct_text(original_text)  # GPT를 사용하여 텍스트 교정

                    #     if original_text_gpt and original_text_gpt.strip() in corrected_map:  # 변환 가능 여부 확인
                    #         t_element.text = corrected_map[original_text_gpt.strip()]  # 변환된 값으로 변경
                #----------------------------------------------------------------
                    # resultMap = gpt_correct_text(original_text)

                    if original_text in corrected_map:
                        # GPT 교정 결과로 치환
                        t_element.text = corrected_map[original_text]
                #----------------------------------------------------------------

                # 수정된 내용을 다시 XML로 직렬화
                file_data = ET.tostring(tree, encoding='utf-8', standalone=False)
                
            except Exception as e:
                print(f"Warning: Parsing {item.filename} failed - {e}")

        # 5) 새로운 zip에 추가
        new_zip.writestr(item, file_data)

    # 기존 zip 마무리
    in_memory_zip.close()
    # 새 zip 마무리
    new_zip.close()
    output_buffer.seek(0)
    return output_buffer.getvalue()

# Excel read -for debug
@app.route("/api/excel_upload", methods=["POST"])
def excel_upload():
    file = request.files["file"]  # XLSX 업로드
    original_bytes = file.read()

    # 1) drawingN.xml에서 텍스트 추출 → GPT 교정 로직
    #   (간단히 "전체 텍스트를 하나로 모아서 GPT에 보낸다"거나,
    #    "문단별로 나눈다"등등 필요에 따라 구현)

    # 예: 특정 텍스트를 "수정했습니다."로 치환
    corrected_map = {
        "地政学リスク": "地政学的リスク"
    }

    # 2) 수정
    modified_bytes = correct_text_box_in_excel(original_bytes, corrected_map)

    # 3) 수정된 XLSX를 반환(다운로드)
    return send_file(
        io.BytesIO(modified_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="annotated.xlsx"
    )

# T-STARヘルプ API --for debug
@app.route('/api/prompt_upload', methods=['POST'])
def prompt_upload():
    try:
        data = request.json
        prompt = data.get("input", "")
        original_text = data.get("original_text", "")

        if not prompt:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        prompt_result = f"""
        Please analyze the provided {original_text} and generate results based on the specified {prompt}.

        **Requirements**:
        1. **Extract relevant information**:
        - Extract only the information that directly answers the {prompt}.
        2. **Process the content**:
        - Process the extracted information to provide a clear and concise response.
        3. **Output in Japanese**:
        - Provide the results in Japanese, strictly based on the {prompt}.
        - Do not include any unrelated information or additional explanations.

        **Output**:
        - The output must be accurate, concise, and fully aligned with the {prompt}.
        - Only provide the response in Japanese.

        **Example**:
        - If the {prompt} is "売上成長率を教えてください", the output should be:
        "2023年の売上成長率は15％です。"

        Ensure the output is accurate, concise, and aligned with the given {prompt} requirements.
        """

        # ChatCompletion Call
        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=32768,
            temperature=0.1,
            seed=42  # 재현 가능한 결과를 위해 seed 설정
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)

        # 수정된 텍스트와 코멘트를 JSON으로 반환
        return jsonify({
            "success": True,
            "original_text": prompt,  # 입력된 원본 텍스트
            "corrected_text": re_answer,  # GPT 모델의 처리 결과
            # "corrections": corrections  # 틀린 부분과 코멘트
        })

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

#------Auto app update API

@app.route('/api/auto_save_cosmos', methods=['POST'])
def auto_save_cosmos():
    try:
        # 요청 데이터 파싱
        data = request.json
        response_data = data['result']
        link_url = data['link']
        container_name = data['containerName']
        file_name_decoding = data['fileName']

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        # Cosmos DB 컨테이너 클라이언트 가져오기)
        container = get_db_connection(container_name)

        # 저장할 아이템 생성
        item = {
            'id': file_name,  # 파일명을 고유 ID로 사용
            'fileName': file_name,
            'result': response_data,
            'link': link_url,
            'updateTime': datetime.utcnow().isoformat(),  # 현재 시간
        }

        # 기존 항목 존재 여부 확인
        existing_item = list(container.query_items(
            query="SELECT * FROM c WHERE c.id = @id",
            parameters=[{"name": "@id", "value": file_name}],
            enable_cross_partition_query=True
        ))

        if not existing_item:
            # 새 항목 생성
            container.create_item(body=item)
            logging.info(f"✅ Cosmos DB Update Success: {file_name}")
        else:
            # 기존 항목 업데이트
            existing_id = existing_item[0]['id']
            item['id'] = existing_id  # 기존 ID 유지
            container.replace_item(item=existing_item[0], body=item)
            logging.info(f"🔄 Cosmos DB update success: {file_name}")

        return jsonify({"success": True, "message": "Data Update Success"}), 200

    except CosmosHttpResponseError as e:
        logging.error(f"❌ Cosmos DB Save error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    except Exception as e:
        logging.error(f"❌ API Save error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    
#----auto app save to blob
@app.route('/api/auto_save_blob', methods=['POST'])
def auto_save_blob():
    try:
        # 파일 가져오기
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "no find file."}), 400

        file = request.files['file']
        blob_name = file.filename
        
        # 컨테이너 클라이언트 가져오기
        container_client = get_storage_container()

        # Blob(파일) 클라이언트 생성
        blob_client = container_client.get_blob_client(blob_name)

        # PDF 파일 업로드
        blob_client.upload_blob(file, overwrite=True)

        # 업로드된 파일의 URL 반환
        file_url = blob_client.url
        logging.info(f"✅ Azure Blob Storage Update Success: {blob_name}")

        return jsonify({"success": True, "url": file_url}), 200

    except Exception as e:
        logging.error(f"❌ Azure Blob Storage update error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

#----auto app save log 
@app.route('/api/auto_save_log_cosmos', methods=['POST','PUT'])
def auto_save_log_cosmos():
    """로그를 Cosmos DB에 저장하는 API"""
    try:
        # Cosmos DB 컨테이너 클라이언트 가져오기
        container = get_db_connection(APPLOG_CONTAINER_NAME)

        # 요청 본문에서 로그 데이터 가져오기
        log_data = request.json
        log_by_date = log_data.get("logs", {})

        # ✅ Cosmos DB에 저장
        for log_id, logs in log_by_date.items():
            existing_logs = list(container.query_items(
                query="SELECT * FROM c WHERE c.id = @log_id",
                parameters=[{"name": "@log_id", "value": log_id}],
                enable_cross_partition_query=True
            ))

            if existing_logs:
                # 기존 로그 문서 업데이트 (logEntries 리스트에 추가)
                existing_log = existing_logs[0]
                existing_log["logEntries"].extend(logs)
                existing_log["timestamp"] = datetime.utcnow().isoformat(),  # 현재 시간
                #update
                container.replace_item(item=existing_log["id"], body=existing_log)
                logging.info(f"🔄 SUCCESS: Update Log Success: {log_id}")
            else:
                # 새로운 로그 문서 생성
                log_data = {
                    "id": log_id,  # YYYYMMDD 형식의 ID
                    "logEntries": logs,  # 로그 리스트
                    "timestamp": datetime.utcnow().isoformat(),  # 현재 시간
                }
                #create
                container.create_item(body=log_data)
                logging.info(f"✅ SUCCESS: Save to Log Success: {log_id}")

        return jsonify({"code": 200, "message": "Logs saved successfully."}), 200

    except Exception as e:
        logging.error(f"❌ ERROR: Save Log Error: {e}")
        return jsonify({"code": 500, "message": "Error saving logs."}), 500


# integeration ruru
INTEGERATION_RURU_CONTAINER_NAME = 'integeration_ruru'
@app.route('/api/integeration_ruru_cosmos', methods=['POST'])
def integeration_ruru_cosmos():
    try:
        # 요청 데이터 파싱
        data = request.json

        base_month = data['Base_Month']
        fundType = data['fundType']
        fcode = data['Fcode']
        org_sheet_name = data['Org_SheetName']
        org_title = data['Org_Title']
        org_text = data['Org_Text']
        org_type = data['Org_Type']
        target_sheet_name = data['Target_SheetName']
        target_text = data['Target_Text']
        target_type = data['Target_Type']
        target_condition = data['Target_Condition']
        result = data['result']
        id = data['id']
        No = data['No']

        # Cosmos DB 컨테이너 클라이언트 가져오기
        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        # 중복 데이터 확인
        query = f"SELECT * FROM c WHERE c.Fcode = '{data['Fcode']}' AND c.Base_Month = '{data['Base_Month']}' AND c.fundType = '{data['fundType']}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        # Cosmos DB에 저장할 아이템 생성
        item = {
            "id": id,
            "No": No,
            "fundType": fundType,
            "Base_Month": base_month,
            "Fcode": fcode,
            "Org_SheetName": org_sheet_name,
            "Org_Title": org_title,
            "Org_Text": org_text,
            "Org_Type": org_type,
            "Target_SheetName": target_sheet_name,
            "Target_Text": target_text,
            "Target_Type": target_type,
            "Target_Condition": target_condition,
            "result": result,
            "updateTime": datetime.utcnow().isoformat(),  # 현재 시간
        }

        # 중복 데이터가 있으면 업데이트, 없으면 삽입
        if items:
            item["_etag"] = items[0]["_etag"]  # 기존 데이터의 etag를 사용하여 업데이트
            container.upsert_item(item)
            # items[0].update(item)
            # container.upsert_item(items[0])
            logging.info("✅ Data updated in Cosmos DB successfully.")
            return jsonify({"success": True, "message": "Data updated successfully."}), 200
        else:
            container.upsert_item(item)
            logging.info("✅ Data inserted into Cosmos DB successfully.")
            return jsonify({"success": True, "message": "Data inserted successfully."}), 200

    except Exception as e:
        logging.error(f"❌ Cosmos DB 저장 오류: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/integeration_ruru_cosmos', methods=['GET'])
def get_integeration_ruru_cosmos():
    # Cosmos DB 연결
    container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

    query = "SELECT * FROM c"
    users = list(container.query_items(query=query, enable_cross_partition_query=True))
    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

# --- ruru test api

@app.route('/api/ruru_search_db', methods=['POST'])
def ruru_search_db():
    try:
        # 요청 데이터 파싱
        data = request.json

        fcode = data.get('fcode')
        base_month = data.get('Base_Month')
        fund_type = data.get('fundType', 'private')  # 기본값으로 'private' 설정

        # Cosmos DB 컨테이너 클라이언트 가져오기)
        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        # DB에서 일치하는 데이터 찾기
        query = f"SELECT * FROM c WHERE c.Fcode = '{fcode}' AND c.Base_Month = '{base_month}' AND c.fundType = '{fund_type}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            # 일치하는 데이터의 result 출력
            # results = [{"id": item["id"], "result": item["result"],"Org_Text":item["Org_Text"],"Org_Type":item["Org_Type"],"Target_Condition":item["Target_Condition"]} for item in items]
            results = [item if item.get("flag") else {"id": item["id"], "result": item["result"],"Org_Text":item["Org_Text"],"Org_Type":item["Org_Type"],"Target_Condition":item["Target_Condition"]} for item in items]
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 404

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/refer_operate', methods=['GET'])
def get_rule():
    try:
        # 요청 데이터 파싱
        data = request.args
        flag = data.get('flag', "")
        fund_type = data.get('fundType', 'private')  # 기본값으로 'private' 설정

        # Cosmos DB 컨테이너 클라이언트 가져오기
        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        # DB에서 일치하는 데이터 찾기
        query = f"SELECT * FROM c WHERE c.flag = '{flag}' AND c.fundType = '{fund_type}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            items_map = list(map(lambda y: dict(filter(lambda x: not x[0].startswith("_"), y.items())), items))
            return jsonify({"success": True, "data": items_map}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 404

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/refer_operate', methods=['POST'])
def insert_rule():
    try:
        data = request.json

        base_month = data.get('Base_Month', '')
        fund_type = data.get('fundType', '')
        fcode = data.get('Fcode', '')
        org_sheet_name = data.get('Org_SheetName', '')
        org_title = data.get('Org_Title', '')
        org_text = data.get('Org_Text', '')
        org_type = data.get('Org_Type', '')
        target_sheet_name = data.get('Target_SheetName', '')
        target_title = data.get('Target_Title', '')
        target_text = data.get('Target_Text', '')
        target_type = data.get('Target_Type', '')
        target_condition = data.get('Target_Condition', '')
        target_consult = data.get('Target_Consult', '')
        id = str(uuid.uuid4())

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        item = {
            "id": id,
            "No": id,
            "fundType": fund_type,
            "Base_Month": base_month,
            "Fcode": fcode,
            "Org_SheetName": org_sheet_name,
            "Org_Title": org_title,
            "Org_Text": org_text,
            "Org_Type": org_type,
            "Target_SheetName": target_sheet_name,
            "Target_Text": target_text,
            "Target_Title": target_title,
            "Target_Type": target_type,
            "Target_Condition": target_condition,
            "Target_Consult": target_consult,
            "flag": "open",
            "updateTime": datetime.now().isoformat()
        }
        container.upsert_item(item)
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

async def get_original(input_data, org_text):
    dt = [
        "文章から原文に類似したテキストを抽出してください",
        "出力は以下のJSON形式でお願いします:",
        "- {'target': '[抽出されたテキスト:]'}",
        "- 類似度は最低でも85％を超える必要があります",
        "- 類似したものがない場合は、空の文字列を返してください",
        "- 類似したものが存在する場合は、最も類似度の高いものを抽出してください",

        f"原文:{org_text}\n文章:{input_data}"
    ]
    input_data = "\n".join(dt)

    question = [
        {"role": "system", "content": "あなたはテキスト抽出アシスタントです"},
        {"role": "user", "content": input_data}
    ]
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=question,
        max_tokens=32768,
        temperature=0,
        seed=42
    )
    answer = response['choices'][0]['message']['content'].strip().strip().replace("`", "").replace("json", "", 1)
    if answer:
        parsed_data = ast.literal_eval(answer)
        similar_content = parsed_data.get("target")
        if similar_content:
            score = SequenceMatcher(None, org_text, similar_content).ratio()
            if score > 0.85:
                return similar_content
        return ""

LOCAL_LINK = "local_link"
@app.route('/api/getaths', methods=['GET'])
def get_local_link():
    try:
        container = get_db_connection(LOCAL_LINK)
        log_data = list(container.query_items(
            query=f"SELECT * FROM c",
            enable_cross_partition_query=True
        ))
        log_map = list(map(lambda y: dict(filter(lambda x: x[1] and not x[0].startswith("_"), y.items())), log_data))
        return jsonify({"success": True, "data": log_map}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 404


@app.route('/api/saveaths', methods=['POST'])
def save_local_link():
    try:
        data = request.json
        commonComment = data.get("commonComment", "")
        individualCheckPath = data.get("individualCheckPath", "")
        individualComment = data.get("individualComment", "")
        individualExcelPath = data.get("individualExcelPath", "")
        individualPdfPath = data.get("individualPdfPath", "")
        meigaramaster = data.get("meigaramaster", "")
        reportData = data.get("reportData", "")
        simu = data.get("simu", "")
        fund_type = data.get("fund_type", "")
        container = get_db_connection(LOCAL_LINK)
        link_data = list(container.query_items(
           query=f"SELECT * FROM c WHERE c.fund_type='{fund_type}'",
            enable_cross_partition_query=True
        ))
        update_data = dict(
                fund_type=fund_type,
                commonComment=commonComment,
                individualCheckPath=individualCheckPath,
                individualComment=individualComment,
                individualExcelPath=individualExcelPath,
                individualPdfPath=individualPdfPath,
                meigaramaster=meigaramaster,
                reportData=reportData,
                simu=simu
        )

        if not link_data:
            update_data.update(id=str(uuid.uuid4()))
            container.upsert_item(update_data)
        else:
            link_data[0].update(update_data)
            container.upsert_item(link_data[0])
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 404


@app.route('/api/log_operate')
def get_log():
    try:
        # 1) page, size 파라미터 읽기 (기본값 page=1, size=15)
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 15))
        file_name = request.args.get('fileName', "")
        log_controller = get_db_connection(LOG_RECORD_CONTAINER_NAME)
        offset = (page - 1) * size
        if file_name:
            file_query = f"SELECT * FROM c WHERE CONTAINS(c.fileName, '{file_name}') OFFSET {offset} LIMIT {size}"
            total_file = list(log_controller.query_items(
                query=file_query,
                enable_cross_partition_query=True
            ))
            name_count = f"SELECT VALUE COUNT(1) FROM c WHERE CONTAINS(c.fileName, '{file_name}')"
            count_result = list(log_controller.query_items(
                query=name_count,
                enable_cross_partition_query=True
            ))[0]
            return jsonify({
                "success": True,
                "data": total_file,
                "total": count_result

            }), 200
        # 2) 전체 카운트 (페이지네이션 위해)
        count_query = "SELECT VALUE COUNT(1) FROM c"
        total_count = list(log_controller.query_items(
            query=count_query,
            enable_cross_partition_query=True
        ))[0]

        # 3) 페이지네이션 쿼리 (ORDER BY DESC, OFFSET, LIMIT)
        query = f"""
                SELECT * FROM c
                ORDER BY c.created_at DESC
                OFFSET {offset} LIMIT {size}
                """
        log_data = list(log_controller.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        log_map = list(map(lambda y: dict(filter(lambda x: x[1] and not x[0].startswith("_"), y.items())), log_data))

        return jsonify({
            "success": True,
            "data": log_map,
            "total": total_count
        }), 200
    
        # return jsonify({"success": True, "data": log_map}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/check_file', methods=['POST'])
def check_file_statue():
    try:
        data = request.json
        file_name = data.get("file_name")
        fund_type = data.get("fund_type")
        comment_type = data.get("comment_type")
        upload_type = data.get("upload_type", "")
        container = get_db_connection(FILE_MONITOR_ITEM)
        file_data = list(container.query_items(
            query=f"SELECT * FROM u WHERE u.file_name = '{file_name}'",
            enable_cross_partition_query=True
        ))
        if file_data:
            file_info = file_data[0]
            if file_info.get("flag") == "success":
                pdf_name = re.sub(r"\.(xlsx|xlsm|xls)", ".pdf", file_name)
                result = {
                    "corrections": file_info.get("corrections", [])
                }
                is_url = file_info.get("link", "")
                link_url = re.sub(r"\.(xlsx|xlsm|xls)", ".pdf", is_url)
                save_to_cosmos(pdf_name, result, link_url, fund_type, comment_type=comment_type, upload_type=upload_type)
                return jsonify({"success": True}), 200
        return jsonify({"success": False}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/file_status_update', methods=['POST'])
def file_update():
    try:
        data = request.json
        id = data.get("id", "")
        flag = data.get("flag", "")
        file_name = data.get("file_name", "")
        link_url = data.get("link", "")
        error_space = data.get("error_space", "")
        if id and flag and file_name:
            container = get_db_connection(FILE_MONITOR_ITEM)
            corrections = list(map(lambda x: dict(
                check_point=x.get("original_text"),
                original_text=x.get("original_text"),
                comment=x.get("original_text"),
                intgr=False,
                page=0,
                reason_type=x.get("reason_type"),
                locations=[{"x0": 0, "x1": 0, "y0": 0, "y1": 0}]
                ), error_space))
            container.upsert_item({"id": id, "flag": flag, "file_name": file_name, "link": link_url, "corrections": corrections})
            return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/file_status_search', methods=['GET'])
def file_search():
    try:
        container = get_db_connection(FILE_MONITOR_ITEM)
        file_data = list(container.query_items(
            query=f"SELECT * FROM u WHERE u.flag = 'wait'",
            enable_cross_partition_query=True
        ))
        if file_data:
            results = []
            for file_info in file_data:
                results.append(file_info)
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "data": []}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/ruru_ask_gpt_enhance', methods=['POST'])
def integrate_enhance():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ 토큰 업데이트 완료")

        data = request.json
        _content = data.get("input", "")
        condition = data.get("Target_Condition", "")
        category = data.get("Org_Type", "")
        consult = data.get("Target_Consult", "")
        base_month = data.get("Base_month", "")
        icon = data.get("icon", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")

        org_text = data.get("Org_Text", "")

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        content = loop.run_until_complete(get_original(_content, org_text))
        
        # content = get_original(_content, org_text)
        if not content:
            return jsonify({
                "success": True,
                "corrections": []  # 틀린 부분과 코멘트
            })



        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")

        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        if condition:
            result_temp = []
            table_list = condition.split("\n")
            for data in table_list:
                if data:
                    if category in ["比率", "配分"]:
                        re_num = re.search(r"([-\d. ]+)(%|％)", content)
                        if re_num:
                            num = re_num.groups()[0]
                            float_num = len(str(num).split(".")[1]) if "." in num else 0
                            old_data = pd.read_json(StringIO(data))
                            result_temp.append(old_data.applymap(
                                lambda x: (str(round(x * 100, float_num)) + "%" if float_num != 0 else str(
                                    int(round(x * 100, float_num))) + "%")
                                if not pd.isna(x) and isinstance(x, float) else x).to_json(force_ascii=False))
                        else:
                            result_temp.append(pd.read_json(StringIO(data)).to_json(force_ascii=False))
                    else:
                        result_temp.append(pd.read_json(StringIO(data)).to_json(force_ascii=False))
            if len(result_temp) > 1:
                result_data = "\n".join(result_temp)
            else:
                result_data = result_temp[0]
        else:
            result_data = ""

        input_list = [
            "以下の内容に基づいて、原文の記述が正しいかどうかを判断してください", "要件:",
            "- 提供された『参考データ』のみに基づいて判断してください、提供されていないデータについては判断する必要はありません",
            "- 最後に原文の記述が正しいかどうかを明確に判断し、文末に『OK』または『NG』を記載してください",
            f"- 現在の参考データは20{base_month[1:3]}年{base_month[3:]}月の参考データです",
            f"- 文中に『先月末』『前月末』『○月末』などの表現があっても、現在の参考データ（月）を基準として判断してください",
            f"原文の判断:'{content}'\n参考データ:\n'{result_data}'",
        ]

        if consult:
            input_list.insert(3, consult)
        input_data = "\n".join(input_list)
        question = [
            {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
            {"role": "user", "content": input_data}
        ]

        response = openai.ChatCompletion.create(
            deployment_id=_deployment_id,  # Deploy Name
            messages=question,
            max_tokens=16384,
            temperature=0,
            seed=42
        )
        answer = response['choices'][0]['message']['content'].strip()
        if answer:
            dt = [
                "以下の分析結果に基づき、原文中の誤りを抽出してください",
                "出力は以下のJSON形式でお願いします:",
                "- [{'original': '[原文中の誤っている部分:]', 'reason': '[理由:]'}]",
                "- 原文の末尾に「OK」がある場合は、空文字列を返してください",
                f"原文:'{content}'\n分析結果:'{answer}'"
            ]
            summarize = "\n".join(dt)
            _question = [
                {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
                {"role": "user", "content": summarize}
            ]
            _response = openai.ChatCompletion.create(
                deployment_id=_deployment_id,  # Deploy Name
                messages=_question,
                max_tokens=16384,
                temperature=0,
                seed=42
            )
            _answer = _response['choices'][0]['message']['content'].strip().replace("`", "").replace("json", "", 1)
            parsed_data = ast.literal_eval(_answer)
            format_data = []
            corrections = []
            for once in parsed_data:
                error_data = once.get("original", "")
                reason = once.get("reason", "")
                intgr_flag = True if error_data else False
                corrections.append({
                    "page": 0,  # 페이지 번호 (0부터 시작, 필요 시 수정)
                    "original_text": error_data,
                    "check_point": content,
                    "comment": reason,
                    "reason_type":reason, # for debug 62
                    "locations": [],  # 뒤에서 실제 PDF 위치(좌표)를 저장할 필드
                    "intgr": intgr_flag, # for debug 62
                })
            else:
                corrections.append({
                    "page": 0,  # 페이지 번호 (0부터 시작, 필요 시 수정)
                    "original_text": "",
                    "check_point": content,
                    "comment": "",
                    "reason_type": "",  # for debug 62
                    "locations": [],  # 뒤에서 실제 PDF 위치(좌표)를 저장할 필드
                    "intgr": False,  # for debug 66
                })

            #5/8 position check
            try:
                pdf_bytes = base64.b64decode(pdf_base64)
                # 위치 정보만 찾아 corrections에 저장
                find_locations_in_pdf(pdf_bytes, corrections)
                

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

            return jsonify({
                "success": True,
                "corrections": corrections  # 틀린 부분과 코멘트
            })
        else:
            return jsonify({
                "success": True,
                "corrections": [{
                    "page": 0,  # 페이지 번호 (0부터 시작, 필요 시 수정)
                    "original_text": "",
                    "check_point": content,
                    "comment": "",
                    "reason_type":"", # for debug 62
                    "locations": [],  # 뒤에서 실제 PDF 위치(좌표)를 저장할 필드
                    "intgr": True, # for debug 62
                }]  # 틀린 부분과 코멘트
            })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 200


@app.route('/api/ruru_ask_gpt', methods=['POST'])
def ruru_ask_gpt():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        _input = data.get("input", "")
        result = data.get("result", "")
        orgtext = data.get("Org_Text", "")
        OrgType = data.get("Org_Type", "")
        TargetCondition = data.get("Target_Condition", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        icon = data.get("icon", "")
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        input = loop.run_until_complete(get_original(_input, orgtext))

        # input = get_original(_input, orgtext)
        if not input:
            return jsonify({
                "success": True,
                "corrections": []  # 틀린 부분과 코멘트
            })


        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")
        icon = data.get("icon", "")

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400


        prompt_result = f"""
You are a professional Japanese financial report proofreader.  
Your task is to **semantically** validate whether the factual content in the `{input}` is consistent with the numerical and financial data provided in `{result}`.  
Only focus on **meaning-level discrepancies** — not textual similarity.

---

### ✅ Validation and Highlighting Rules:

1. Compare the values and facts in `{result}` with the claims made in `{input}`.
2. If a sentence or phrase in `{input}` contains a value (e.g., "騰落率", "ベンチマークに対して") that is **semantically consistent** with the value in `{result}`, do **not highlight it**.
3. If the value mentioned in `{input}` **differs from** the corresponding value in `{result}` (in meaning or number), highlight that specific phrase.
4. Use the following format for incorrect parts:
```html
<span style="background-color:#ace4e6;color:red;">[Wrong phrase]</span>  
(<span>提示: [Field or Reason] <s style="color:red">[Wrong value]</s> → [Correct value]</span>)

5. If before includes the same meaning or exact phrase as after, even partially, treat it as correct and do not output anything.
6. If all values are semantically correct, return nothing (empty).
7. Use {OrgType} and {TargetCondition} only as context for understanding — but your judgment should be based solely on {input} vs {result}.

Input:
    Input Text:
    {input}

    Original Type:
    {OrgType}

    Target Condition:
    {TargetCondition}

    Result:
    {result}

Return only HTML output. Do not explain or add comments.

        
        """  
        # ChatCompletion Call
        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."
                "This includes not only Japanese text but also English abbreviations (英略語), "
                "foreign terms (外来語),and specialized terminology (専門用語)."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=32768,
            temperature=0,
            seed=42  # 재현 가능한 결과를 위해 seed 설정
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)

        # add the write logic
        # 틀린 부분 찾기
        corrections = extract_corrections(re_answer,input)

        # 엑셀 처리
        if excel_base64:
            try:
                excel_bytes_decoding = base64.b64decode(excel_base64)
                modified_bytes = correct_text_box_in_excel(excel_bytes_decoding,resutlmap)

                # 3) 수정된 XLSX를 반환(다운로드)
                return send_file(
                    io.BytesIO(modified_bytes),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name="annotated.xlsx"
                )
            except Exception as e:
                return jsonify({
                    "success": False,
                    "error": str(e)
                })

        if pdf_base64:
            try:
                pdf_bytes = base64.b64decode(pdf_base64)
                # 위치 정보만 찾아 corrections에 저장
                find_locations_in_pdf(pdf_bytes, corrections)
                
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # 수정된 텍스트와 코멘트를 JSON으로 반환
        return jsonify({
            "success": True,
            "corrections": corrections  # 틀린 부분과 코멘트
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
# 611 opt - debug new prompt
def extract_text_from_base64_pdf(pdf_base64: bytes) -> list:
    # Base64 -> PDF bytes
    # pdf_bytes = base64.b64decode(pdf_base64)

    # 메모리에서 PDF 열기
    pdf_document = fitz.open(stream=pdf_base64, filetype="pdf")

    text_all = []
    keyword_pages = []
    page_list = []
    for page_num in range(pdf_document.page_count):
        page = pdf_document.load_page(page_num)

        full_text = page.get_text()

        keyword_pos = -1
        for keyword in ["組入銘柄解説", "組入銘柄","組入上位10銘柄の解説"]:
            keyword_pos = full_text.find(keyword)
            if keyword_pos != -1:
                keyword_pages.append(page_num)
                break

        if page_num in keyword_pages:
            # text_all.append(full_text)
            page_text = full_text
            

        else:
            blocks = page.get_text("blocks")  # (x0, y0, x1, y1, "text", block_no, block_type)
            blocks.sort(key=lambda b: b[1])
            page_text = "".join(block[4] for block in blocks)
            # text_all.append(page_text)

        # page_list.append(("".join(text_all), page_num))

        # ✅ 한 페이지 단위로만 저장
        page_list.append((page_text, page_num))
            
    return page_list


# add pre-half logic
half_to_full_map = {
    '%': '％',
    '@': '＠',
    '&': '＆',
    '!': '！',
    '?': '？',
    '#': '＃',
    '$': '＄',
    '(': '（',
    ')': '）',
    '+': '＋'
}
def convert_halfwidth_to_fullwidth_safely(text):
    # 이미 변환된 부분 (修正理由 포함) 보존
    protected_blocks = {}
    
    def protect_span(match):
        key = f"__PROTECT_{len(protected_blocks)}__"
        protected_blocks[key] = match.group(0)
        return key

    # ① 보호할 블록 감추기
    text = re.sub(r'<span[^>]*?>修正理由:.*?</span>\)', protect_span, text)

    # ② 남은 텍스트만 변환
    def replace_half(match):
        char = match.group(0)
        full = half_to_full_map[char]
        return (
            f'<span style="color:red;">{full}</span> '
            f'(<span>修正理由: 半角記号を全角に統一 '
            f'<s style="background:yellow;color:red">{char}</s> → {full}</span>)'
        )

    pattern = re.compile('|'.join(map(re.escape, half_to_full_map.keys())))
    text = pattern.sub(replace_half, text)

    # ③ 보호한 블록 되돌리기
    for key, val in protected_blocks.items():
        text = text.replace(key, val)

    return text

def get_num(num):
    if num:
        num_str = str(num)
        num_len = len(num_str)
        num_list = []
        for i in range(num_len, 0, -3):
            if i - 3 < 0:
                num_r = 0
            else:
                num_r = i - 3
            num_list.insert(0, num_str[num_r: i])
        return ",".join(num_list)
    return ""


# async call ,need FE promises
def opt_common(input, prompt_result, pdf_base64, pageNumber, re_list, word_list, rule_list, rule1_list, rule3_list):  
    # ChatCompletion Call
    response = openai.ChatCompletion.create(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "You are a Japanese text extraction tool capable of accurately extracting the required text."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=32768,
        temperature=0,
        seed=42  # 재현 가능한 결과를 위해 seed 설정
    )
    answer = response['choices'][0]['message']['content'].strip().replace("`", "").replace("json", "", 1).replace("\n", "")
    parsed_data = ast.literal_eval(answer)
    combine_corrections = []
    for data in parsed_data:
        _re_rule = ".{,2}"
        _original_re = regcheck.search(f"{_re_rule}{re.escape(data["original"])}{_re_rule}", input)
        if _original_re:
            _original_text = _original_re.group()
        else:
            _original_text = data["original"]
        combine_corrections.append({
            "page": pageNumber,
            "original_text": _original_text,  # 전체 입력
            "comment": f'{_original_text} → {data["correct"]}', 
            "reason_type": data["reason"],
            "check_point": _original_text,
            "locations": [],
            "intgr": False,  # for debug 62
        })
    if rule_list:
        for rule_result in rule_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": str(rule_result),  # 전체 입력
                "comment": f"{str(rule_result)} → 当月の投資配分",
                "reason_type": "誤字脱字",
                "check_point": str(rule_result),
                "locations": [],
                "intgr": False,  # for debug 62
            })

    if re_list:
        for re_result in re_list:
            correct = get_num(re_result)
            combine_corrections.append({
                "page": pageNumber,
                "original_text": str(re_result),  # 전체 입력
                "comment": correct,
                "reason_type": "数値千位逗号分隔修正",
                "check_point": str(re_result),
                "locations": [],
                "intgr": False,  # for debug 62
            })

    if rule1_list:
        for rule1_result in rule1_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": rule1_result,  # 전체 입력
                "comment": f"{rule1_result} →  ",
                "reason_type": "削除",
                "check_point": rule1_result,
                "locations": [],
                "intgr": False,  # for debug 62
            })

    if rule3_list:
        for rule3_result in rule3_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": rule3_result,  # 전체 입력
                "comment": f"{rule3_result} → {rule3_result[1:]}",
                "reason_type": "削除",
                "check_point": rule3_result,
                "locations": [],
                "intgr": False,  # for debug 62
            })


    if word_list:
        for word_result in word_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": word_result,  # 전체 입력
                "comment": f"{word_result} → 値上がりし",
                "reason_type": "動詞固定用法",
                "check_point": word_result,
                "locations": [],
                "intgr": False,  # for debug 62
            })


    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            # 위치 정보만 찾아 corrections에 저장
            find_locations_in_pdf(pdf_bytes, combine_corrections)
            
        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    # 수정된 텍스트와 코멘트를 JSON으로 반환
    return jsonify({
        "success": True,
        "corrections": combine_corrections,  # 틀린 부분과 코멘트
    })

async def opt_common_wording(file_name,fund_type,input,prompt_result,excel_base64,pdf_base64,resutlmap,upload_type,comment_type,icon,pageNumber):
    # ChatCompletion Call
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "あなたは曖昧な表現を定型語に変換する、厳格な金融校正AIです。出力形式・修正ルールはすべて厳守してください。"},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=32768,
        temperature=0,
        seed=42  # 재현 가능한 결과를 위해 seed 설정
    )
    answer = response['choices'][0]['message']['content'].strip()
    re_answer = remove_code_blocks(answer)

    # add the write logic
    # 틀린 부분 찾기
    corrections = find_corrections(re_answer,input,pageNumber)

    corrections_wording = find_corrections_wording(input,pageNumber)

    combine_corrections = corrections + corrections_wording

    # add half-full logic 603
    # test_ = convert_halfwidth_to_fullwidth_safely(corrections)

    # 엑셀 처리
    if excel_base64:
        try:
            excel_bytes_decoding = base64.b64decode(excel_base64)
            modified_bytes = correct_text_box_in_excel(excel_bytes_decoding,resutlmap)

            # 3) 수정된 XLSX를 반환(다운로드)
            return send_file(
                io.BytesIO(modified_bytes),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="annotated.xlsx"
            )
        except Exception as e:
            return jsonify({
                "success": False,
                "error": str(e)
            })


    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            # 위치 정보만 찾아 corrections에 저장
            find_locations_in_pdf(pdf_bytes, combine_corrections)
            
        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    # 수정된 텍스트와 코멘트를 JSON으로 반환
    return jsonify({
        "success": True,
        "corrections": combine_corrections,  # 틀린 부분과 코멘트
        "debug_re_answer":re_answer, #610 debug
    })

@app.route('/api/opt_typo', methods=['POST'])
def opt_typo():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input = data.get("input", "")

        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        icon = data.get("icon", "")
        pageNumber = data.get('pageNumber',0)

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        prompt_result = get_prompt("\"" + input + "\"")
        async def run_tasks():
            tasks = [handle_result(once) for once in prompt_result]
            return await asyncio.gather(*tasks)

        results = asyncio.run(run_tasks())
        sec_input = "\n".join(results)

        dt = [
            "以下の分析結果に基づき、原文中の誤りを抽出してください。",
            "- originalには必ず全文や長い文ではなく、**reason_typeで指摘されている最小限の誤りポイント（単語や助詞など）**のみを記載してください。",
            "- 1単語またはごく短いフレーズ単位でoriginalを抽出してください。",
            "- originalはreason_typeの説明に該当する部分のみを抽出してください（例：『など』の後に助詞『の』が必要→originalは必ず『など』）。",
            "出力は以下のJSON形式でお願いします:",
            "- [{'original': '[原文中の誤っている最小単位の部分]', 'correct': '[正しいテキスト]', 'reason': '[理由:]'}]",
            "- 分析結果に修正部分がある場合は、必ず空のリストを返さないでください。",
            "【例】",
            "reason_type: '年表記は4桁（西暦）に統一'",
            "原文: \"22年の経済成長率は-1～0.5の範囲で推移しました。\"",
            "出力例:",
            "[",
            "  {",
            "    \"original\": \"22年\",",
            "    \"correct\": \"2022年\",",
            "    \"reason\": \"年表記は4桁（西暦）に統一\"",
            "  }",
            "]",
            "reason_type: '例示『など』の後には助詞『の』が必要。『など海外主要中銀』は文法的に不自然なため。'",
            "原文: \"など海外主要中銀による\"",
            "出力例:",
            "[",
            "  {",
            "    \"original\": \"など\",",
            "    \"correct\": \"などの\",",
            "    \"reason\": \"例示『など』の後には助詞『の』が必要\"",
            "  }",
            "]",
            f"原文:'{input}'\n分析結果:'{sec_input}'"

        ]
        sec_prompt = "\n".join(dt)
        re_list = regcheck.findall(r"(\d{4,})[人種万円兆億]", input)
        word_list = regcheck.findall(r".{,2}値上がり(?!し).{,2}", input)
        rule_list = regcheck.findall(r"当月投資配分", input)
        rule1_list = regcheck.findall(r"【(先月の投資環境|先月の運用経過|今後の運用方針)】", input)
        rule3_list = regcheck.findall(r"-[\d.％]{4,6}下落", input)

        _content = opt_common(input, sec_prompt, pdf_base64, pageNumber, re_list, word_list, rule_list, rule1_list, rule3_list)
        return _content

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

async def handle_result(prompt_result):
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system",
            "content": "You are a professional Japanese business document proofreader specialized in financial and public disclosure materials."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=32768,
        temperature=0,
        seed=42  # 재현 가능한 결과를 위해 seed 설정
    )
    answer = response['choices'][0]['message']['content'].strip()
    return answer

def get_prompt(corrected):
    example_0 = "'original': '月間ではほぼ変わらずなりました。', 'correct': '月間ではほぼ変わらずとなりました。', 'reason': '誤字'"
    example_1 = "'original': '経剤成長', 'correct': '経済成長', 'reason': '誤字'"
    example_10 = "'original': '子供たちは公園で自由にあそぼれますか。', 'correct': '子供たちは公園で自由にあそばれますか。', 'reason': '動詞活用の誤り（「遊ばれる」→「遊ぼれる」）'"
    example_11 = "'original': '我々は新しいプロジェクトに取り組みし、成果を上げました。'"
    example_110 = "'original': 'セクター配分において特化型（物流施設）をアンダーウェイト（参考指数と比べ低めの投資比率）したことなどがプラスに寄与しました。', 'correct': 'セクター配分において特化型（物流施設）をアンダーウェイト（参考指数と比べ低めの投資比率）としたことなどがプラスに寄与しました。', 'reason': '動詞活用の誤り（「遊ばれる」→「遊ぼれる」）'"
    example_111 = "'original': '電子部品や通信機器などの製造・販売を行なうグローバルで事業を展開する電子モジュール・部品メーカー。', 'correct': '電子部品や通信機器などの製造・販売を行なうグローバルに事業を展開する電子モジュール・部品メーカー。', 'reason': 'グローバルは「に」を使用する'"
    example_2 = "'original': '今後はトランプ次期米大統領が掲げる減税や規制緩和の政策が米景気を押し上げることが、市場の下支えになると考えています。引き続き、FRBによる金融政策や新政権の政策により影響を受けるセクターなどを注視しながら、銘柄を選定して運用を行ないます', 'correct': '今後はトランプ次期米大統領が掲げる減税や規制緩和の政策が米景気を押し上げることが、市場の下支えになると考えています。引き続き、FRBによる金融政策や新政権の政策により影響を受けるセクターなどを注視しながら、銘柄を選定して運用を行ないます。', 'reason': '文法誤用'"
    example_4 = "'original': '半導体メーカー。マイクロコントローラーや 関連の複合信号製品', 'correct': '半導体メーカー。マイクロコントローラーや関連の複合信号製品', 'reason': '不要スペース削除'"
    example_5 = "'original': '東南アジアや欧州など市場調査開始して', 'correct': '東南アジアや欧州などの市場調査を開始して', 'reason': '欧州など“の”市場を省略不可'"
    example_6 = "'original': '東南アジアや欧州など市場調査開始して', 'correct': '東南アジアや欧州などの市場調査を開始して', 'reason': '市場調査”を”開始省略不可'"
    example_60 = "'original': 'ECB（欧州中央銀行）など海外主要中銀による', 'correct': 'ECB（欧州中央銀行）などの海外主要中銀による', 'reason': 'など”の”海外主要省略不可'"
    example_61 = "'original': '気候変動への対応には新たな技術開発や制度改革の必要性が指摘されています。', 'correct': '気候変動への対応は新たな技術開発や制度改革の必要性が指摘されています。', 'reason': '「には」は誤用のため、「に」を削除し、「は」だけ使用する'"
    example_62 = "'original': '中旬から下旬かけては', 'correct': '中旬から下旬にかけては', 'reason': '助詞「に」の脱落修正'"
    example_70 = "'original': '○月間の基準価額（分配金再投資）の騰落率は、毎月分配型が0.37％、年2回決算型は0.36％の上昇となり、参考指数の騰落率（0.58％の上昇）を下回りました。', 'correct': '○月間の基準価額（分配金再投資）の騰落率は、毎月分配型が0.37％の上昇、年2回決算型は0.36％の上昇となり、参考指数の騰落率（0.58％の上昇）を下回りました。', 'reason': 'Aが◯%、Bは△%の上昇の場合、「の上昇」がBだけにかかっていて、Aにもつけた方がわかりやすいため。'"
    prompt_list = [
        f"""
        **Typographical Errors (脱字・誤字) Detection**
        -Detect any missing characters (脱字) or misused characters (誤字) that cause unnatural expressions or misinterpretation.
        **Proofreading Requirements**：
        - Detect and correct all genuine missing characters (脱字) or misused characters (誤字) that cause grammatical errors or change the intended meaning.
        - Always detect and correct any incorrect conjugations, misused readings, or wrong kanji/verb usage, even if they superficially look natural.
        - Do not point out stylistic variations, natural auxiliary expressions, or acceptable conjugations unless they are grammatically incorrect.
        - Confirm that each kanji matches the intended meaning precisely.
        - Detect cases where non-verb terms are incorrectly used as if they were verbs.
        - Do **not** treat orthographic variants involving okurigana omission or abbreviation（e.g., 書き換え vs 書換え, 読み取る vs 読取る, 取り込む vs 取込）as typographical errors

        **missing Example*：
        {example_0}  ”と”を脱字しました
        {example_1}  The kanji '剤' was incorrectly used instead of '済', resulting in a wrong word formation.
        {example_10} The verb "遊ぶ" was incorrectly conjugated into a non-existent form "あそぼれる" instead of the correct passive form "あそばれる".
        {example_110}  "と"を省略したら、「アンダーウェイト」は名詞であり、動詞のように「〜した」と活用するのは文法的に誤りです。
        {example_111}
        **correct Example*：
        {example_11}
        "取り組みし"は自然な連用形表現のため、修正不要'
        {example_70}
        """,
        f"""
        **Punctuation (句読点) Usage Check**
        -Detect missing, excessive, or incorrect use of punctuation marks (、。).
        **Proofreading Requirements**：
        -Ensure sentences correctly end with「。」where appropriate.
        -Avoid redundant commas「、」in unnatural positions.
        -Maintain standard business writing style.
        **Example**：
        {example_2}
        """,
        f"""
        **Omission of Particles (助詞の省略・誤用) Detection**
        - Detect omissions and misuses of grammatical particles (助詞), especially「の」「を」「に」, that lead to structurally incorrect or unnatural expressions.
        **Proofreading Requirements**:
        - Carefully examine whether all necessary particles—particularly「の」「を」「に」—are correctly used in every sentence.
        - Do not tolerate the omission of any structurally required particle, even if the sentence appears understandable or natural overall.
        - Focus on grammatical correctness, not perceived readability.
        - In long texts, perform sentence-by-sentence proofreading to ensure no required particle is missing at any position.
        - If a particle should be present according to standard Japanese grammar but is omitted, it must be explicitly identified and corrected.

        **Example**：
        {example_61}
        {example_5}   
        {example_6}     
        {example_60}
        {example_62} 
        """,
        f"""
        **Monetary Unit(金額表記) Check**
        -Proofreading Requirements：
        -Ensure currency units (円、兆円、億円) are correctly used.
        """
    ]

    for target_prompt in prompt_list:
        if "助詞の省略" in target_prompt:
            special_word = "- **動詞の連用形や文中の接続助詞前の活用形は正しい表現として認め、文末形などへの変更を求めないこと。**"
        else:
            special_word = ""
        common_result = f"""
        You are a professional Japanese business document proofreader specialized in financial and public disclosure materials. 

        あなたは金融機関の対外発表文書に特化した、日本語ビジネス文書の校正専門家です。  
        必ず以下のポリシーを守って校正を行ってください：  
        文法的正確性だけでなく、ビジネス文書としての自然さ・厳密さを重視する。  
        金融機関向け公式発表として恥ずかしくない完璧な日本語表現を目指す。  
        曖昧な推測は禁止。必ず明確な根拠とルールに基づき指摘する。

        以下の校正基準を厳守すること：  
        - 修正を行った場合、必ず**文法的に成立する自然な文に仕上げる**こと。修正後に不自然または誤用となる表現は禁止。
        - 修正は**明示的にルール内で定義された語句・表現**に限ること。それ以外は**勝手に「よりよい表現」として書き換えない**。
        - **原文に明確な問題がない限り、修正を加えないこと。**
        {special_word}
        - 校正により変更された結果、文全体の意味や論理、文法が破綻しないよう、**必ず前後関係を考慮すること。**
        - 常用外漢字に関する内容は校正不要です。
        - 送り仮名の一致については校正不要です。
        - 回答は200字以内に制限してください。

        **Proofreading Targets：**
        {corrected}

        {target_prompt}

        **Please check each sentence **individually** for okurigana consistency.**
        """
        yield common_result



def detect_hyogai_kanji(input_text, hyogaiKanjiList):
    corrected_map = {}
    for char in input_text:
        if char in hyogaiKanjiList:
            # 常用外漢字の読みまたは代替語をここでは仮に「？」としています。
            # 実際には、文脈に応じて適切な読みや代替語を設定する必要があります。
            replacement = f"<span style=\"color:red;\">?</span> (<span>修正理由: 常用外漢字の使用 <s style=\"background:yellow;color:red\">{char}</s> → ?</span>)"
            corrected_map[char] = replacement
            input_text = input_text.replace(char, replacement) # 逐次的に置換

    return input_text

@app.route('/api/opt_kanji', methods=['POST'])
def opt_kanji():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input = data.get("input", "")

        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        tenbrend = data.get("tenbrend", [])
        icon = data.get("icon", "")
        pageNumber = data.get('pageNumber',0)

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400
        

        corrections = find_corrections_wording(input, pageNumber,tenbrend,fund_type)
        
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            find_locations_in_pdf(pdf_bytes, corrections)

        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    
        return jsonify({
                "success": True,
                "corrections": corrections  # 틀린 부분과 코멘트
            })


    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def loop_in_ruru(input):
    ruru_all =[
        {
            "category": "表記の統一 (Standardized Notation)",
            "rule_id": "1.1",
            "description": "基準価額の騰落率に関する表現の統一および数値の四捨五入を行うこと。指定された表現に厳密に従う。",
            "requirements": [
                {
                    "condition": "騰落率が 0.00％ / 0.0％ / 0％ の場合",
                    "correction": "『騰落率は変わらず』の代わりに、以下のいずれかに修正する:\n- 『基準価額（分配金再投資）は前月末から変わらず』\n- 『前月末と同程度』"
                },
                {
                    "condition": "騰落率の数値が小数第3位まである（例：0.546％）",
                    "correction": "小数第2位で四捨五入（round-half-up）し、『0.55%』のように修正する"
                },
                {
                    "condition": "ファンドとベンチマーク（参考指数）の騰落率を比較する場合",
                    "correction": "上記の四捨五入処理後の値で比較し、同じ場合は『騰落率は同程度となりました』と記述する"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "ファンドの騰落率は0.546％",
                    "output": "ファンドの騰落率は<span style=\"color:red;\">0.546％</span> (<span>修正理由: 四捨五入 <s style=\"background:yellow;color:red\">0.546％</s> →0.55%</span>)"
                },
                {
                    "input": "ベンチマークの騰落率は0.546％",
                    "output": "ベンチマークの騰落率は<span style=\"color:red;\">0.546％</span> (<span>修正理由: 四捨五入 <s style=\"background:yellow;color:red\">0.546％</s> →0.55%</span>)"
                },
                {
                    "input": "騰落率は同じでした。",
                    "output": "<span style=\"color:red;\">騰落率は同じでした。</span> (<span>修正理由: 騰落率が同じ <s style=\"background:yellow;color:red\">騰落率は同じでした。</s> →騰落率は同程度となりました。</span>)"
                },
                {
                    "input": "0.00％となりました",
                    "output": "<span style=\"color:red;\">0.00％となりました</span> (<span>修正理由: 表記の修正 <s style=\"background:yellow;color:red\">0.00％となりました</s> → 前月末から変わらず</span>)"
                },
                {
                    "input": "0.0％でした。",
                    "output": "<span style=\"color:red;\">0.0％でした。</span> (<span>修正理由: 表記の修正 <s style=\"background:yellow;color:red\">0.0％でした。</s> → 前月末から変わらず</span>)"
                }
            ]
        },
        {
            "category": "表記の統一 (Standardized Notation)",
            "rule_id": "1.3",
            "description": "年の記載が省略された日付表記（例：1月31日）に対して、文中に存在する最新の『XXXX年度』を基準に対応する西暦年（XXXX年）を補完する。",
            "requirements": [
                {
                "condition": "『XXXX年度』という記載が文中に存在し、かつ『○月○日』という形式で年が省略された日付がある場合",
                "correction": "該当する『○月○日』の前に『XXXX年』を補完し、『XXXX年○月○日』の形式に統一する（XXXXは直前の年度から抽出）"
                },
                {
                "condition": "既に『XXXX年○月○日』のように年が明記されている場合",
                "correction": "変更は行わない（既に正しい形式）"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                "input": "2024年度の予想経常利益は前年度比4.4％増、1月31日時点の見通しです。",
                "output": "2024年度の予想経常利益は前年度比4.4％増、<span style=\"color:red;\">1月31日</span> (<span>修正理由: 年が省略されていた日付に 2024年 を補完 <s style=\"background:yellow;color:red\">1月31日</s> → 2024年1月31日</span>)"
                },
                {
                "input": "2025年度は同5.8％増（同上）となることが予想されます。2月1日に発表予定。",
                "output": "2025年度は同5.8％増（同上）となることが予想されます。<span style=\"color:red;\">2月1日</span> (<span>修正理由: 年が省略されていた日付に 2025年 を補完 <s style=\"background:yellow;color:red\">2月1日</s> → 2025年2月1日</span>)"
                },
                {
                "input": "2023年度の実績は2023年12月31日時点の情報です。",
                "output": "2023年度の実績は2023年12月31日時点の情報です。"
                }
            ]
        },
        {
            "category": "記号処理 (Symbol Handling)",
            "rule_id": "2.1",
            "description": "文中の「※」記号は、すべてHTMLの上付き形式<sup>※</sup>に変換する。視認性向上のため、注釈や脚注の目的で使用。",
            "requirements": [
                {
                    "condition": "単語の末尾に「※」が付いている場合（例：重要事項※）",
                    "correction": "上付き形式に変換：例）重要事項<sup>※</sup>"
                },
                {
                    "condition": "文末に「※」がある場合",
                    "correction": "上付き形式に変換：例）〜とされています<sup>※</sup>"
                },
                {
                    "condition": "「※」が単独で文中に存在する場合",
                    "correction": "<sup>※</sup> に変換"
                },
                {
                    "condition": "「※」が複数回出現する場合",
                    "correction": "すべて個別に変換する（例：リスク※や費用※ → リスク<sup>※</sup>や費用<sup>※</sup>）"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "重要事項※",
                    "output": "重要事項<sup>※</sup> (<span>修正理由: 単語の末尾にある「※」を上付きに変換 <s style=\"background:yellow;color:red\">重要事項※</s> → <sup>※</sup></span>)"
                },
                {
                    "input": "リスク※や費用※",
                    "output": "リスク<sup>※</sup>や費用<sup>※</sup> (<span>修正理由: 単語末尾の「※」を上付きに変換 <s style=\"background:yellow;color:red\">リスク※や費用※</s> → リスク<sup>※</sup>や費用<sup>※</sup></span>)"
                },
                {
                    "input": "※",
                    "output": "<span style=\"color:red;\">※</span> (<span>修正理由: 単独の「※」を上付きに変換 <s style=\"background:yellow;color:red\">※</s> → <sup>※</sup></span>)"
                }
            ],
            "notes": "文脈の意味が変わらないように注意し、自然な文章になるように変換すること。一括置換ではなく、文全体を見て慎重に変換を行うこと。"
        },
        {
            "category": "数値記号の統一 (Numeric Sign Consistency)",
            "rule_id": "3.1",
            "description": "収益率・騰落率などにおいて、正の数値には明示的に「+」を付与して統一性を保つ。既に「+」「−」が付いているものは変更しない。",
            "requirements": [
                {
                    "condition": "収益率、騰落率などで、正の数値に符号（+）が付いていない場合",
                    "correction": "符号（+）を付与する（例：4.04％ → +4.04％）"
                },
                {
                    "condition": "すでに「+」や「−」が付いている数値",
                    "correction": "変更しない"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「＋」「-」の明示的統一'",
            "Examples": [
                {
                    "input": "○月間の基準価額の騰落率は4.04％、ベンチマークの騰落率は4.53％、超過収益は-0.49％となりました。",
                    "output": "○月間の基準価額の騰落率は<span style=\"color:red;\">4.04％</span> (<span>修正理由: 「＋」「-」の明示的統一 <s style=\"background:yellow;color:red\">4.04％</s> → +4.04％</span>)、ベンチマークの騰落率は<span style=\"color:red;\">4.53％</span> (<span>修正理由: 「＋」「-」の明示的統一 <s style=\"background:yellow;color:red\">4.53％</s> → +4.53％</span>)、超過収益は-0.49％となりました。"
                },
                {
                    "input": "前月比は0.00％で、先月は-1.23％でした。",
                    "output": "前月比は<span style=\"color:red;\">0.00％</span> (<span>修正理由: 「＋」「-」の明示的統一 <s style=\"background:yellow;color:red\">0.00％</s> → +0.00％</span>)で、先月は-1.23％でした。"
                }
            ],
            "notes": "対象数値は一般的に％が後ろに付く収益や成長値などに限定。整数・小数とも対象（例：5％、0.00％、1.234％ など）文章内に複数該当がある場合もすべて個別に対応する"
        },
        {
            "category": "カタカナ表記の統一",
            "rule_id": "4.1",
            "description": "外来語などカタカナ語の表記ゆれを文書内で統一する。",
            "requirements": [
                {
                    "condition": "カタカナ語が複数表記で出現する場合",
                    "correction": "最初に出現した表記を基準に統一。ただし、外部の正準リスト（辞書等）で誤りが明示されている場合はその正準形に従う。"
                },
                {
                    "condition": "表記が文書内で混在している（例：サステナブル／サスティナブル）",
                    "correction": "最初の出現語（あるいは標準リスト準拠）に合わせて他を統一する"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「カタカナ表記の統一'",
            "Examples": [
                {
                    "input": "このファンドはサステナブルな投資を目指します。なお、近年はサスティナブル経営が重視されています。",
                    "output": "このファンドは<span style=\"color:red;\">サステナブル</span> (<span>修正理由: カタカナ表記の統一 <s style=\"background:yellow;color:red\">サステナブル</s> → サスティナブル</span>)な投資を目指します。なお、近年はサスティナブル経営が重視されています。"
                },
                {
                    "input": "エンターテイメントの分野では、エンターテインメント企業の影響が大きい。",
                    "output": "エンターテイメントの分野では、<span style=\"color:red;\">エンターテインメント</span> (<span>修正理由: カタカナ表記の統一 <s style=\"background:yellow;color:red\">エンターテインメント</s> → エンターテイメント</span>)企業の影響が大きい。"
                }
            ],
            "notes": "最初に出現したカタカナ語を文書内基準とする。ただし、外部で定義された標準用語がある場合はそちらを優先。自動処理時は、同義語グループを辞書で管理し、最初の使用形にマッピングする処理が必要。"
        },
        {
            "category": "経済指標の月明記",
            "rule_id": "5.1",
            "description": "経済指標が特定の時期に基づくものである場合、その月を明記する。",
            "requirements": [
                {
                    "condition": "経済指標（例: PMI, CPI, 雇用統計など）が登場し、時期が文脈から特定可能だが、文中に月が記載されていない場合",
                    "correction": "「○月の」などの形で時期を明記する。国名が必要な場合は併記する。"
                }
            ],
            "output_format": "修正箇所は <span style=\"color:red;\">修正前</span> (<span>修正理由: 月が明記されていないため、明記が必要 <s style=\"background:yellow;color:red\">修正前</s> → 修正後</span>) の形式で表示",
            "Examples": [
                {
                    "input": "製造業PMI（購買担当者景気指数）は上昇しました。",
                    "output": "<span style=\"color:red;\">製造業PMI（購買担当者景気指数）は上昇しました。</span> (<span>修正理由: 月が明記されていないため、明記が必要 <s style=\"background:yellow;color:red\">製造業PMI（購買担当者景気指数）は上昇しました。</s> → 10月の製造業PMI（購買担当者景気指数）は上昇しました。</span>)"
                }
            ],
            "notes": "月の明記は可能な限り前方に記載（例:『4月のCPI』）。日本国外の指標の場合、必要に応じて『米国の』なども追加する。"
        },
        {
            "category": "加速表現の強調",
            "rule_id": "5.2",
            "description": "「加速」という語が登場する場合は常に強調して修正提案する（曖昧性排除目的）。",
            "requirements": [
                {
                    "condition": "文中に「加速」が含まれる場合",
                    "correction": "文意が明確でも、常に<span>修正理由: 「加速」を提示する</span>形式でマーキングして明示的に注意喚起を行う"
                }
            ],
            "output_format": "修正箇所は <span style=\"color:red;\">修正前</span> (<span>修正理由: 「加速」を提示する <s style=\"background:yellow;color:red\">修正前</s> → 修正後</span>) の形式で表示",
            "Examples": [
                {
                    "input": "景気は前月から加速しました。",
                    "output": "<span style=\"color:red;\">景気は前月から加速しました。</span> (<span>修正理由: 「加速」を提示する <s style=\"background:yellow;color:red\">景気は前月から加速しました。</s> → 景気は前月から加速しました。</span>)"
                }
            ],
            "notes": "「加速」は経済文脈で曖昧な場合が多いため、常に修正候補としてマーク。可能であれば、何がどのように加速したのか（例: 回復ペース、下落幅、景気回復の速度など）を明記するよう提案。"
        },
        {
            "category": "会計期間の区切り記号統一",
            "rule_id": "6.1",
            "description": "会計期間表記において「～」は禁止し、必ず「ー」（全角ハイフン）に統一する。",
            "requirements": [
                {
                    "condition": "表記中に「～」が含まれている場合",
                    "correction": "「～」を「ー」に置換する"
                }
            ],
            "output_format": "修正箇所は <span style=\"color:red;\">修正前</span> (<span>修正理由: 会計期間表記において「～」は禁止、「ー」に修正 <s style=\"background:yellow;color:red\">修正前</s> → 修正後</span>) の形式で表示",
            "Examples": [
                {
                    "input": "6～8月期",
                    "output": "<span style=\"color:red;\">6～8月期</span> (<span>修正理由: 会計期間表記において「～」は禁止、「ー」に修正 <s style=\"background:yellow;color:red\">6～8月期</s> → 6ー8月期</span>)"
                }
            ]
        },
        {
            "category": "カレンダー年採用国の年度明記追加",
            "rule_id": "6.2",
            "description": "カレンダー年（1月～12月）採用国の年度表記には、具体的な期間を括弧で明記する。",
            "requirements": [
                {
                    "condition": "「年度」のみ記載で月の範囲がない場合",
                    "correction": "「（yyyy年1月-12月）」を追加"
                },
                {
                    "condition": "「上期」や「下期」がある場合",
                    "correction": "期間を対応した形で括弧内に追加（例: 上期→1月-6月、下期→6月-12月）"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "ブラジルの2021年度予算",
                    "output": "ブラジルの2021年度（2021年1月-12月）予算"
                },
                {
                    "input": "ブラジルの2021年下期予算",
                    "output": "ブラジルの2021年下期（2021年6月-12月）予算"
                }
            ]
        },
        {
            "category": "年度または四半期表現の年明示義務",
            "rule_id": "6.3",
            "description": "1～3月発表の開示資料において、「今年度」「来年度」など曖昧な年度表現は使わず、西暦＋年度に統一する。",
            "requirements": [
                {
                    "condition": "「今年度」「来年度」などの年度表現が使われている場合（1～3月発表文脈）",
                    "correction": "実際の西暦を用いて「yyyy年度」に置換する"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "今年度の第1四半期にあたる",
                    "output": "<span style=\"color:red;\">今年度の第1四半期にあたる</span> (<span>修正理由: 1～3月発表は西暦＋年度に統一 <s style=\"background:yellow;color:red\">今年度の第1四半期にあたる</s> → 2024年度の第1四半期にあたる</span>)"
                },
                {
                    "input": "来年度の予想",
                    "output": "<span style=\"color:red;\">来年度の予想</span> (<span>修正理由: 1～3月発表は西暦＋年度に統一 <s style=\"background:yellow;color:red\">来年度の予想</s> → 2025年度の予想</span>)"
                }
            ]
        },
        {
            "category": "四半期表現の月期統一",
            "rule_id": "6.4",
            "description": "会計期間が四半期であっても、「〇-〇月期」の形式で表現を統一する。",
            "requirements": [
                {
                    "condition": "「第1四半期（5月21日～8月20日）」など不定期な期間表現がある場合",
                    "correction": "「～」は「ー」に直し、可能な限り「〇-〇月期」形式に置換"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '会計期間表記は「ー」区切りに統一し、月期に直す'",
            "Examples": [
                {
                    "input": "第1四半期（5月21日～8月20日）",
                    "output": "<span style=\"color:red;\">第1四半期（5月21日～8月20日）</span> (<span>修正理由: 会計期間表記は「ー」区切りに統一し、月期に直す <s style=\"background:yellow;color:red\">第1四半期（5月21日～8月20日）</s> → 6-8月期</span>)"
                }
            ]
        },
        {
            "category": "年表記の統一",
            "rule_id": "YearNotation-1",
            "description": "年表記は常に4桁（西暦）で統一し、省略表記は不可。",
            "requirements": [
                {
                    "condition": "年を表現する場合",
                    "correction": "必ず4桁の西暦（例：2022年）で表記。2桁（例：'22年'）は禁止。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "22年の経済成長率は",
                    "output": "<span style=\"color:red;\">22年の経済成長率は</span> (<span>修正理由: 年表記は4桁（西暦）に統一 <s style=\"background:yellow;color:red\">22年の経済成長率は</s> → 2022年の経済成長率は</span>)"
                }
            ]
        },
        {
            "category": "前年比較表現の統一",
            "rule_id": "YearNotation-2",
            "description": "前年比較の表現は「前年同月比」または「前年同期比」に統一。通年比較のみ「前年比」可。",
            "requirements": [
                {
                    "condition": "前年との比較を表現する場合",
                    "correction": "月次・期次データは「前年同月比」または「前年同期比」に統一。通年比較のみ「前年比」可。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '月次データの比較は「前年同月比」に統一'",
            "Examples": [
                {
                    "input": "1月のCPIは前年比で上昇しました。",
                    "output": "1月のCPIは <span style=\"color:red;\">前年比</span> (<span>修正理由: 月次データの比較は「前年同月比」に統一 <s style=\"background:yellow;color:red\">前年比</s> → 前年同月比</span>) で上昇しました。"
                },
                {
                    "input": "2023年のGDPは前年比+3.0％となりました。",
                    "output": "修正不要。この場合通年比較なので「前年比」使用可。"
                }
            ]
        },
        {
            "category": "年跨ぎ指標の年表記",
            "rule_id": "YearNotation-3",
            "description": "複数年にまたがる指標（CPI, PMI, GDP等）は、最初の出現時のみ年を明記し、以降は文脈が明確な場合は月のみで可。",
            "requirements": [
                {
                    "condition": "異なる年にまたがる月を記載する場合",
                    "correction": "最初の月にのみ年を明記し、以降は文脈が明確なら月のみで可。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                { 
                    "input":"",
                    "output":""
                }
            ]
        },
        {
            "category": "年推定ルール",
            "rule_id": "YearNotation-4",
            "description": "文脈から年を推定する場合のルール。",
            "requirements": [
                {
                    "condition": "直前に4桁の年（例：2025年）が明記されている場合",
                    "correction": "以降の月はその年を基準とする。"
                },
                {
                    "condition": "年が明記されていない場合",
                    "correction": "現在月を基準に推定。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '年表記の初出が不明確なためメッセージ提示'",
            "Examples": [
                {
                    "input": "12月のCPIは～",
                    "output": "<span style=\"color:red;\">12月のCPIは～</span> (<span>修正理由: 年表記の初出が不明確なためメッセージ提示 <s style=\"background:yellow;color:red\">12月のCPIは～</s> → 2024年12月のCPIは～</span>)"
                },
                {
                    "input": "1月のユーロ圏PMIは～",
                    "output": "<span style=\"color:red;\">1月のユーロ圏PMIは～</span> (<span>修正理由: 年表記の初出が不明確なためメッセージ提示 <s style=\"background:yellow;color:red\">1月のユーロ圏PMIは～</s> → 2025年1月のユーロ圏PMIは～</span>)"
                },
                {
                    "input": "10-12月期のGDPは～",
                    "output": "<span style=\"color:red;\">10-12月期のGDPは～</span> (<span>修正理由: 年表記の初出が不明確なためメッセージ提示 <s style=\"background:yellow;color:red\">10-12月期のGDPは～</s> → 2024年10-12月期のGDPは～</span>)"
                }
            ],
            "notes": "常に現在の年・月を動的に判定すること（システム時刻や手動設定の基準月を利用）。月範囲（例：3-5月期, 10-12月期）は最終月を抽出して年を推定。年が明記されている場合は推定しない（例：2024年12月のCPIはそのまま）。" 
        },
        {
            "category": "レンジ表記の統一",
            "rule_id": "RangeNotation-1",
            "description": "数値レンジ（範囲）を表現する際は、パーセントの場合両端に必ず「％」を付ける。",
            "requirements": [
                {
                    "condition": "パーセントのレンジ（範囲）を表現する場合",
                    "correction": "両端に「％」を付けて表記（例: -1％～0.5％）"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'レンジ表記においては両端に％を付ける必要あり'",
            "Examples": [
                {
                    "input": "物価上昇率は-1～0.5の範囲で推移しました。",
                    "output": "物価上昇率は<span style=\"color:red;\">-1～0.5</span> (<span>修正理由: レンジ表記においては両端に％を付ける必要あり <s style=\"background:yellow;color:red\">-1～0.5</s> → -1％～0.5％</span>) の範囲で推移しました。"
                }
            ],
            "notes": "「-1～0.5％」のような表現は誤りであり、「-1％～0.5％」と記載すること。"
        },
        {
            "category": "InvestmentEnvironmentMonthEndExpression",
            "rule_id": "先月末",
            "description": "When mentioning 'last month's investment environment' ('先月の投資環境') in the article, apply the following correction rule,When referring to '先月末' inside the '先月の投資環境' section, if it matches the month prior to the report creation standard month, it must be expressed as '前月末' or '●月末'. Simply saying '先月末' is not acceptable.",
            "requirements": [
                {
                    "condition": "先月末",
                    "correction": "前月末"
                }
            ],
            "Example": [
                {
                    "Input": "先月末時点での株式市場は～",
                    "Output": "<span style=\"color:red;\">先月末</span> (<span>修正理由: 投資環境記載時は「前月末」または「●月末」と記載するルール <s style=\"background:yellow;color:red\">先月末</s> → 前月末</span>) 時点での株式市場は～"
                }
            ]
        },
        {
            "category": "AvoidDirectMentionOfSpecificCompanyNames(個別企業名の表記)",
            "rule_id": "AvoidDirectMentionOfSpecificCompanyNames",
            "description": "When describing the investment environment or economic conditions, avoid mentioning specific company names directly. ",
            "requirements": [
                {
                    "condition": "Expressions must be generalized without citing the company's exact name.",
                    "correction": ""
                },
                {
                    "condition": "Instead of using the company name, describe it by category, industry, or region.",
                    "correction": ""
                }
            ],
            "Example": {
                "Input": "スイス金融大手クレディ・スイスは～",
                "Output": "スイスの大手金融グループ<span style=\"color:red;\">クレディ・スイス</span> (<span>修正理由: 個別企業名の直接記載は禁止。カテゴリや地域に一般化する必要あり。 <s style=\"background:yellow;color:red\">クレディ・スイス</s> → スイスの大手金融グループ</span>) は～"
            }
        },
        {
            "category": "PositiveAndNegativeContributionExpressions",
            "rule_id": "PositiveAndNegativeContributionExpressions",
            "description": "When describing positive or negative contributions (寄与/影響), apply the following rules",
            "requirements": [
                {
                    "condition": "For positive effects, either プラスに寄与 or プラスに影響 is acceptable.",
                    "correction": ""
                },
                {
                    "condition": "Alternatively, you may rephrase as ～プラス要因となる.",
                    "correction": ""
                },
                {
                    "condition": "For negative effects, 寄与 must not be used. Always use マイナスに影響 instead.",
                    "correction": ""
                },
                {
                    "condition": "Alternatively, rephrase as ～マイナス要因となる.",
                    "correction": ""
                },
                {
                    "condition": "Detect incorrect usage according to these standards and mark appropriately.",
                    "correction": ""
                }
            ],
            "Example": {
                "Input": "政策の遅れが景気にマイナスに寄与した。",
                "Output": "<span style=\"color:red;\">マイナスに寄与</span> (<span>修正理由:マイナス方向は「影響」表記に統一する <s style=\"background:yellow;color:red\">マイナスに寄与</s> → マイナスに影響</span>) "
            }
        },
        {
            "category": "YieldMovementdescription",
            "rule_id": "YieldMovementdescription",
            "description": "When describing the movement of yields (利回り), ensure that the inverse relationship with prices is properly reflected.",
            "requirements": [
                {
                    "condition": "If yields rise, it implies that prices fall.",
                    "correction": ""
                },
                {
                    "condition": "If yields fall, it implies that prices rise.",
                    "correction": ""
                },
                {
                    "condition": "If this inverse relationship is not mentioned where necessary, highlight and prompt for correction.",
                    "correction": ""
                },
                {
                    "condition": "利回りは「上昇（価格は下落）」または「低下（価格は上昇）」と表記。",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "利回りの低下",
                    "Output": "<span style=\"color:red;\">利回りの低下</span> (<span>修正理由:利回りと価格の逆相関関係を明記する必要あり <s style=\"background:yellow;color:red\">利回りの低下</s> → 価格は上昇</span>)"
                },
                {
                    "Input": "利回りの上昇",
                    "Output": "<span style=\"color:red;\">利回りの上昇</span> (<span>修正理由:利回りと価格の逆相関関係を明記する必要あり <s style=\"background:yellow;color:red\">利回りの上昇</s> → 価格は下落</span>)"
                }
            ]
        },
        {
            "category": "CorrectUsageOfTeikaAndGeraku",
            "rule_id": "CorrectUsageOfTeikaAndGeraku",
            "description": "When describing changes in yields, prices, or interest rates, apply the following word choice rules strictly",
            "requirements": [
                {
                    "condition": "利回り（Yield）",
                    "correction": "use 低下 for decline, not 下落."
                },
                {
                    "condition": "価格（Price）",
                    "correction": "use 下落 for decline, not 低下."
                },
                {
                    "condition": "金利（Interest Rate）",
                    "correction": "use 低下 for decline, not 下落."
                },
                {
                    "condition": "利回りは「上昇（価格は下落）」または「低下（価格は上昇）」と表記。",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "米国債利回りが下落しました。",
                    "Output": "<span style=\"color:red;\">米国債利回りが下落しました。</span> (<span>修正理由:利回りには「低下」を使用 <s style=\"background:yellow;color:red\">米国債利回りが下落しました。</s> → 米国債利回りが低下しました。</span>)"
                },
                {
                    "Input": "株価が低下しました。",
                    "Output": "<span style=\"color:red;\">株価が低下しました。</span> (<span>修正理由:価格には「下落」を使用 <s style=\"background:yellow;color:red\">株価が低下しました。</s> → 株価が下落しました。</span>)"
                },
                {
                    "Input": "金利が下落しました。",
                    "Output": "<span style=\"color:red;\">金利が下落しました。</span> (<span>修正理由:金利には「低下」を使用 <s style=\"background:yellow;color:red\">金利が下落しました。</s> → 金利が低下しました。</span>)"
                }
            ]
        },
        {
            "category": "FundInflowsAndOutflowsExpression",
            "rule_id": "CorrectUsageOfTeikaAndGeraku",
            "description": "When describing fund flows related to investors (especially foreign investors), apply the following standardized expression",
            "requirements": [
                {
                    "condition": "外国人投資家の資金流出",
                    "correction": "外国人投資家からの資金流入"
                },
                {
                    "condition": "Ensure consistency in the wording.",
                    "correction": ""
                },
                {
                    "condition": "If incorrect or incomplete expressions are found, highlight and annotate accordingly.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '資金流出入表現の統一'",
            "Example": {
                "Input": "外国人投資家の資金流入が続きました。",
                "Output": "外国人投資家<span style=\"color:red;\">の資金流入</span> (<span>修正理由:資金流出入表現の統一 <s style=\"background:yellow;color:red\">の資金流入</s> → からの資金流入</span>) が続きました。"
            }
        },
        {
            "category": "PortfoliodescriptionStandardization",
            "rule_id": "PortfoliodescriptionStandardization",
            "description": "When describing portfolio composition, apply the following rule",
            "requirements": [
                {
                    "condition": "Expressions like \"●●への組み入れ\" are prohibited.",
                    "correction": "外国人投資家からの資金流入"
                },
                {
                    "condition": "Always use \"●●の組み入れ\" instead.",
                    "correction": ""
                },
                {
                    "condition": "However, expressions like \"●●への投資比率\" are acceptable and do not need correction.",
                    "correction": ""
                },
                {
                    "condition": "If incorrect expressions are found, highlight and annotate accordingly.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'ポートフォリオ表現の統一'",
            "Example": {
                "Input": "株式への組み入れを拡大しました。",
                "Output": "株式<span style=\"color:red;\">への組み入れ</span> (<span>修正理由:ポートフォリオ表現の統一 <s style=\"background:yellow;color:red\">への組み入れ</s> → の組み入れ</span>) を拡大しました。"
            }
        },
        {
            "category": "CurrencyNotationStandardization",
            "rule_id": "CurrencyNotationStandardization",
            "description": "Throughout the article, unify the notation of currency names consistently.",
            "requirements": [
                {
                    "condition": "If different expressions referring to the same currency are found, choose the first appearing form as the standard.",
                    "correction": "外国人投資家からの資金流入"
                },
                {
                    "condition": "Other variants must be corrected to match the first occurrence.",
                    "correction": ""
                },
                {
                    "condition": "However, expressions like \"●●への投資比率\" are acceptable and do not need correction.",
                    "correction": ""
                },
                {
                    "condition": "If incorrect expressions are found, highlight and annotate accordingly.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '通貨表記の統一'",
            "Example": {
                "Input": "米ドル建て資産とUSD建て資産に投資しました。",
                "Output": "米ドル建て資産と<span style=\"color:red;\">USD建て資産</span> (<span>修正理由:通貨表記の統一 <s style=\"background:yellow;color:red\">USD建て資産</s> → 米ドル建て資産</span>) に投資しました。"
            }
        },
        {
            "category": "ZeroPercentCompositionNotation",
            "rule_id": "ZeroPercentCompositionNotation",
            "description": "When describing a composition ratio of 0%, use either 「0％程度」 or 「ゼロ％程度」",
            "requirements": [
                {
                    "condition": "Both expressions are acceptable, but direct expressions like just \"0%\" without 「程度」 should be corrected.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '構成比0％の表記統一'",
            "Example": {
                "Input": "当ファンドの構成比は0％です。",
                "Output": "当ファンドの構成比は<span style=\"color:red;\">0％</span> (<span>修正理由:構成比0％の表記統一 <s style=\"background:yellow;color:red\">0％</s> → 0％程度</span>) です。"
            }
        },

        {
            "category": "TerminologyConsistency_Sentiment",
            "rule_id": "TerminologyConsistency_Sentiment",
            "description": "Both 「先高感」 and 「先高観」 are acceptable expressions.",
            "requirements": [
                {
                    "condition": "TerminologyConsistency_Sentiment",
                    "correction": "先高感"
                },
                {
                    "condition": "TerminologyConsistency_Sentiment",
                    "correction": "先高"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Example": [
                {
                    "Input": "金利の先高感が高まりました。",
                    "Output": "No correction needed."
                },
                {
                    "Input": "金利の先高観が意識されました。",
                    "Output": "No correction needed."
                }
            ]
        },
        {
            "category": "TerminologyConsistency_Calm",
            "rule_id": "TerminologyConsistency_Calm",
            "description": "If the usage does not match the context, correct it according to the appropriate meaning.修正理由: 意味の誤用",
            "requirements": [
                {
                    "condition": "Use 「沈静」 when referring to natural calming down over time.",
                    "correction": ""
                },
                {
                    "condition": "Use 「鎮静」 when referring to intentional or artificial suppression (e.g., medical treatment, intervention).",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "市場は徐々に鎮静していった。",
                    "Output": "市場は徐々に <span style=\"color:red;\">沈静</span> (<span>修正理由: 意味の誤用 <s style=\"background:yellow;color:red\">鎮静</s> → 沈静</span>) していった。"
                },
                {
                    "Input": "医療チームは患者の暴動を沈静させた。",
                    "Output": "医療チームは患者の暴動を <span style=\"color:red;\">鎮静</span> (<span>修正理由: 意味の誤用 <s style=\"background:yellow;color:red\">沈静</s> → 鎮静</span>) させた。"
                }
            ]
        },
        {
            "category": "TerminologyConsistency_BenchmarkIndex",
            "rule_id": "TerminologyConsistency_BenchmarkIndex",
            "description": "The notation of each index must be consistent throughout the document. 修正理由:メッセージ提示（指数名称の表記統一）",
            "requirements": [
                {
                    "condition": "If multiple variations of the same index name appear (e.g., abbreviated and full names), this inconsistency must be detected.",
                    "correction": ""
                },
                {
                    "condition": "Correction should not be automatically performed",
                    "correction": ""
                },
                {
                    "condition": "Instead, highlight and prompt the user with a correction suggestion as a message.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Example": {
                "Input": "本ファンドのベンチマークはMSCIインド指数です。",
                "Output": "<span style=\"color:red;\">MSCIインド指数</span> (<span>修正理由: 表記統一の確認が必要 <s style=\"background:yellow;color:red\">MSCIインド指数</s> → MSCIインド・インデックス</span>)"
            }
        },
        {
            "category": "TerminologyConsistency_DowJones",
            "rule_id": "TerminologyConsistency_DowJones",
            "description": "Always use the full official name 「ダウ平均株価」）",
            "requirements": [
                {
                    "condition": "If multiple variations of the same index name appear (e.g., abbreviated and full names), this inconsistency must be detected.",
                    "correction": ""
                },
                {
                    "condition": "Abbreviations like 「ダウ平均」 or 「NYダウ」 are not allowed when mentioning it as a benchmark or reference index.",
                    "correction": ""
                },
                {
                    "condition": "Instead, highlight and prompt the user with a correction suggestion as a message.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "米国市場ではダウ平均が上昇しました。",
                    "Output": "米国市場では <span style=\"color:red;\">ダウ平均</span> (<span>修正理由:正式名称「ダウ平均株価」に統一 <s style=\"background:yellow;color:red\">ダウ平均</s> → ダウ平均株価</span>) が上昇しました。"
                },
                {
                    "Input": "NYダウも同様にプラスとなりました。",
                    "Output": "<span style=\"color:red;\">NYダウ</span> (<span>修正理由:正式名称「ダウ平均株価」に統一 <s style=\"background:yellow;color:red\">NYダウ</s> → ダウ平均株価</span>) も同様にプラスとなりました。"
                }
            ]
        },
        {
            "category": "TerminologyConsistency_Others",
            "rule_id": "TerminologyConsistency_Others",
            "description": "Ensure the following specialized terms are consistent",
            "requirements": [
                {
                    "condition": "- \"平均乖離率\" should be used instead of \"平均かい離率\" or other variants.",
                    "correction": "平均かい離率"
                },
                {
                    "condition": "- \"運用報告書\" should not be abbreviated or altered.",
                    "correction": ""
                },
                    {
                    "condition": "- \"営業日\" must be spelled correctly without deviations.",
                    "correction": ""
                },
                    {
                    "condition": "- \"日経平均株価\" must always be used fully; \"日経平均\" alone is not acceptable when used as a benchmark.",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "平均かい離率は10％でした。",
                    "Output": "<span style=\"color:red;\">平均かい離率</span> (<span>修正理由:用語統一 <s style=\"background:yellow;color:red\">平均かい離率</s> → 平均乖離率</span>) は10％でした。"
                },
                {
                    "Input": "営業日数は多かった。",
                    "Output": "変更不要（正しい表記）"
                }
            ]
        },
        {
            "category": "Attractive_Terms_Check",
            "rule_id": "Attractive_Terms_Check",
            "description": "If the above expressions are found, highlight them and provide a message prompt to the user.Do not automatically delete or correct the expression. These expressions require careful usage: They can be used if a sufficient basis is provided. However, when referring directly to funds (ファンド), their use is prohibited even with justification.",
            "requirements": [
                {
                    "condition": "魅力的な",
                    "correction": "魅力的な"
                },
                {
                    "condition": "投資妙味",
                    "correction": "投資妙味"
                },
                {
                    "condition": "割高.",
                    "correction": "割高"
                },
                {
                    "condition": "割高感.",
                    "correction": "割高感"
                },
                {
                    "condition": "割安.",
                    "correction": "割安"
                },
                {
                    "condition": "割安感.",
                    "correction": "割安感"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "魅力的な",
                    "Output": "<span style=\"color:red;\">魅力的な</span> (<span>修正理由: メッセージ提示 <s style=\"background:yellow;color:red\">魅力的な</s> → 魅力的な</span>)"
                },
                {
                    "Input": "また、現在の価格は割安感があると判断できます。",
                    "Output": "<span style=\"color:red;\">割安感</span> (<span>修正理由: メッセージ提示 <s style=\"background:yellow;color:red\">割安感</s> → 割安感</span>)"
                }
            ]
        },
        {
            "category": "Prohibited_Expression_Must",
            "rule_id": "Prohibited_Expression_Must",
            "description": "Avoid any expression implying a definitive judgment (such as \"必ず～\") about future investment performance, economic indicators, corporate earnings, etc. unless explicit supporting evidence is provided.If detected, mark the problematic phrase and issue a message prompt.",
            "requirements": [
                {
                    "condition": "「必ず～」のように、将来の運用成績や経済指標・企業業績等に対する断定的な判断を示す表現は、根拠が明示されていない限り使用不可。",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '断定的な判断表現、根拠明示なし'",
            "Example": {
                "Input": "このファンドは必ず利益を生み出します。",
                "Output": "このファンドは <span style=\"color:red;\">必ず</span> (<span>修正理由:断定的な判断表現、根拠明示なし</span>) 利益を生み出します。"
            }
        },
        {
            "category": "Prohibited_or_Cautioned_Expressions_Selling_and_Limited",
            "rule_id": "Prohibited_or_Cautioned_Expressions_Selling_and_Limited",
            "description": "Detect the above expressions and highlight them. Issue a message prompt to the user. No forced rewriting, only alert.",
            "requirements": [
                {
                    "condition": "断定表現のためメッセージ提示",
                    "correction": ""
                },
                {
                    "condition": "利益確定の売り",
                    "correction": ""
                },
                {
                    "condition": "利食い売り",
                    "correction": ""
                },
                {
                    "condition": "限定的",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '断定表現のためメッセージ提示'",
            "Examples": [
                {
                    "Input": "株価の下落により利益確定の売りが出ました。",
                    "Output": "<span style=\"color:red;\">利益確定の売り</span> (<span>修正理由: 断定表現のためメッセージ提示 <s style=\"background:yellow;color:red\">利益確定の売り</s> → 利益確定の売り</span>)"
                },
                {
                    "Input": "投資家心理が弱まり、利食い売りが続きました。",
                    "Output": "<span style=\"color:red;\">利食い売り</span> (<span>修正理由: 断定表現のためメッセージ提示 <s style=\"background:yellow;color:red\">利食い売り</s> → 利食い売り</span>)"
                },
                {
                    "Input": "政策効果は限定的でした。",
                    "Output": "<span style=\"color:red;\">限定的</span> (<span>修正理由: 曖昧表現のためメッセージ提示 <s style=\"background:yellow;color:red\">限定的</s> → 限定的</span>)"
                }
            ]
        },
        {
            "category": "Prohibited_or_Cautioned_Expressions_Prediction_and_Psychology",
            "rule_id": "Prohibited_or_Cautioned_Expressions_Prediction_and_Psychology",
            "description": "Detect these expressions and highlight them to issue a message prompt. No forced rewriting, only alert the user. Exception: Do NOT trigger warning if part of fixed financial terms such as 「市場予想」 or 「市場心理」.",
            "requirements": [
                {
                    "condition": "予想",
                    "correction": ""
                },
                {
                    "condition": "心理",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '主体不明確'",
            "Examples": [
                {
                    "Input": "景気回復への予想が高まった。",
                    "Output": "<span style=\"color:red;\">予想</span> (<span>修正理由: 主体不明確 <s style=\"background:yellow;color:red\">予想</s> → 予想</span>)"
                },
                {
                    "Input": "利上げ観測を受けて心理が改善した。",
                    "Output": "<span style=\"color:red;\">心理</span> (<span>修正理由: 主体不明確 <s style=\"background:yellow;color:red\">心理</s> → 心理</span>)"
                }
            ]
        },
        {
            "category": "Prohibited_or_Cautioned_Expressions_Rise_and_Decline_Factors",
            "rule_id": "Prohibited_or_Cautioned_Expressions_Rise_and_Decline_Factors",
            "description": "When \"上昇\" or \"下落\" appears, check the entire paragraph context. Verify if a causal explanation (background or reason) is clearly described within the paragraph. If missing, highlight the term and issue a user prompt. Do not modify the original sentence; only annotate and prompt. 上昇・下落要因の記載漏れ、メッセージ提示",
            "requirements": [
                {
                    "condition": "上昇",
                    "correction": ""
                },
                {
                    "condition": "下落",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '上昇・下落要因の記載漏れ、メッセージ提示'",
            "Examples": [
                {
                    "Input": "最近、米国経済において賃金上昇と物価高が続いており、景気過熱感が指摘されています。米国のインフレ懸念の高まりを背景に、株式市場は下落しました。金融当局は今後も利上げを続ける見通しです。",
                    "Output": "（エラーなし:前文に要因記載あり）"
                },
                {
                    "Input": "米国の景気動向が注目されています。株式市場は下落しました。今後の動向を見守る必要があります。",
                    "Output": "<span style=\"color:red;\">下落</span> (<span>修正理由: 上昇・下落要因の記載漏れ、メッセージ提示 <s style=\"background:yellow;color:red\">下落</s> → 下落</span>)"
                }
            ]
        },
        {
            "category": "Replacement Rules for Verb-Type Expressions（動詞・活用形を含む表現の置き換え）",
            "rule_id": "Replacement Rules for Verb-Type Expressions（動詞・活用形を含む表現の置き換え）",
            "description":"Certain terms or expressions require more than simple string or regex-based replacement. These are called dynamically varying expressions, which include but are not limited to: Register-sensitive expressions (e.g., polite/humble language variations) Compound phrases or abbreviations that appear in flexible forms When the term to be replaced is a verb, the system must detect and process all conjugated or inflected forms. Do not use rigid pattern matching. Ensure grammatical accuracy after replacement. In general, all such replacements must be done in a context-sensitive manner, ensuring the result remains grammatically and semantically correct",
            "requirements": [
                {
                    "condition": "先月/前月, 1ヵ月前について言及する場合は、「前月」を使用。",
                    "correction": "前月"
                },
                {
                    "condition": "前期比○％ ,replacement 前期比年率○％ , 基本、期間比較の伸率は、「年率」記載にして下さい。",
                    "correction": ""
                },
                {
                    "condition": "「第◯四半期」という表現を使用する際には、各四半期に対応する月（1〜3月、4〜6月、7〜9月、10〜12月）を明記する必要がある。特に「18年第4四半期」などのように年号や対象月が不明確な場合は、「2018年10〜12月期」などの具体的な年月期間に必ず言い換えること。",
                    "correction": ""
                },
                {
                    "condition": "～に賭ける to ～を予想して ,日本語の使い型変換を注意すべき",
                    "correction": "～を予想して"
                },
                {
                    "condition": "「出遅れ感」「割高感」などの“〜感”を含む表現は、読み手に主観的・感情的印象を与える恐れがあるため注意が必要です。とくに「出遅れ感」は相場観を含む表現であり、客観的事実の報告が求められる金融レポートには不適切な場合があります。",
                    "correction": ""
                },
                {
                    "condition": "そのため、以下のいずれかの対応を行うこと:\n- 相場観を表に出す意図がない場合 → 削除\n- 相場観を意図的に述べる場合 → 「〜と考えます」「〜との見方があります」などに修正",
                    "correction": ""
                },
                {
                    "condition": "約、程度、前後は同時に使わない。",
                    "correction": ""
                },
                {
                    "condition": "「（割安に）放置」などの表現は、主観的な割安感に基づいているにもかかわらず、省略形によって根拠や判断主体が曖昧になるため、読者に正確な意味が伝わりにくい。このような場合には、「割安感のある銘柄」や「割安と評価される」など、より明確で客観的な表現へと言い換えること。",
                    "correction": ""
                },
                {
                    "condition": "「横ばい」という表現は、期間中の価格・利回り等の値動きが非常に小さい場合に限定して使用すること。一方で、期間中に一定の変動があったものの、最終的に開始時点と同程度の水準に戻った場合には、「ほぼ変わらず」「同程度となる」などの表現を使用する。誤って「横ばい」と記述すると、値動きがなかったような誤認を与える可能性があるため、事実に基づいた正確な表現選択が求められる。",
                    "correction": "ほぼ変わらず"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "18年第4四半期の売上は前年同期比で増加しました。",
                    "Output": "<span style=\"color:red;\">18年第4四半期</span> (<span>修正理由: 「18年第4四半期」では西暦か和暦か不明確であり、また期間（月）が具体的に示されていないため、「2018年10〜12月期」と明確な年月表記に修正する必要があります。 <s style=\"background:yellow;color:red\">18年第4四半期</s> → 2018年10〜12月期</span>)"
                },
                {
                    "Input": "～に賭ける",
                    "Output": "～を予想して"
                },
                {
                    "Input": "資源関連株は出遅れ感から買われました。",
                    "Output": "<span style=\"color:red;\">出遅れ感</span> (<span>修正理由: 「出遅れ感」は主観的な相場観を含む表現であり、事実報告を重視する文書では「〜との見方」などの客観的表現に言い換える必要があります。 <s style=\"background:yellow;color:red\">出遅れ感</s> → 相対的に割安との見方</span>)"
                },
                {
                    "Input": "約○％程度",
                    "Output": "約○％"
                },
                {
                    "Input": "一部の小型株は（割安に）放置されています。",
                    "Output": "<span style=\"color:red;\">（割安に）放置</span> (<span>修正理由: 「（割安に）放置」という省略的な表現は、主観的評価である割安感を文中に明確に示す必要があるため、より客観的な言い換えが必要です。 <s style=\"background:yellow;color:red\">（割安に）放置</s> → 割安感のある水準にあると見られています。</span>)"
                },
                {
                    "Input": "当作成期を通してみると債券利回りは横ばいでした。",
                    "Output": "<span style=\"color:red;\">横ばい</span> (<span>修正理由: 期間中に一定の変動幅が確認されており、「横ばい」という表現は実態と合わないため、「ほぼ変わらず」とするのが適切。 <s style=\"background:yellow;color:red\">横ばい</s> → ほぼ変わらず</span>)"
                }
            ]
        },
        {
            "category": "行って来い ⇒ 「上昇（下落）したのち下落（上昇）」等へ書き換える",
            "rule_id": "行って来い ⇒ 「上昇（下落）したのち下落（上昇）」等へ書き換える",
            "description": "The expression “行って来い” is informal and vague. It must not be used in formal financial documents or reports intended for external audiences.Replace it with a precise description of the price movement, such as: “上昇したのち下落した” 下落したのち上昇した Always use fact-based, objective wording that clearly describes the market movement.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「行って来い」は曖昧かつ口語的な表現であり、正式な金融文書では具体的な値動きを明記する必要があります。'",
            "Examples": [
                {
                    "Input": "相場は行って来いの展開となりました。",
                    "Output": "<span style=\"color:red;\">行って来い</span> (<span>修正理由: 「行って来い」は曖昧かつ口語的な表現であり、正式な金融文書では具体的な値動きを明記する必要があります。 <s style=\"background:yellow;color:red\">行って来い</s> → 一時上昇したものの、その後下落し、前日と同水準で終了しました。</span>)"
                }
            ]
        },
        {
            "category": "The phrase format “○○大手” (e.g., “通信大手”) is not preferred in formal writing.",
            "rule_id": "The phrase format “○○大手” (e.g., “通信大手”) ",
            "description": "Always rephrase using the structure “大手 + [industry/company type]”, such as: “大手通信会社” “大手不動産企業” “大手半導体メーカー” The “大手” modifier should appear before the industry or company type.If the noun is a company name, ensure “大手” precedes the type, not the sector.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "通信大手が新サービスを発表しました。",
                    "output": "<span style=\"color:red;\">通信大手</span> (<span>修正理由: 「通信大手」という表現は語順が不自然であり、正式なビジネス文書では「大手通信会社」のように「大手」を先に置く表現が適切です。 <s style=\"background:yellow;color:red\">通信大手</s> → 大手通信会社</span>)"
                }
            ]
        },
        {
            "category": "The term “ローン” (katakana loanword) ",
            "rule_id": "rule_12",
            "description": "Replace “ローン” with “貸し付け” or “融資” in general.Do not flag when part of fixed financial terms such as “住宅ローン”, “教育ローン”, “マイカーローン” This maintains formal tone and correct financial terminology.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "企業向けのローンが増加しています。",
                    "output": "<span style=\"color:red;\">ローン</span> (<span>修正理由: 「ローン」は口語的で曖昧な表現であり、金融用語としては「融資」と明確に表記するのが適切です。ただし、「住宅ローン」などの固定用語はこの限りではありません。 <s style=\"background:yellow;color:red\">ローン</s> → 融資</span>)"
                },
                {
                    "input": "住宅ローンの金利が今月から引き下げられました。",
                    "output": "<span style=\"color:red;\">住宅ローン</span> (<span>修正理由: 「住宅ローン」は一般的かつ正式な金融商品名であり、修正の必要はありません。</span>)"
                }
            ]
        },
        {
            "category": "Use “格付け” (with okurigana) when the word stands alone or is used as a verb/noun phrase, and “格付” (without okurigana) in compounds or technical terms.",
            "rule_id": "rule_13",
            "description": "Use “格付け” for standalone usage.""Use “格付” for compound words such as “格付機関”, “格付別”.""Do not mix both styles within the same document. Ensures consistency with financial industry standards.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "格付",
                    "output": "<span style=\"color:red;\">格付</span> (<span>修正理由: 単独使用の場合は送り仮名「け」を付けて「格付け」とするのが適切です。 <s style=\"background:yellow;color:red\">格付</s> → 格付け</span>)"
                },
                {
                    "input": "格付機関",
                    "output": "<span style=\"color:red;\">格付機関</span> (<span>修正理由: 「格付機関」は専門用語として定着しており、複合語では送り仮名を付けないのが適切です。</span>)"
                }
            ]
        },
        {
            "category": "買い入れの書き方",
            "rule_id": "rule_14",
            "description": "Use “国債買い入れ” (with okurigana) when the word stands alone or is used as a verb/noun phrase. Use “国債買入” (without okurigana) when part of a compound word or technical term (e.g., 国債買入オペ). Do not mix both styles within the same document.Ensures consistency with financial industry standards and clarity",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "引締の書き方",
            "rule_id": "rule_15",
            "description": "Use “引き締め” (with okurigana) when the word stands alone or is used as a verb/noun phrase. Use “引締” (without okurigana) when part of a compound word or technical term (e.g., 引締策). Do not mix both styles within the same document. Ensures consistency with financial industry standards and clarity.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "積極姿勢/消極姿勢の表現修正",
            "rule_id": "rule_16",
            "description": "Expressions such as “積極姿勢” or “消極姿勢” are too abstract and vague for formal financial writing. Always replace with more specific, descriptive investment actions, e.g.: “長めとした” (increased allocation/duration) “組入を増やした” (added to portfolio) “保有を減らした” (reduced exposure) Ensure the expression clearly conveys actual investment action or policy",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「積極姿勢」は抽象的で読み手に具体的な投資行動が伝わらないため、「長めとした」などの明確な表現に置き換える必要があります。'",
            "Examples": [
                {
                    "input": "積極姿勢",
                    "output": "<span style=\"color:red;\">積極姿勢</span> (<span>修正理由: 「積極姿勢」は抽象的で読み手に具体的な投資行動が伝わらないため、「長めとした」などの明確な表現に置き換える必要があります。 <s style=\"background:yellow;color:red\">積極姿勢</s> → 長め</span>)"
                }
            ]
        },
        {
            "category": "ハト派／タカ派（金融政策に関する表現",
            "rule_id": "rule_17",
            "description": "Expressions “ハト派” (dovish) and “タカ派” (hawkish) are metaphorical and colloquial. Replace with objective descriptions of monetary policy stance, such as:“金融緩和を重視する姿勢” (emphasis on monetary easing) “金融引き締めに積極的” (proactive on tightening) If the replacement conveys the same intent (stance on easing/tightening), it is acceptable.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「ハト派」は比喩的かつ俗語的表現のため、金融政策スタンスを客観的に表す文に修正する必要があります。'",
            "Examples": [
                {
                    "input": "ハト派",
                    "output": "<span style=\"color:red;\">ハト派</span> (<span>修正理由: 「ハト派」は比喩的かつ俗語的表現のため、金融政策スタンスを客観的に表す文に修正する必要があります。 <s style=\"background:yellow;color:red\">ハト派</s> → 金融緩和を重視する姿勢</span>)"
                }
            ]
        },
        {
            "category": "表記の統一",
            "rule_id": "consistent_terminology",
            "description": "Ensure that the writing format of financial terms is consistent throughout the report.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "ひらがなを表記するもの",
            "rule_id": "hiragana_usage",
            "description": "Ensure the report follows the rules for hiragana notation. Replace kanji that are not commonly used with their appropriate hiragana form.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "一部かな書き等で表記するもの",
            "rule_id": "kana_notation_nonstandard_kanji",
            "description":"Ensure non-standard or difficult kanji are replaced with kana in accordance with the standard writing format.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "一般的な送り仮名など",
            "rule_id": "okurigana_usage",
            "description": "Verify that the correct and standardized okurigana is used consistently in the report.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "Foreign Exchange Market Trend Analysis",
            "rule_id": "rule_forex_trend_analysis",
            "description": "Revise to either '円高' or '円安' based on context.",
            "requirements": [
                {
                    "condition": "Occurs when the value of the yen increases relative to other currencies.",
                    "correction": "多くの通貨が対円で下落した"
                },
                {
                    "condition": "Occurs when the value of the yen increases relative to other currencies.",
                    "correction": "ドル円が下落した"
                },
                {
                    "condition": "Occurs when the value of the yen increases relative to other currencies.",
                    "correction": "対米ドルで円の価値が上昇した"
                },
                {
                    "condition": "Occurs when the value of the yen declines relative to other currencies.",
                    "correction": "多くの通貨が対円で上昇した"
                },
                {
                    "condition": "Occurs when the value of the yen declines relative to other currencies.",
                    "correction": "ドル円が上昇した"
                },
                {
                    "condition": "Occurs when the value of the yen declines relative to other currencies.",
                    "correction": "対米ドルで円の価値が下落した"
                },
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                },
                {
                    "input": "",
                    "output": ""
                }
            ],
            "Notes": "The replacement depends on the surrounding context, so parsing of nearby sentences is required."
        },
        {
            "category": "レポートのデータ部との整合性確認",
            "rule_id": "レポートのデータ部との整合性確認",
            "description": "Ensure the textual descriptions in the main body are logically and factually consistent with the data section. No content-related discrepancies should exist.",
            "requirements": [
                {
                    "condition": "",
                    "correction": ""
                },
                {
                    "condition": "",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "",
                    "output": ""
                }
            ]
        },
        {
            "category": "文頭の「〇」印と一文字目の間隔を統一",
            "rule_id": "文頭の「〇」印と一文字目の間隔を統一",
            "description": "- When a sentence or paragraph includes multiple “○” bullet-style markers, the spacing between each “○” and the following text must be **uniform**. - Do not impose a fixed rule like “no space” or “one half-width space”; instead, **adopt the spacing style used in the first occurrence** and apply it consistently throughout the sentence or document.- If spacing after “○” varies (e.g., mixed usage of no space, half-width, and full-width), flag it as a formatting inconsistency and suggest unifying.",
            "requirements": [
                {
                    "condition": "Adopt spacing from the **first occurrence** and apply it consistently.",
                    "correction": "全ての〇の後に、最初の〇に準拠したスペース（例：なし or 半角）を統一する。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「〇」の後に半角空白を入れる形式が最初に使用されているため、すべての「〇」後の間隔を統一する必要があります。'",
            "Examples": [
                {
                    "input": "〇　　利益率は横ばいとなった。",
                    "output": "〇利益率は横ばいとなった"
                }
            ]
        },
        {
            "category": "年度表記",
            "rule_id": "年度表記",
            "description": "Japanese financial documents must standardize how symbols, punctuation, and numbers are written. Two-digit year expressions (e.g., 22年)",
            "requirements": [
                {
                    "condition": "Convert to four-digit year format (e.g., 2022年).",
                    "correction": "22年 → 2022年"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "22年",
                    "output": "2022年"
                }
            ],
            "Notes": "Avoid direct modification in proper nouns or quoted references unless clearly incorrect."
        },
        {
            "category": "Special Formatting Rules",
            "rule_id": "Special Formatting Rules",
            "description": "Japanese financial documents must standardize how symbols, punctuation, and numbers are written. Do not modify names of people, places, or organizations unless clearly incorrect. Delete unnecessary characters without replacing them with others. Maintain intra-text spaces unless at start or end of the sentence.",
            "requirements": [
                {
                    "condition": "Proper Nouns",
                    "correction": "ベッセント氏 → ✅ (No correction)"
                },
                {
                    "condition": "Redundant Characters",
                    "correction": "ユーロ圏域内の景気委 → ユーロ圏域内の景気"
                },
                {
                    "condition": "Preserving Spaces",
                    "correction": "月の 前半は米国の 債券利回りの上昇 につれて → ✅ (No correction)"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "ベッセント氏",
                    "output": "No correction"
                },
                {
                    "input": "ユーロ圏域内の景気委",
                    "output": "ユーロ圏域内の景気"
                },
                {
                    "input": "月の 前半は米国の 債券利回りの上昇 につれて",
                    "output": "No correction"
                }
            ],
            "Notes": "Avoid direct modification in proper nouns or quoted references unless clearly incorrect."
        }
    ]
    
    for ruru_split in ruru_all:
        result = f"""
        You are a professional Japanese text proofreading assistant specialized in high-precision document validation.
        Your task is to carefully and strictly proofread the provided Japanese report based on the detailed rules specified below.

        The proofreading targets include:

        Important:
        All subsequent instructions are divided into sections.
        Each section must be strictly followed without omission.
        You are prohibited from making subjective judgments or skipping steps, even if an error seems minor.
        Always prioritize rule adherence over general readability or aesthetic preference.
        Final Output Requirements:
        Use the specified correction format for each detected error (e.g., using <span style="color:red;"> tags).
        Preserve the original sentence structure and paragraph formatting unless explicitly instructed otherwise.
        If no corrections are needed for a section, explicitly state "No errors detected" (検出された誤りなし).
        Follow all instructions strictly and proceed only according to the rules provided.:
        
        **Report Content to Proofread:**
        {input}

        **Proofreading Requirements:**
        {ruru_split}

        **Output Requirements:**
        1. **Return only structured correction results as a Python-style list of dictionaries:**
        - Format:
            'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason for correction'
        - Example:
            'original': '月間ではほぼ変わらずなりました。', 'correct': '月間ではほぼ変わらずとなりました。', 'reason': '誤字'
        
        2. **Each dictionary must include:**
            - 'original': the original incorrect text
            - 'correct': the corrected text
            - 'reason': a concise explanation for the correction
        3. **Do not include any explanation, HTML tags, or narrative. Only return the data in this dictionary format.**
        4. **Maintain the original document structure internally during processing, but the output should only contain corrections in the required format.**
    
        """
        yield result

@app.route('/api/opt_wording', methods=['POST'])
def opt_wording():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input = data.get("input", "")
        pdf_base64 = data.get("pdf_bytes", "")

        fund_type = data.get("fund_type", "public")  # 기본값은 'public'
        file_name_decoding = data.get("file_name", "")
        icon = data.get("icon", "")
        comment_type = data.get("comment_type", "")
        upload_type = data.get("upload_type", "")
        pageNumber = data.get('pageNumber',0)
        

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400

        prompt_result = loop_in_ruru("\"" + input + "\"")
        async def run_tasks():
            tasks = [handle_result(once) for once in prompt_result]
            return await asyncio.gather(*tasks)

        results = asyncio.run(run_tasks())
        sec_input = "\n".join(results)

        dt = [
            "以下の分析結果に基づき、原文中の誤りを抽出してください",
            "出力は以下のJSON形式でお願いします:",
            "- [{'original': '[原文中の誤っている部分:]', 'correct': '[誤り部分を正しい部分のテキストに修正:]', 'reason': '[理由:]'}]",
            "- 分析結果が正しい場合は、空のリストを返します",
            f"原文:'{input}'\n分析結果:'{sec_input}'"
        ]
        sec_prompt = "\n".join(dt)

        _content = opt_common(input, sec_prompt, pdf_base64, pageNumber,False,False,False,False,False)
        
        return _content
    
        # loop = asyncio.new_event_loop()
        # asyncio.set_event_loop(loop)
        # _content = loop.run_until_complete(opt_common_wording(file_name,fund_type,input,prompt_result,excel_base64,pdf_base64,resutlmap,upload_type,comment_type,icon,pageNumber))
        
        # return _content

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    

# ruru_update_save_corrections
@app.route('/api/save_corrections', methods=['POST'])
def save_corrections():
    try:
        data = request.get_json()
        corrections = data.get('corrections','')
        fund_type = data.get("fund_type",'')
        # link = data.get("link",'')
        file_name_decoding = data.get('file_name','')
        icon = data.get('icon','')

        # URL 디코딩
        file_name = urllib.parse.unquote(file_name_decoding)

        # 1. 필수 필드 검증
        if not file_name or not isinstance(corrections, list):
            return jsonify({"success": False, "error": "file_name과 corrections(list)가 필요합니다."}), 400
        
        # 컨테이너 이름 결정
        container_name = f"{fund_type}_Fund"
        # 2. Cosmos DB 연결
        container = get_db_connection(container_name)

        # 기존 항목 존재 여부 확인
        existing_item = list(container.query_items(
            query="SELECT * FROM c WHERE c.id = @id",
            parameters=[{"name": "@id", "value": file_name}],
            enable_cross_partition_query=True
        ))

        # src_corrections = existing_item[0].get("result", {"corrections": []})["corrections"]
        # corrections.extend(src_corrections)

        # 기존 corrections 가져오기
        existing_corrections = []
        if existing_item:
            result = existing_item[0].get("result", {})
            existing_corrections = result.get("corrections", [])

        # 🔁 중복 제거 (dict list 기준, 'check_point' + 'comment' 기준 등으로)
        def dict_key(d):
            return (d.get('check_point'), d.get('comment'))

        merged_corrections = {dict_key(c): c for c in (corrections + existing_corrections)}
        final_corrections = list(merged_corrections.values())

        # 새 데이터 생성
        item = {
            'id': file_name,
            'fileName': file_name,
            'icon': icon,
            "result": {
                "corrections": final_corrections
            },
            'updateTime': datetime.utcnow().isoformat(),
        }
        

        if not existing_item:
            # 새 항목 생성
            container.create_item(body=item)
            logging.info(f"✅ Cosmos DB Update Success: {file_name}")
        else:
            # 기존 항목 업데이트
            existing_item[0].update(item)
            container.upsert_item(existing_item[0])

            logging.info(f"🔄 Cosmos DB update success: {file_name}")

        return jsonify({"success": True, "message": "Data Update Success"}), 200
    
    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB Error: {str(e)}")
        return jsonify({"success": False, "error": "DB Error"}), 500
    except Exception as e:
        logging.error(f"Server error: {str(e)}")
        return jsonify({"success": False, "error": "Server error"}), 500
    
#th for test api

@app.route("/api/integrated_test", methods=["POST"])
def integrated_test():
    data = request.json
    prompt = data.get("prompt", "")
    input_data = data.get("input_data", "")
    flag_type = data.get("flag_type", "")
    base64_img = data.get("base64_img", "")
    token = token_cache.get_token()
    openai.api_key = token

    if prompt and input_data:
        question = [
            {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
            {"role": "user", "content": input_data}
        ]
        if flag_type == "picture":
            question.append({"role": "user", "content": [{"type": "text", "text": input_data},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_img}"}}]})
        else:
            question.append({"role": "user", "content": input_data})

        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=question,
            max_tokens=32768,
            temperature=0,
            seed=42
        )
        answer = response['choices'][0]['message']['content'].strip()
        return jsonify({"response_ai": answer})

@app.after_request
def after_request(response):
    try:
        data = request.json
        error_text = ""
        file_name = data.get("file_name", "")
        log_controller = get_db_connection(LOG_RECORD_CONTAINER_NAME)
        response_json = json.loads(response.get_data(as_text=True))
        if response.status_code > 200:
            response_json = json.loads(response.get_data(as_text=True))
            error_text = response_json.get("error", "")
        if file_name:
            if re.search(r"save_corrections|write_upload_save", request.path) or (
                    "check_file" in request.path and response_json.get("success", False)):
                log_info = {
                    "id": str(uuid.uuid4()),
                    "fileName": file_name,
                    "path": request.path,
                    "ip_address": request.remote_addr,
                    "result": "NG" if response.status_code > 200 else "OK",
                    "error_text": error_text,
                    "created_at": datetime.now(UTC).isoformat(),
                }
                log_controller.upsert_item(log_info)

    finally:
        return response


#10铭柄新追加

# PDF 容器路径
PDF_DIR = "https://nriazureaistudio.blob.core.windows.net/1225-container"

def copy_row_style(ws, source_row_idx, target_row_idx):
    """
    将 source_row_idx 的样式复制到 target_row_idx 行（包括字体、边框、填充、对齐方式、数字格式等）
    """
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

def write_wrapped_stock_cell(ws, row, col, stock_value):
    """
    写入 stock 到 Excel 单元格，自动在英日分界处换行并设置 wrap_text。
    """
    if not stock_value:
        return

    # ✅ 在英文(ASCII)和日文之间插入换行
    stock_value = re.sub(r'([a-zA-Z0-9]+)([^\x00-\x7F])', r'\1\n\2', stock_value)

    cell = ws.cell(row=row, column=col, value=stock_value)
    cell.alignment = Alignment(wrap_text=True)
                
def extract_pdf_table(pdf_path):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入上位10銘柄の解説" in text or "組入上位銘柄の解説" in text:
                tables += page.extract_tables()
    return tables

def extract_pdf_table_special(pdf_path):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入銘柄解説" in text:
                tables += page.extract_tables()
    return tables
#清除样式
# def clean_text(text):
#     if not text:
#         return ""
#     return re.sub(r'\s+', '', text.replace('\n', '').strip())
def clean_text(text):
    if not text:
        return ""
    # 全角转半角，并去掉换行、空白符（含全角空格）
    text = jaconv.z2h(text, kana=False, digit=True, ascii=True)
    return re.sub(r'[\s\u3000]+', '', text.strip())
#获取决算月
def get_prev_month_str():
    today = datetime.today()
    prev_month_date = (today.replace(day=1) - timedelta(days=1))
    return prev_month_date.strftime("%Y%m")

#往10铭柄的履历表里写
def insert_tenbrend_history(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)
    

    for record in diff_rows:
        history_item = {
            "filename": record["filename"],
            "id": str(uuid.uuid4()),
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)
        
def insert_tenbrend_history42(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元1": record["元1"],
            "新1": record["新1"],
            "元2": record["元2"],
            "新2": record["新2"],
            "元3": record["元3"],
            "新3": record["新3"],
            "元4": record["元4"],
            "新4": record["新4"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)
        
def insert_tenbrend_history41(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "fcode": record["fcode"],
            "filename": record["filename"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)
        
def insert_tenbrend_history5(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元解決すべき社会的課題": record["元解決すべき社会的課題"],
            "新解決すべき社会的課題": record["新解決すべき社会的課題"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)

# 🚩读取源文件并更新 diff_rows，往10铭柄的excel中写入
def update_excel_with_diff_rows(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        # no = row["no"]
        try:
            no = int(row["no"])
        except (KeyError, TypeError, ValueError):
            no = 0 
        months = row["months"]


        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        headers = {cell.value: idx for idx, cell in enumerate(ws[3])}
        stock_col = headers.get("組入銘柄") or headers.get("銘柄")
        desc_col = headers.get("組入銘柄解説") or headers.get("銘柄解説")
        no_col = headers.get("No.")
        months_col = headers.get("決算月")

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        fcode_col = headers.get("Fコード")
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)

def find_column_by_keyword(header_row, keywords):
    """
    在 header_row 中查找包含关键字的列索引
    """
    for idx, cell in enumerate(header_row):
        title = str(cell.value).strip() if cell.value else ""
        for key in keywords:
            if key in title:
                return idx
    return None

def update_excel_with_diff_rows4(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_esg = row["新最高益更新回数"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["最高益更新回数"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client =get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)
    
def update_excel_with_diff_rows42(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_1 = row["新1"]
        new_2 = row["新2"]
        new_3 = row["新3"]
        new_4 = row["新4"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[2]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)
            ws.cell(row=insert_idx, column=6, value=new_1)
            ws.cell(row=insert_idx, column=7, value=new_2)
            ws.cell(row=insert_idx, column=8, value=new_3)
            ws.cell(row=insert_idx, column=9, value=new_4)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)
            ws.cell(row=target_row_idx, column=6, value=new_1)
            ws.cell(row=target_row_idx, column=7, value=new_2)
            ws.cell(row=target_row_idx, column=8, value=new_3)
            ws.cell(row=target_row_idx, column=9, value=new_4)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)
    
def update_excel_with_diff_rows41(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説","組入発行体解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由","脱炭素社会の実現への貢献と企業評価のポイント"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)
    
def update_excel_with_diff_rows_shang(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_esg = row["新上場年月"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["上場年月"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)

def update_excel_with_diff_rows5(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:

        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_keti = row["新解決すべき社会的課題"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        keti_col = find_column_by_keyword(header_row, ["解決すべき社会的課題", "業種","投資分野","分野","目指すインパクト"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由","社会的課題の解決と利益成長を両立させるポイント"])
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)
    


#10铭柄的check
def check_tenbrend(filename, fund_type):
    try:
        fcode = os.path.basename(filename).split("_")[0]

        if fund_type == 'private':
            TENBREND_CONTAINER_NAME = 'tenbrend_private'
        else:
            TENBREND_CONTAINER_NAME = 'tenbrend'

        container = get_db_connection(TENBREND_CONTAINER_NAME)

        query = "SELECT c.sheetname FROM c WHERE CONTAINS(c.fcode, @fcode)"
        parameters = [{"name": "@fcode", "value": fcode}]
        result = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))

        if not result:
            return []

        sheetname = result[0]["sheetname"]
        pdf_url = f"{PDF_DIR}/{filename}"

        if sheetname == "過去分整理3列":
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return []
            if fcode in ["140193","140386","140675","140565-6","140655-6","140695-6","180291-2","180295-8"]:
                tables = extract_pdf_table_special(io.BytesIO(pdf_response.content))
            else:
                    
                tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return []
            
            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return []

            wb = load_workbook(filename=io.BytesIO(excel_response.content))
            ws = wb.active

            seen_stocks = set()
            unique_rows = []
            for table in tables:
                for row in table:
                    if len(row) < 3:
                        continue
                    if (row[1] and ('組入銘柄' in row[1] or '銘柄' in row[1])) and \
                            ((row[2] and '銘柄解説' in row[2]) or (len(row) > 3 and row[3] and '銘柄解説' in row[3])):
                        continue
                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    if not row[1]:
                        pdf_stock = clean_text(row[2])
                    else:
                        
                        pdf_stock = clean_text(row[1])
                    if not row[2]:
                        pdf_desc = clean_text(row[3])
                    else:
                        pdf_desc = clean_text(row[2])

                    if pdf_stock and not pdf_desc:
                        alt_desc = clean_text(row[3]) if len(row) > 3 else ""
                        if alt_desc:
                            pdf_desc = alt_desc
                        else:
                            continue

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break
            # ✅ Excel 最后一行写入
            for row in unique_rows:
                ws.append(row)

            output_stream = io.BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)
            container_client = get_storage_container()
            blob_client = container_client.get_blob_client("10mingbing.xlsx")
            blob_client.upload_blob(output_stream, overwrite=True)
            # ✅ 与 Cosmos DB 比对并插入必要记录
            diff_rows = []
            for stock, desc in unique_rows:
                # 查询当前项是否存在
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                # return matched[0]["組入銘柄解説"]
                
                if matched:
                    old_desc = clean_text(matched[0]["組入銘柄解説"])
                    
                    if old_desc != desc:
                        # ✅ 差异更新
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "no": 0,
                            "months": "",
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "分類": "銘柄解説更新あり"
                        })
                else:
                    # ✅ 新規插入
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),  # 减1月
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "コメント": "",
                        "分類": "新規銘柄"
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "no": max_no + 1,
                        "months" : get_prev_month_str(),
                        "新組入銘柄解説": desc,
                        "分類": "新規銘柄"
                    })
            insert_tenbrend_history(diff_rows)
            update_excel_with_diff_rows(diff_rows, fund_type)

            return diff_rows or []

        elif sheetname == "過去分整理4列ESG一緒":
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return []

            tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return []
        

            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return []

            wb = load_workbook(filename=io.BytesIO(excel_response.content))
            ws = wb.active

            seen_stocks = set()
            unique_rows = []


            for table in tables:
                header_found = False
                for row in table:
                    if len(row) < 4:
                        continue
                    
                    if (row[1] == "組入銘柄" and
                        "最高益更新回数" in row[2] and
                        "組入銘柄解説" in row[3]):
                        header_found = True
                        continue
                    if not header_found:
                        continue
                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    pdf_stock = clean_text(row[1])
                    pdf_esg = clean_text(row[2])
                    pdf_desc = clean_text(row[3])

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc, pdf_esg])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break

            # ✅ Excel 最后一行写入（调试用）
            for row in unique_rows:
                ws.append(row)

            output_stream = io.BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)
            container_client = get_storage_container()
            blob_client = container_client.get_blob_client("10mingbing.xlsx")
            blob_client.upload_blob(output_stream, overwrite=True)

            # ✅ 比对逻辑
            diff_rows = []
            for stock, desc, esg in unique_rows:
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                if matched:
                    old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                    old_esg = clean_text(matched[0].get("最高益更新回数", ""))

                    classify = None
                    if old_desc != desc or old_esg != esg:
                        classify = "銘柄解説更新あり"

                    if classify:
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        matched_item["最高益更新回数"] = esg
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "元最高益更新回数": old_esg,
                            "新最高益更新回数": esg,
                            "分類": classify,
                            "no": matched_item.get("no", 0),
                            "months": matched_item.get("months", ""),
                            "分類": "銘柄解説更新あり"
                        })

                else:
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "最高益更新回数": esg,
                        "コメント": ""
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "新組入銘柄解説": desc,
                        "最高益更新回数": "",
                        "新最高益更新回数": esg,
                        "分類": "新規銘柄",
                        "no": max_no + 1,
                        "months": get_prev_month_str()
                    })

            insert_tenbrend_history(diff_rows)
            update_excel_with_diff_rows4(diff_rows, fund_type)

            return diff_rows or []

        elif sheetname == "過去分整理4列+4列〇二行":
            return handle_sheet_plus42(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理4列＆（4+1）二行":
            return handle_sheet_plus41(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理4列上場年月":
            return handle_sheet_plus4(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理5列＆（5+1）":
            return handle_sheet_plus5(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname in ["300355","300469","300481"]:
            return handle_sheet_plus_si4(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname in ["300449","300462","300387"]:
            return handle_sheet_plus_si5(pdf_url, fcode, sheetname, fund_type, container, filename)
            
        else:
            return []

    except Exception as e:
        return f"❌ check_tenbrend error: {str(e)}"


def handle_sheet_plus42(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return []

        tables = extract_pdf_table(io.BytesIO(pdf_response.content))
        if not tables:
            return []

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return []

        wb = load_workbook(filename=io.BytesIO(excel_response.content))
        ws = wb.active

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = [row for table in tables for row in table]

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = clean_text(row1[1])
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue

            v1 = clean_text(row1[3])
            v2 = clean_text(row1[4])
            v3 = clean_text(row1[5])
            v4 = clean_text(row1[6])
            desc = clean_text(row1[7]) if len(row1) > 7 else ""
            esg = clean_text(row2[7]) if len(row2) > 7 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, v1, v2, v3, v4, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, v1, v2, v3, v4, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_1 = clean_text(matched[0].get("1", ""))
                old_2 = clean_text(matched[0].get("2", ""))
                old_3 = clean_text(matched[0].get("3", ""))
                old_4 = clean_text(matched[0].get("4", ""))
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg or old_1 != v1 or old_2 != v2 or old_3 != v3 or old_4 != v4:
                    matched_item = matched[0]
                    matched_item.update({
                        "1": v1,
                        "2": v2,
                        "3": v3,
                        "4": v4,
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新1": v1,
                        "元1": old_1,
                        "新2": v2,
                        "元2": old_2,
                        "新3": v3,
                        "元3": old_3,
                        "新4": v4,
                        "元4": old_4,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "1": v1,
                    "2": v2,
                    "3": v3,
                    "4": v4,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                    "コメント": "",
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新1": v1,
                    "元1": "",
                    "新2": v2,
                    "元2": "",
                    "新3": v3,
                    "元3": "",
                    "新4": v4,
                    "元4": "",
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history42(diff_rows)
        update_excel_with_diff_rows42(diff_rows, fund_type)

        return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_plus42 error: {str(e)}"
    
def extract_excel_table(file_like, sheet_name="銘柄解説"):
    try:
        # 支持传入 BytesIO 或本地路径
        df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:C", dtype=str)
    except Exception as e:
        print(f"❌ Excel 读取失败: {e}")
        return []

    df = df.reset_index(drop=True)
    results = []

    for i in range(len(df) - 1):
        index_val = str(df.iloc[i, 0]).strip()

        if index_val in [str(n) for n in range(1, 11)]:  # 只处理1~10
            stock = clean_text(df.iloc[i, 1])
            desc = clean_text(df.iloc[i, 2])
            esg = clean_text(df.iloc[i + 1, 2])
            if stock:
                results.append([stock, desc, esg])

    return results


def handle_sheet_plus41(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        if fcode in ['140752','140302-3']:
            # 将 .pdf 替换为 .xlsx 作为 Excel 文件路径
            excel_url = pdf_url.replace(".pdf", ".xlsx")

            # 下载 Excel 文件内容
            response = requests.get(excel_url)
            if response.status_code != 200:
                return []

            # 转为 BytesIO 对象传给 extract_excel_table
            excel_file = io.BytesIO(response.content)
            tables = extract_excel_table(excel_file, sheetname)
        else:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return []
            
            tables = extract_structured_tables(io.BytesIO(pdf_response.content))
        if not tables:
            return []

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return []

        wb = load_workbook(filename=io.BytesIO(excel_response.content))
        ws = wb.active

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = clean_text(row1[1])
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue
            if fcode in ["140793-6","140764-5"]:
                desc = clean_text(row1[3]) if len(row1) > 2 else ""
                esg = clean_text(row2[3]) if len(row2) > 2 else "" 
            else:
                
                desc = clean_text(row1[2]) if len(row1) > 2 else ""
                esg = clean_text(row2[2]) if len(row2) > 2 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg :
                    matched_item = matched[0]
                    matched_item.update({
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                    "コメント": "",
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history41(diff_rows)
        update_excel_with_diff_rows41(diff_rows, fund_type)

        return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_plus41 error: {str(e)}"
    
#往10铭柄的履历表里写
def insert_tenbrend_history4(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元上場年月": record["元上場年月"],
            "新上場年月": record["新上場年月"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)

    
def handle_sheet_plus4(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return []

            tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return []
        
            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return []

            wb = load_workbook(filename=io.BytesIO(excel_response.content))
            ws = wb.active

            seen_stocks = set()
            unique_rows = []


            for table in tables:
                header_found = False
                for row in table:
                    if len(row) < 4:
                        continue

                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    pdf_stock = clean_text(row[1])
                    
                    pdf_desc = clean_text(row[2])
                    pdf_esg = re.sub(r"(\d{4})年(\d{1,2})月", lambda m: f"{m.group(1)}/{int(m.group(2))}/1", clean_text(row[3]))

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc, pdf_esg])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break

            # ✅ Excel 最后一行写入（调试用）
            for row in unique_rows:
                ws.append(row)

            output_stream = io.BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)
            container_client = get_storage_container()
            blob_client = container_client.get_blob_client("10mingbing.xlsx")
            blob_client.upload_blob(output_stream, overwrite=True)
            
            # ✅ 比对逻辑
            diff_rows = []
            for stock, desc, esg in unique_rows:
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                if matched:
                    old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                    old_esg = clean_text(matched[0].get("上場年月", ""))

                    classify = None
                    if old_desc != desc or old_esg != esg:
                        classify = "銘柄解説更新あり"

                    if classify:
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        matched_item["上場年月"] = esg
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "元上場年月": old_esg,
                            "新上場年月": esg,
                            "分類": classify,
                            "no": matched_item.get("no", 0),
                            "months": matched_item.get("months", ""),
                            "分類": "銘柄解説更新あり"
                        })

                else:
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "上場年月": esg,
                        "コメント": ""
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "新組入銘柄解説": desc,
                        "元上場年月": "",
                        "新上場年月": esg,
                        "分類": "新規銘柄",
                        "no": max_no + 1,
                        "months": get_prev_month_str()
                    })

            insert_tenbrend_history4(diff_rows)
            update_excel_with_diff_rows_shang(diff_rows, fund_type)

            return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_4plus41 error: {str(e)}"

def handle_sheet_plus5(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return []

        tables = extract_structured_tables(io.BytesIO(pdf_response.content))
        if not tables:
            return []

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return []


        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = clean_text(row1[1])
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue
            keti = clean_text(row1[2]) if len(row1) > 3 else ""
            if fcode in ["140793-6","140406-7","180340-1"]:
                desc = clean_text(row1[3]) if len(row1) > 3 else ""
                esg = clean_text(row2[3]) if len(row2) > 3 else ""
            else:
                desc = clean_text(row1[4]) if len(row1) > 3 else ""
                esg = clean_text(row2[4]) if len(row2) > 3 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, keti, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, keti,desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_keti = clean_text(matched[0].get("解決すべき社会的課題", ""))
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg :
                    matched_item = matched[0]
                    matched_item.update({
                        "解決すべき社会的課題": keti,
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新解決すべき社会的課題": keti,
                        "元解決すべき社会的課題": old_keti,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "解決すべき社会的課題": keti,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新解決すべき社会的課題": keti,
                    "元解決すべき社会的課題": "",
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history5(diff_rows)
        update_excel_with_diff_rows5(diff_rows, fund_type)

        return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_plus5 error: {str(e)}"

#私募相关的处理
def insert_tenbrend_history_si4(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def update_excel_with_diff_si4(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        no_col = find_column_by_keyword(header_row, ["No","NO"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)

def handle_sheet_plus_si4(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return []

            tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return []
        
            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return []

            seen_stocks = set()
            unique_rows = []

            
            for table in tables:
                for row in table:
                    if len(row) < 3:
                        continue

                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    pdf_stock = clean_text(row[1])
                    if not row[2]:
                        pdf_desc = clean_text(row[3]) if len(row) > 3 else ""
                    else:
                        pdf_desc = clean_text(row[2])

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break
            
            # ✅ 比对逻辑
            diff_rows = []
            for stock, desc in unique_rows:
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                if matched:
                    old_desc = clean_text(matched[0].get("組入銘柄解説", ""))

                    classify = None
                    if old_desc != desc :
                        classify = "銘柄解説更新あり"

                    if classify:
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "分類": classify,
                            "no": matched_item.get("no", 0),
                            "months": matched_item.get("months", ""),
                            "分類": "銘柄解説更新あり"
                        })

                else:
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "コメント": ""
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "新組入銘柄解説": desc,
                        "分類": "新規銘柄",
                        "no": max_no + 1,
                        "months": get_prev_month_str()
                    })

            insert_tenbrend_history_si4(diff_rows)
            update_excel_with_diff_si4(diff_rows, fund_type)

            return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_4plus41 error: {str(e)}"

def insert_tenbrend_history_si5(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "新社会的課題": record["新社会的課題"],
            "元社会的課題": record["元社会的課題"],
            "新コメント": record["新コメント"],
            "元コメント": record["元コメント"],
            "新ESGコメント": record["新ESGコメント"],
            "元ESGコメント": record["元ESGコメント"],
            "分類": record["分類"],
            "no":record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)

def update_excel_with_diff_si5(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))
    
    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_keti = row["新社会的課題"]
        new_desc = row["新コメント"]
        new_esg = row["新ESGコメント"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]
        
        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄","銘柄名"])
        keti_col = find_column_by_keyword(header_row, ["社会的課題", "目指すインパクト"])
        desc_col = find_column_by_keyword(header_row, ["コメント"])
        esg_col = find_column_by_keyword(header_row, ["ESGコメント"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)
    
def clean_text_si(text):
    return text.replace("\n", "").replace(" ", " ").strip()

def split_by_numbered_blocks(text_block):
    parts = re.split(r'\n?(?=\s*([1-9]|10)[^\d])', text_block)
    combined_parts = []
    i = 1
    while i < len(parts):
        num = parts[i].strip()
        content = parts[i+1].strip() if i+1 < len(parts) else ""
        combined_parts.append(f"{num} {content}")
        i += 2

    results = []
    for part in combined_parts:
        match = re.match(r"^([1-9]|10)[\s ]*([^\s\d]{2,30})[\s ]*([\s\S]+)", part)
        if match:
            no = match.group(1)
            company = clean_text_si(match.group(2))
            description = clean_text_si(match.group(3))
            # 如果三项都不为空，再加入
            if no and company and description:
                results.append([no, company, description])
    return results

def extract_structured_tables(pdf_input):
    all_rows = []
    with pdfplumber.open(pdf_input) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入上位10銘柄の解説" not in text:
                continue

            tables = page.extract_tables()
            if not tables:
                continue

            for table in tables:
                if len(table) == 1 and len(table[0]) == 1:
                    # 说明是合并文本块（单元格中是段落文字）
                    text_block = table[0][0]
                    rows = split_by_numbered_blocks(text_block)
                    all_rows.extend(rows)
                else:
                    # 普通表格结构
                    for row in table:
                        cleaned_row = [clean_text_si(cell) if cell else "" for cell in row]
                        # ✅ 过滤掉字段数量少于3的行
                        if len(cleaned_row) >= 3:
                            all_rows.append(cleaned_row)
    return all_rows
    
def handle_sheet_plus_si5(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return []

        tables = extract_structured_tables(io.BytesIO(pdf_response.content))
        if not tables:
            return []
        
        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return []


        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = clean_text(row1[1])
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue
            keti = clean_text(row1[2]) if len(row1) > 3 else ""
            desc = clean_text(row1[3]) if len(row1) > 3 else ""
            esg = clean_text(row2[3]) if len(row2) > 3 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, keti, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, keti,desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_keti = clean_text(matched[0].get("社会的課題", ""))
                old_desc = clean_text(matched[0].get("コメント", ""))
                old_esg = clean_text(matched[0].get("ESGコメント", ""))
                if old_desc != desc or old_esg != esg :
                    matched_item = matched[0]
                    matched_item.update({
                        "社会的課題": keti,
                        "コメント": desc,
                        "ESGコメント": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新社会的課題": keti,
                        "元社会的課題": old_keti,
                        "新コメント": desc,
                        "元コメント": old_desc,
                        "新ESGコメント": esg,
                        "元ESGコメント": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "社会的課題": keti,
                    "コメント": desc,
                    "ESGコメント": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新社会的課題": keti,
                    "元社会的課題": "",
                    "新コメント": desc,
                    "元コメント": "",
                    "新ESGコメント": esg,
                    "元ESGコメント": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history_si5(diff_rows)
        update_excel_with_diff_si5(diff_rows, fund_type)

        return diff_rows or []

    except Exception as e:
        return f"❌ handle_sheet_plussi5 error: {str(e)}"



app = WsgiToAsgi(app)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True) # 启用HTTPS, ssl_context='adhoc'
